"""
Clarisys Store Firewall Policy API
Evaluates proposed requests against security standards only.
Accepts any source/destination IP or FQDN and returns a standards-based verdict.
"""
import csv
import io
import json
import xml.etree.ElementTree as ET
import os
import hashlib
import html
import hmac
import http.client
import statistics
import threading
import re
import subprocess
import tempfile
import time
from contextlib import asynccontextmanager
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import deque
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Literal
from urllib.parse import urlparse

from fastapi import Body, Depends, FastAPI, File, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, StreamingResponse
from pydantic import BaseModel, EmailStr, Field, field_validator, model_validator

from api.auth import CallerIdentity, current_settings as auth_settings, require_scope
from api import decision_history as _decision_history
from api.audit_store import get_audit_store, make_event
from api.decision_history import (
    append_decision_history,
    force_prune as decision_history_force_prune,
    get_decision_lifecycle,
    list_recent_decisions,
    prune_stats as decision_history_prune_stats,
    set_decision_lifecycle,
)
from api.atomic_io import atomic_write_json
from api.logging_setup import (
    RequestTimer,
    bind_request_context,
    clear_request_context,
    configure_logging,
    get_logger,
    new_request_id,
)
from api.rate_limit import get_limiter, settings_for_logging as rate_limit_settings_for_logging
from api.roi_metrics import (
    record_rule_processed,
    record_opa_request,
    get_prometheus_metrics,
    get_current_metrics,
    update_opa_cache_metrics,
    request_latency_histogram,
    stream_duration_histogram,
)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent.parent / "policy"
DATA_FILE   = BASE_DIR / "data.json"
POLICY_FILE = BASE_DIR / "request_standards.rego"
COMPLIANCE_MAPPINGS_FILE = BASE_DIR / "compliance_mappings.json"
OPA_BINARY  = os.environ.get("OPA_BINARY", "/usr/local/bin/opa")

# ── Runtime config (env-driven, safe defaults) ────────────────────────────────
APP_ENV               = os.environ.get("APP_ENV", "development").lower()
IS_PRODUCTION         = APP_ENV == "production"
IS_TESTING            = os.environ.get("TESTING", "false").lower() == "true"
# Cap request body size — covers /audit/csv and bulk JSON payloads.
# Default 5 MiB: 500-rule cap × ~2 KB/row leaves headroom; override via env.
MAX_REQUEST_BODY_BYTES = int(os.environ.get("MAX_REQUEST_BODY_BYTES", str(5 * 1024 * 1024)))
_STATE_DIR = Path(os.environ.get("FIREWALL_API_STATE_DIR", str(Path.home() / ".firewall-api")))
EVIDENCE_DIR = Path(os.environ.get("EVIDENCE_DIR", "/tmp/firewall-evidence"))
EVIDENCE_INDEX_FILE = EVIDENCE_DIR / "index.jsonl"
EVIDENCE_RETENTION_DAYS = int(os.environ.get("EVIDENCE_RETENTION_DAYS", "365"))
SLO_STATE_FILE = Path(os.environ.get("SLO_STATE_FILE", str(_STATE_DIR / "slo-metrics.json")))

SLO_ALERT_ERROR_RATE_THRESHOLD = float(os.environ.get("SLO_ALERT_ERROR_RATE_THRESHOLD", "0.02"))
SLO_ALERT_P95_MS_THRESHOLD = int(os.environ.get("SLO_ALERT_P95_MS_THRESHOLD", "1500"))
SLO_ALERT_OPA_UNAVAILABLE_THRESHOLD = int(os.environ.get("SLO_ALERT_OPA_UNAVAILABLE_THRESHOLD", "1"))
SLO_ALERT_SLACK_FAILURES_THRESHOLD = int(os.environ.get("SLO_ALERT_SLACK_FAILURES_THRESHOLD", "1"))
SLO_ALERT_DIGEST_BACKLOG_THRESHOLD = int(os.environ.get("SLO_ALERT_DIGEST_BACKLOG_THRESHOLD", "100"))

# ── Logging (configured at import time so module-level loggers are usable) ────
configure_logging()
log = get_logger("api")

_IPV4 = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
_EXTERNAL_IP_RANGES = (
    ("10.", False),
    ("192.168.", False),
    ("172.16.", False),
    ("172.17.", False),
    ("172.18.", False),
    ("172.19.", False),
    ("172.20.", False),
    ("172.21.", False),
    ("172.22.", False),
    ("172.23.", False),
    ("172.24.", False),
    ("172.25.", False),
    ("172.26.", False),
    ("172.27.", False),
    ("172.28.", False),
    ("172.29.", False),
    ("172.30.", False),
    ("172.31.", False),
)

CONTROL_STANDARD_MAP = {
    "Enc-Transit": ["ISO 27001", "CIS v8.1", "Clarisys NFR", "PCI-DSS"],
    "CIS_4.8": ["CIS v8.1", "ISO 27001", "Clarisys NFR"],
    "Cloud-08 / CIS_12.2": ["CIS v8.1", "ISO 27001", "Clarisys NFR"],
    "IAM-8 / Cloud-09 / CIS_8.2": ["ISO 27001", "CIS v8.1", "Clarisys NFR"],
    "CIS_13.6": ["CIS v8.1", "ISO 27001"],
    "Data-01": ["Clarisys NFR", "ISO 27001"],
    "Data-10": ["Clarisys NFR", "ISO 27001"],
    "Data-11": ["Clarisys NFR", "ISO 27001"],
}

ALLOWED_STANDARDS = {"ISO 27001", "CIS v8.1", "PCI-DSS"}

_POLICY_SIGNING_KEY = os.environ.get("POLICY_SIGNING_KEY", "").encode("utf-8")
_POLICY_WEBHOOK_URLS = [
    url.strip() for url in os.environ.get("POLICY_WEBHOOK_URLS", "").split(",") if url.strip()
]
_SLACK_WEBHOOK_URLS = [
    url.strip() for url in os.environ.get("SLACK_WEBHOOK_URLS", "").split(",") if url.strip()
]
_SLACK_HIGH_PRIORITY_WEBHOOK_URLS = [
    url.strip() for url in os.environ.get("SLACK_HIGH_PRIORITY_WEBHOOK_URLS", "").split(",") if url.strip()
]
_SLACK_LOW_PRIORITY_WEBHOOK_URLS = [
    url.strip() for url in os.environ.get("SLACK_LOW_PRIORITY_WEBHOOK_URLS", "").split(",") if url.strip()
]
_SLACK_INCLUDE_JSON_DETAILS = os.environ.get("SLACK_INCLUDE_JSON_DETAILS", "false").lower() == "true"
_SLACK_DEDUP_WINDOW_SECONDS = int(os.environ.get("SLACK_DEDUP_WINDOW_SECONDS", "0"))
_SLACK_DIGEST_MODE = os.environ.get("SLACK_DIGEST_MODE", "false").lower() == "true"
_SLACK_DIGEST_WINDOW_SECONDS = int(os.environ.get("SLACK_DIGEST_WINDOW_SECONDS", "3600"))
_SLACK_API_BASE_URL = os.environ.get("SLACK_API_BASE_URL", "").strip().rstrip("/")
_SLACK_MESSAGE_FORMAT = os.environ.get("SLACK_MESSAGE_FORMAT", "verbose").strip().lower()
_SLACK_STATE_FILE = Path(os.environ.get("SLACK_STATE_FILE", str(_STATE_DIR / "slack-state.json")))
_SLACK_SEND_ONLY_DENY = os.environ.get("SLACK_SEND_ONLY_DENY", "false").lower() == "true"
_SLACK_REALTIME_MIN_RISK = os.environ.get("SLACK_REALTIME_MIN_RISK", "LOW").strip().upper()
_SLACK_MAX_ALERTS_PER_MINUTE = int(os.environ.get("SLACK_MAX_ALERTS_PER_MINUTE", "0"))
_SLACK_DIGEST_AUTO_FLUSH = os.environ.get("SLACK_DIGEST_AUTO_FLUSH", "true").lower() == "true"
_SLACK_DIGEST_FLUSH_INTERVAL_SECONDS = int(os.environ.get("SLACK_DIGEST_FLUSH_INTERVAL_SECONDS", "30"))
_SLACK_WEBHOOK_MAX_RETRIES = int(os.environ.get("SLACK_WEBHOOK_MAX_RETRIES", "3"))
_SLACK_WEBHOOK_BACKOFF_BASE_SECONDS = float(os.environ.get("SLACK_WEBHOOK_BACKOFF_BASE_SECONDS", "0.5"))
_SLACK_WEBHOOK_BACKOFF_MULTIPLIER = float(os.environ.get("SLACK_WEBHOOK_BACKOFF_MULTIPLIER", "2.0"))
_SLACK_WEBHOOK_TIMEOUT_SECONDS = float(os.environ.get("SLACK_WEBHOOK_TIMEOUT_SECONDS", "5.0"))
_OPA_CB_FAILURE_THRESHOLD = int(os.environ.get("OPA_CB_FAILURE_THRESHOLD", "3"))
_OPA_CB_COOLDOWN_SECONDS = int(os.environ.get("OPA_CB_COOLDOWN_SECONDS", "30"))

# Window for treating Slack last_error_at as "recent" in /health?verbose=true.
# Kept short so a transient blip resolves itself before ops sees it on a
# triage view, but long enough that a real outage stays visible across a
# typical 5-minute alerting cycle.
_HEALTH_SLACK_RECENT_FAILURE_WINDOW_SECONDS = int(
    os.environ.get("HEALTH_SLACK_RECENT_FAILURE_WINDOW_SECONDS", "900")
)

# Some read-only endpoints are polled by automation/link previews and can
# create dashboard noise. Keep these out of request/SLO trend metrics.
_SLO_EXCLUDED_GET_PATH_PREFIXES = tuple(
    p.strip()
    for p in os.environ.get(
        "SLO_EXCLUDED_GET_PATH_PREFIXES",
        "/decisions/lifecycle/,/metrics,/notifications/slack/metrics",
    ).split(",")
    if p.strip()
)

_SLACK_DEDUP_LOCK = threading.Lock()
_SLACK_DEDUP_CACHE: dict[str, float] = {}
_SLACK_DIGEST_LOCK = threading.Lock()
_SLACK_DIGEST_STATE = {
    "window_start": 0.0,
    "items": [],
}
_SLACK_RATE_LOCK = threading.Lock()
_SLACK_RATE_STATE = {
    "window_start": 0.0,
    "sent": 0,
}
_SLACK_METRICS_LOCK = threading.Lock()
_SLACK_METRICS = {
    "decision_notifications_sent": 0,
    "batch_notifications_sent": 0,
    "notifications_dedup_suppressed": 0,
    "dispatch_successes": 0,
    "dispatch_failures": 0,
    "last_error": None,
    "last_error_at": None,
    "digest_notifications_sent": 0,
    "digest_items_buffered": 0,
    "policy_suppressed": 0,
    "rate_limited": 0,
    "dispatch_latency_count": 0,
    "dispatch_latency_sum_ms": 0.0,
    "dispatch_latency_last_ms": 0.0,
    "dispatch_latency_max_ms": 0.0,
}
_SLACK_LATENCIES_MS = deque(maxlen=5000)

_SLO_LOCK = threading.Lock()
_SLO_COUNTERS = {
    "requests_total": 0,
    "requests_error": 0,
    "requests_rate_limited": 0,
    "decisions_total": 0,
    "decisions_deny": 0,
    "opa_unavailable": 0,
    "failed_standard_ms_nfr_total": 0,
    "failed_standard_iso_27001_total": 0,
    "failed_standard_cis_v81_total": 0,
    "failed_standard_pci_dss_total": 0,
}
_SLO_LATENCIES_MS = deque(maxlen=5000)
# Separate per-endpoint deques for stream total-duration p95 (populated on
# generator exhaustion, not on StreamingResponse creation like _SLO_LATENCIES_MS).
_STREAM_LATENCIES_MS: dict[str, deque] = {
    "/evaluate/bulk/stream": deque(maxlen=500),
    "/intake/evaluate/bulk/stream": deque(maxlen=500),
}
_EVIDENCE_LOCK = threading.Lock()

_STATE_WRITE_METRICS_LOCK = threading.Lock()
_STATE_WRITE_METRICS = {
    "slo": {"success": 0, "failure": 0},
    "slack": {"success": 0, "failure": 0},
}

_CB_LOCK = threading.Lock()
_CB_FAILURES = 0
_CB_OPEN_UNTIL = 0.0


def _is_synthetic_request(request: Request) -> bool:
    synthetic_header = request.headers.get("x-monitoring-synthetic", "")
    return synthetic_header.lower() in {"1", "true", "yes"}


def _exclude_from_request_metrics(request: Request) -> bool:
    request_id = getattr(request.state, "request_id", None) or request.headers.get("x-request-id", "")
    if isinstance(request_id, str) and request_id.lower().startswith("test-"):
        return True

    if _is_synthetic_request(request):
        return True

    if request.method.upper() != "GET":
        return False
    path = request.url.path
    return any(path.startswith(prefix) for prefix in _SLO_EXCLUDED_GET_PATH_PREFIXES)

def _slack_digest_flush_loop() -> None:
    interval = max(5, _SLACK_DIGEST_FLUSH_INTERVAL_SECONDS)
    while True:
        try:
            _emit_pending_digest_if_due()
        except Exception:
            log.exception("slack.digest_flush_loop_failed")
        time.sleep(interval)


def _startup_runtime_checks() -> None:
    _load_slo_state()
    _load_slack_state()
    _audit_path_writable_check()
    _ensure_evidence_store_ready()
    if _SLACK_DIGEST_MODE and _SLACK_DIGEST_AUTO_FLUSH:
        threading.Thread(target=_slack_digest_flush_loop, daemon=True).start()


@asynccontextmanager
async def _app_lifespan(_: FastAPI):
    _startup_runtime_checks()
    yield


# ── App ────────────────────────────────────────────────────────────────────────
app = FastAPI(
    title="Clarisys Store Firewall Policy API",
    description=(
        "Submit any source IP/FQDN, destination IP/FQDN, protocol, and port. "
        "Returns an immediate ACCEPTABLE or DENY verdict evaluated only against "
        "the Clarisys security standards framework rather than the existing firewall ruleset. "
        "Standards: CIS v8.1 IG3, ISO 27001, PCI-DSS 4.1."
    ),
    version="3.0.0",
    lifespan=_app_lifespan,
    # Disable interactive docs in production; re-enable behind auth later.
    docs_url=None if IS_PRODUCTION else "/docs",
    redoc_url=None if IS_PRODUCTION else "/redoc",
    openapi_url=None if IS_PRODUCTION else "/openapi.json",
)


# ── Hardening middleware ──────────────────────────────────────────────────────
_SECURITY_HEADERS = {
    "X-Content-Type-Options": "nosniff",
    "X-Frame-Options": "DENY",
    "Referrer-Policy": "no-referrer",
    "Cache-Control": "no-store",
    "Content-Security-Policy": "default-src 'none'; frame-ancestors 'none'",
    "Permissions-Policy": "geolocation=(), microphone=(), camera=()",
}


@app.middleware("http")
async def _rate_limit_middleware(request: Request, call_next):
    """Enforce rate limits per caller and endpoint."""
    limiter = get_limiter()
    if not limiter.settings.enabled:
        return await call_next(request)

    # Determine caller ID: prefer auth sub, fallback to IP
    caller_id = getattr(request.state, "caller_id", None)
    if not caller_id:
        caller_id = request.client.host if request.client else "unknown"
        request.state.caller_id = caller_id

    path = request.url.path
    allowed, info = limiter.is_allowed(caller_id, path)
    if not allowed:
        if not _exclude_from_request_metrics(request):
            with _SLO_LOCK:
                _SLO_COUNTERS["requests_total"] += 1
                _SLO_COUNTERS["requests_rate_limited"] += 1
            _save_slo_state()
        log.warning(
            "request.rate_limited",
            caller_id=caller_id,
            endpoint=path,
            limit=info.get("limit"),
        )
        return PlainTextResponse(
            "Rate limit exceeded.",
            status_code=429,
            headers={
                "Retry-After": str(info.get("reset_in_secs", 60)),
                "X-RateLimit-Limit": str(info.get("limit", 0)),
                "X-RateLimit-Remaining": str(info.get("remaining", 0)),
            },
        )

    response = await call_next(request)
    # Stamp rate limit headers on successful response
    if allowed:
        response.headers["X-RateLimit-Limit"] = str(info.get("limit", 0))
        response.headers["X-RateLimit-Remaining"] = str(info.get("remaining", 0))
    return response


@app.middleware("http")
async def _enforce_body_size_and_security_headers(request: Request, call_next):
    """Reject oversized requests up-front, attach a request_id, and stamp
    security headers on every reply. Body-size enforcement uses Content-Length
    when present; we deliberately do not buffer the body to count bytes for
    chunked uploads — that path is rejected by the Content-Length check.
    """
    # Per-request correlation id (honour client header if present)
    request_id = request.headers.get("x-request-id") or new_request_id()
    request.state.request_id = request_id
    bind_request_context(
        request_id=request_id,
        method=request.method,
        path=request.url.path,
    )
    timer = RequestTimer()
    request.state.timer = timer

    try:
        content_length = request.headers.get("content-length")
        if content_length is not None:
            try:
                if int(content_length) > MAX_REQUEST_BODY_BYTES:
                    log.warning("request.rejected", reason="body_too_large", limit=MAX_REQUEST_BODY_BYTES)
                    return PlainTextResponse(
                        "Request body too large.",
                        status_code=413,
                        headers={**_SECURITY_HEADERS, "X-Request-Id": request_id},
                    )
            except ValueError:
                log.warning("request.rejected", reason="invalid_content_length")
                return PlainTextResponse(
                    "Invalid Content-Length.",
                    status_code=400,
                    headers={**_SECURITY_HEADERS, "X-Request-Id": request_id},
                )

        try:
            response = await call_next(request)
        except Exception:
            log.exception("request.unhandled")
            raise

        for key, value in _SECURITY_HEADERS.items():
            response.headers.setdefault(key, value)
        response.headers.setdefault("X-Request-Id", request_id)
        if IS_PRODUCTION:
            response.headers.setdefault(
                "Strict-Transport-Security",
                "max-age=63072000; includeSubDomains",
            )
        log.info(
            "request.completed",
            status_code=response.status_code,
            elapsed_ms=timer.elapsed_ms(),
        )
        if not _exclude_from_request_metrics(request):
            _record_slo(response.status_code, timer.elapsed_ms())
            # Per-endpoint latency histogram for Prom-side percentiles.
            # Use the registered route template (bounded label cardinality)
            # rather than request.url.path; falls back to "unmatched" for
            # 404s so the metric stays usable for routing diagnostics.
            route = request.scope.get("route")
            endpoint_label = getattr(route, "path", None) or "unmatched"
            request_latency_histogram.labels(endpoint=endpoint_label).observe(
                timer.elapsed_ms() / 1000.0
            )
        return response
    finally:
        clear_request_context()

INTAKE_EVALUATE_EXAMPLES = {
    "standard_allow": {
        "summary": "Standard allow request",
        "description": "A straightforward production request for an HTTPS integration.",
        "value": {
            "app_id": "ap-A1234",
            "portfolio": "Finance & Payroll",
            "environment": "production",
            "requested_by": "alice@example.com",
            "expires_at": "2027-03-01",
            "project_reference": "CHG0012345",
            "source_name": "payroll-app",
            "destination_name": "hmrc-api",
            "destination_port": 443,
            "protocol": "TCP",
            "action": "ALLOW",
            "business_justification": "Required to submit payroll data to HMRC via their API gateway.",
        },
    },
    "any_protocol": {
        "summary": "ANY protocol request",
        "description": "Demonstrates an ANY protocol intake request evaluated through the standards policy.",
        "value": {
            "app_id": "ap-B9876",
            "portfolio": "Technology",
            "environment": "production",
            "requested_by": "bob@example.com",
            "expires_at": "2026-11-01",
            "project_reference": "PRJ-444",
            "source_name": "legacy-app",
            "destination_name": "internal-db",
            "destination_port": 1433,
            "protocol": "ANY",
            "action": "ALLOW",
            "business_justification": "Legacy app requires broad protocol access to internal database.",
        },
    },
    "deny_request": {
        "summary": "Explicit deny request",
        "description": "Shows that DENY requests are still evaluated through the same standards path.",
        "value": {
            "app_id": "ap-C2468",
            "portfolio": "Retail Platforms",
            "environment": "production",
            "requested_by": "carol@example.com",
            "expires_at": "2026-12-15",
            "project_reference": "CHG0099999",
            "source_name": "store-kiosk",
            "destination_name": "payment-switch",
            "destination_port": 443,
            "protocol": "TCP",
            "action": "DENY",
            "business_justification": "Block this legacy path while retaining a formally reviewed and time-bounded rule record.",
        },
    },
}

INTAKE_BULK_EVALUATE_EXAMPLES = {
    "mixed_batch": {
        "summary": "Mixed intake batch",
        "description": "Two logical requests: one standard HTTPS flow and one ANY-protocol legacy integration.",
        "value": {
            "requests": [
                INTAKE_EVALUATE_EXAMPLES["standard_allow"]["value"],
                INTAKE_EVALUATE_EXAMPLES["any_protocol"]["value"],
            ]
        },
    },
    "deny_and_allow_batch": {
        "summary": "Allow and deny in one batch",
        "description": "Demonstrates that a batch can mix ALLOW and DENY requests and still return itemized verdicts.",
        "value": {
            "requests": [
                INTAKE_EVALUATE_EXAMPLES["standard_allow"]["value"],
                INTAKE_EVALUATE_EXAMPLES["deny_request"]["value"],
            ]
        },
    },
}


# ── Request schema ─────────────────────────────────────────────────────────────
class TrafficRequest(BaseModel):
    request_id: str | None = Field(
        None,
        max_length=128,
        examples=["req-0001"],
        description=(
            "Optional caller-supplied identifier for this request. "
            "Echoed back in the corresponding results[i].request so callers "
            "can correlate verdicts with their original bulk items."
        ),
    )
    source: str = Field(
        ...,
        examples=["10.157.26.5"],
        description="Source IPv4 address or FQDN (e.g. 10.157.26.5 or host.example.com).",
    )
    destination: str = Field(
        ...,
        examples=["10.221.126.33"],
        description="Destination IPv4 address or FQDN.",
    )
    protocol: str = Field(
        ...,
        examples=["tcp"],
        description="Network protocol: tcp, udp, or icmp.",
    )
    port: int = Field(
        0,
        ge=0,
        le=65535,
        examples=[443],
        description="Destination port (0–65535). Use 0 for icmp.",
    )
    log: str = Field(
        "all",
        examples=["all"],
        description="Proposed logging mode for the new request: all, utm, or no_log.",
    )
    action: str = Field(
        "accept",
        examples=["accept"],
        description="Proposed firewall action. Defaults to accept for new permit requests.",
    )
    encryption_required: bool | None = Field(
        None,
        description="Whether encryption is explicitly required for this proposed request.",
    )
    tls_version_minimum: str | None = Field(
        None,
        examples=["1.2"],
        description="Minimum TLS version if encryption is required.",
    )
    source_interface: str = Field(
        "proposed-src",
        description="Logical source interface label used for standards segmentation checks.",
    )
    destination_interface: str = Field(
        "proposed-dst",
        description="Logical destination interface label used for standards segmentation checks.",
    )
    data_classification: str | None = Field(
        None,
        examples=["Confidential"],
        description="Optional data classification for the proposed flow: Public, Internal, Confidential, Highly Confidential.",
    )
    approved_external_sharing: bool = Field(
        False,
        description="Set true if the request involves approved external data sharing.",
    )
    contract_reference: str | None = Field(
        None,
        description="Contract / DPA reference required for approved external sharing.",
    )
    standards: list[str] = Field(
        default_factory=lambda: ["ISO 27001", "CIS v8.1", "PCI-DSS"],
        description=(
            "Standards to evaluate for this request. "
            "Choose from ISO 27001, CIS v8.1, or PCI-DSS."
        ),
    )

    @field_validator("protocol")
    @classmethod
    def validate_protocol(cls, v: str) -> str:
        allowed = {"tcp", "udp", "icmp", "any"}
        v = v.lower()
        if v not in allowed:
            raise ValueError(f"protocol must be one of {sorted(allowed)}")
        return v

    @field_validator("log")
    @classmethod
    def validate_log(cls, v: str) -> str:
        allowed = {"all", "utm", "no_log", "log_all_sessions", "log_violation_traffic"}
        if v not in allowed:
            raise ValueError(f"log must be one of {sorted(allowed)}")
        return v

    @field_validator("action")
    @classmethod
    def validate_action(cls, v: str) -> str:
        allowed = {"accept", "deny"}
        if v not in allowed:
            raise ValueError(f"action must be one of {sorted(allowed)}")
        return v

    @field_validator("source", "destination")
    @classmethod
    def not_empty(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("must not be empty")
        return v.strip()

    @field_validator("request_id")
    @classmethod
    def validate_request_id(cls, v: str | None) -> str | None:
        if v is None:
            return None
        stripped = v.strip()
        if not stripped:
            return None
        if not re.fullmatch(r"[A-Za-z0-9._:\-]+", stripped):
            raise ValueError(
                "request_id may only contain letters, digits, '.', '_', ':' or '-'"
            )
        return stripped

    @field_validator("standards")
    @classmethod
    def validate_standards(cls, v: list[str]) -> list[str]:
        canonical: list[str] = []
        for standard in v:
            if not isinstance(standard, str):
                raise ValueError("standards must be strings")
            normalized = standard.strip()
            if not normalized:
                continue
            lookup = {
                "iso 27001": "ISO 27001",
                "cis v8.1": "CIS v8.1",
                "pci-dss": "PCI-DSS",
            }.get(normalized.lower(), normalized)
            if lookup not in ALLOWED_STANDARDS:
                raise ValueError(f"standards must be chosen from {sorted(ALLOWED_STANDARDS)}")
            canonical.append(lookup)

        return list(dict.fromkeys(canonical)) if canonical else ["ISO 27001", "CIS v8.1", "PCI-DSS"]


# ── Response schema ────────────────────────────────────────────────────────────
class StandardsVerdict(BaseModel):
    decision_id: str | None = None
    verdict: str = Field(..., description='"ACCEPTABLE" or "DENY"')
    allow: bool
    reason: str = Field(..., description="Standards-based reason for the verdict.")
    overall_status: str = Field(..., description="COMPLIANT or NON-COMPLIANT.")
    overall_risk: str = Field(..., description="LOW, MEDIUM, HIGH, or CRITICAL.")
    failed_standards: list[str] = Field(default_factory=list)
    failed_controls: list[str] = Field(default_factory=list)
    framework_clauses: dict[str, list[str]] = Field(default_factory=dict)
    control_clause_mappings: dict[str, dict[str, list[str]]] = Field(default_factory=dict)
    violations_count: int = 0
    violations: list[dict] = Field(default_factory=list)
    request: dict = Field(..., description="Echo of the submitted request.")
    policy_version: str | None = None
    policy_hash: str | None = None
    policy_signature: str | None = None


class HealthResponse(BaseModel):
    status: str
    opa_available: bool
    data_file_loaded: bool


class BulkRequest(BaseModel):
    requests: list[TrafficRequest] = Field(
        ...,
        min_length=1,
        max_length=500,
        description="Array of proposed traffic requests to evaluate (1–500 items).",
    )


class BulkStreamRequest(BaseModel):
    requests: list[TrafficRequest] = Field(
        ...,
        min_length=1,
        max_length=5000,
        description="Array of proposed traffic requests to stream evaluate (1–5000 items).",
    )


class BulkSummary(BaseModel):
    total: int
    acceptable: int
    denied: int
    by_failed_standard: dict[str, int] = Field(default_factory=dict)
    by_failed_control: dict[str, int] = Field(default_factory=dict)
    failed_controls: list[str] = Field(default_factory=list)
    overall_status: str = Field(..., description="COMPLIANT if every request is acceptable, otherwise NON-COMPLIANT.")


class BulkResponse(BaseModel):
    summary: BulkSummary
    results: list[StandardsVerdict]


# ── Intake schema ──────────────────────────────────────────────────────────────
_BASE_RISK_SCORES = {
    "LOW": 25,
    "MEDIUM": 50,
    "HIGH": 75,
    "CRITICAL": 100,
}

class IntakeRequest(BaseModel):
    intake_mode: Literal["logical"] = Field(
        "logical",
        description='Intake mode. Currently only "logical" is supported.',
    )
    request_id: str | None = Field(
        None,
        max_length=128,
        examples=["req-0001"],
        description=(
            "Optional caller-supplied identifier for this intake item. "
            "Echoed back in the corresponding results[i].intake so callers can "
            "correlate verdicts with their original bulk items."
        ),
    )
    app_id: str = Field(
        ...,
        pattern=r"^ap-[A-Za-z0-9]+$",
        examples=["ap-A1234"],
        description="CMDB application identifier (e.g. ap-A1234).",
    )
    portfolio: str = Field(
        ...,
        min_length=2,
        examples=["Finance & Payroll"],
        description="Business portfolio that owns the application.",
    )
    environment: str = Field(
        ...,
        examples=["production"],
        description="Target environment (e.g. production, staging, development).",
    )
    requested_by: EmailStr = Field(
        ...,
        examples=["alice@example.com"],
        description="Email address of the requestor.",
    )
    expires_at: date = Field(
        ...,
        examples=["2027-05-15"],
        description="Date the access should expire. Maximum 12 months from today.",
    )
    project_reference: str = Field(
        ...,
        min_length=2,
        examples=["CHG0012345"],
        description="Change or project reference this request relates to.",
    )
    source_name: str = Field(
        ...,
        min_length=2,
        examples=["payroll-app"],
        description="Logical name of the source application.",
    )
    destination_name: str = Field(
        ...,
        min_length=2,
        examples=["hmrc-api"],
        description="Logical name of the destination service.",
    )
    destination_port: int | None = Field(
        None,
        ge=0,
        le=65535,
        examples=[443],
        description="Destination port. Required for TCP and UDP; omit for ICMP.",
    )
    protocol: str = Field(
        ...,
        examples=["TCP"],
        description="Protocol: TCP, UDP, ICMP, or ANY. ANY is accepted as input and evaluated through the standards policy.",
    )
    action: str = Field(
        "ALLOW",
        examples=["ALLOW"],
        description="Requested action: ALLOW or DENY.",
    )
    business_justification: str = Field(
        ...,
        min_length=20,
        examples=["Required to submit payroll data to HMRC via their API gateway."],
        description="Business reason for the access request (minimum 20 characters).",
    )

    @field_validator("protocol")
    @classmethod
    def validate_protocol(cls, v: str) -> str:
        allowed = {"TCP", "UDP", "ICMP", "ANY"}
        v = v.upper()
        if v not in allowed:
            raise ValueError(f"protocol must be one of {sorted(allowed)}")
        return v

    @field_validator("action")
    @classmethod
    def validate_action(cls, v: str) -> str:
        allowed = {"ALLOW", "DENY"}
        v = v.upper()
        if v not in allowed:
            raise ValueError(f"action must be one of {sorted(allowed)}")
        return v

    @field_validator("expires_at")
    @classmethod
    def validate_expiry(cls, v: date) -> date:
        max_date = date.today() + timedelta(days=365)
        if v > max_date:
            raise ValueError("expires_at must be within 12 months of today")
        if v <= date.today():
            raise ValueError("expires_at must be in the future")
        return v

    @field_validator("request_id")
    @classmethod
    def validate_request_id(cls, v: str | None) -> str | None:
        if v is None:
            return None
        stripped = v.strip()
        if not stripped:
            return None
        if not re.fullmatch(r"[A-Za-z0-9._:\-]+", stripped):
            raise ValueError(
                "request_id may only contain letters, digits, '.', '_', ':' or '-'"
            )
        return stripped

    standards: list[str] = Field(
        default_factory=lambda: ["ISO 27001", "CIS v8.1", "PCI-DSS"],
        description=(
            "Standards to evaluate for this request. "
            "Choose from ISO 27001, CIS v8.1, or PCI-DSS."
        ),
    )

    @field_validator("standards")
    @classmethod
    def validate_standards(cls, v: list[str]) -> list[str]:
        canonical: list[str] = []
        for standard in v:
            if not isinstance(standard, str):
                raise ValueError("standards must be strings")
            normalized = standard.strip()
            if not normalized:
                continue
            lookup = {
                "iso 27001": "ISO 27001",
                "cis v8.1": "CIS v8.1",
                "pci-dss": "PCI-DSS",
            }.get(normalized.lower(), normalized)
            if lookup not in ALLOWED_STANDARDS:
                raise ValueError(f"standards must be chosen from {sorted(ALLOWED_STANDARDS)}")
            canonical.append(lookup)

        return list(dict.fromkeys(canonical)) if canonical else ["ISO 27001", "CIS v8.1", "PCI-DSS"]

    @model_validator(mode="after")
    def port_required_for_tcp_udp(self) -> "IntakeRequest":
        if self.protocol in {"TCP", "UDP"} and self.destination_port is None:
            raise ValueError("destination_port is required for TCP and UDP protocols")
        return self


class IntakeVerdict(BaseModel):
    """Standards verdict enriched with the logical intake metadata."""
    decision_id: str | None = None
    verdict: str
    allow: bool
    reason: str
    overall_status: str
    overall_risk: str
    risk_score: int = Field(
        ...,
        description="Numeric risk score derived directly from the standards overall_risk band.",
    )
    failed_standards: list[str] = Field(default_factory=list)
    failed_controls: list[str] = Field(default_factory=list)
    framework_clauses: dict[str, list[str]] = Field(default_factory=dict)
    control_clause_mappings: dict[str, dict[str, list[str]]] = Field(default_factory=dict)
    violations_count: int = 0
    violations: list[dict] = Field(default_factory=list)
    intake: dict = Field(..., description="Echo of the submitted intake request.")
    policy_version: str | None = None
    policy_hash: str | None = None
    policy_signature: str | None = None


class LifecycleUpdate(BaseModel):
    status: str
    notes: str | None = None
    expires_at: str | None = None


class IntakeBulkRequest(BaseModel):
    requests: list[IntakeRequest] = Field(
        ...,
        min_length=1,
        max_length=500,
        description="Array of logical intake requests to evaluate (1–500 items).",
    )


class IntakeBulkStreamRequest(BaseModel):
    requests: list[IntakeRequest] = Field(
        ...,
        min_length=1,
        max_length=5000,
        description="Array of logical intake requests to stream evaluate (1–5000 items).",
    )


class IntakeBulkSummary(BaseModel):
    total: int
    acceptable: int
    denied: int
    total_risk_score: int
    max_risk_score: int
    by_failed_standard: dict[str, int] = Field(default_factory=dict)
    by_failed_control: dict[str, int] = Field(default_factory=dict)
    failed_controls: list[str] = Field(default_factory=list)
    overall_status: str = Field(..., description="COMPLIANT if every request is acceptable, otherwise NON-COMPLIANT.")


class IntakeBulkResponse(BaseModel):
    summary: IntakeBulkSummary
    results: list[IntakeVerdict]


# ── Helpers ────────────────────────────────────────────────────────────────────
def _is_internal_ip(value: str) -> bool:
    if not _IPV4.match(value):
        return False
    if value.startswith("10."):
        return True
    if value.startswith("192.168."):
        return True
    return any(value.startswith(prefix) for prefix, _ in _EXTERNAL_IP_RANGES)


def _is_external_destination(value: str) -> bool:
    if _IPV4.match(value):
        return not _is_internal_ip(value)
    return True


def _infer_encryption_required(request: TrafficRequest) -> bool:
    if request.encryption_required is not None:
        return request.encryption_required
    return request.protocol == "tcp" and request.port == 443


def _infer_tls_minimum(request: TrafficRequest, encryption_required: bool) -> str | None:
    if request.tls_version_minimum:
        return request.tls_version_minimum
    return "1.2" if encryption_required and request.protocol == "tcp" and request.port == 443 else None


def _build_standards_input(request: TrafficRequest) -> dict:
    encryption_required = _infer_encryption_required(request)
    tls_minimum = _infer_tls_minimum(request, encryption_required)

    return {
        "source": request.source,
        "destination": request.destination,
        "protocol": request.protocol,
        "port": request.port,
        "log": request.log,
        "action": request.action,
        "encryption_required": encryption_required,
        "tls_version_minimum": tls_minimum,
        "source_interface": request.source_interface,
        "destination_interface": request.destination_interface,
        "data_classification": request.data_classification,
        "approved_external_sharing": request.approved_external_sharing,
        "contract_reference": request.contract_reference,
        "standards": request.standards,
        "destination_is_external": _is_external_destination(request.destination),
    }


def _collect_failed_standards(violations: list[dict], selected_standards: list[str] | None = None) -> list[str]:
    standards: set[str] = set()
    for violation in violations:
        if not isinstance(violation, dict):
            continue
        control = violation.get("control")
        for standard in CONTROL_STANDARD_MAP.get(control, []):
            standards.add(standard)
        raw_standard = violation.get("standard")
        if raw_standard and isinstance(raw_standard, str):
            for part in raw_standard.split("/"):
                standards.add(part.strip().replace("_", " "))
    if selected_standards:
        allowed = set(selected_standards)
        standards = standards & allowed
    return sorted(standards)


def _control_matches_standards(control: str | None, selected_standards: list[str] | None) -> bool:
    """Return True if a control maps to at least one of the selected standards."""
    if not selected_standards or not control:
        return True
    allowed = set(selected_standards)
    mapped = set(CONTROL_STANDARD_MAP.get(control, []))
    if mapped:
        return bool(mapped & allowed)
    return True  # unknown controls pass through


def _filter_violations_by_standards(violations: list[dict], selected_standards: list[str] | None) -> list[dict]:
    """Keep only violations whose control maps to at least one selected standard."""
    if not selected_standards:
        return violations
    return [v for v in violations if _control_matches_standards(v.get("control") if isinstance(v, dict) else None, selected_standards)]


def _filter_controls_by_standards(controls: list[str], selected_standards: list[str] | None) -> list[str]:
    """Keep only controls that map to at least one selected standard."""
    if not selected_standards:
        return controls
    return sorted(c for c in controls if _control_matches_standards(c, selected_standards))


def _canonical_standard_key(standard: str) -> str | None:
    lookup = {
        "clarisys nfr": "failed_standard_ms_nfr_total",
        "ms nfr": "failed_standard_ms_nfr_total",
        "iso 27001": "failed_standard_iso_27001_total",
        "cis v8.1": "failed_standard_cis_v81_total",
        "pci-dss": "failed_standard_pci_dss_total",
        "pci dss": "failed_standard_pci_dss_total",
    }
    return lookup.get(standard.strip().lower())


def _build_reason(request: TrafficRequest | None, decision: dict) -> str:
    if decision.get("compliant"):
        if request is not None:
            requested_standards = ", ".join(request.standards)
            return f"Permitted: the proposed request is compliant with {requested_standards} controls."
        return "Permitted: the proposed request is compliant with the applicable security standards controls."

    violations = decision.get("violations", [])
    selected = request.standards if request is not None else None
    violations = _filter_violations_by_standards(violations, selected)
    failed_standards = _collect_failed_standards(violations, selected_standards=selected)
    failed_controls = _filter_controls_by_standards(
        sorted({v.get("control") for v in violations if isinstance(v, dict) and v.get("control")}),
        selected,
    )

    standards_text = ", ".join(failed_standards[:4]) if failed_standards else "security standards"
    mappings_data = _load_compliance_mappings().get("controls", {})
    ctrl_labels = [mappings_data.get(c, {}).get("title", c) for c in failed_controls[:3]]
    controls_text = ", ".join(ctrl_labels) if ctrl_labels else "control violations"
    return f"Requires remediation due to {standards_text} failures: {controls_text}."


def _collect_remediations(violations: list[dict]) -> list[str]:
    remediations: list[str] = []
    for violation in violations:
        if not isinstance(violation, dict):
            continue
        remediation = violation.get("remediation")
        if isinstance(remediation, str):
            cleaned = remediation.strip()
            if cleaned:
                remediations.append(cleaned)
    return list(dict.fromkeys(remediations))


def _risk_scores() -> dict[str, int]:
    file_path = Path(os.environ.get("RISK_MODEL_FILE", str(BASE_DIR / "risk_model.json")))
    fallback = {"LOW": 25, "MEDIUM": 50, "HIGH": 75, "CRITICAL": 100}
    if not file_path.exists():
        return fallback
    try:
        payload = json.loads(file_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return fallback
    if not isinstance(payload, dict):
        return fallback
    bands = payload.get("bands")
    if not isinstance(bands, dict):
        return fallback
    resolved = {}
    for key, value in bands.items():
        if isinstance(key, str) and isinstance(value, int):
            resolved[key.upper()] = value
    return {**fallback, **resolved}


def _policy_metadata() -> dict[str, str | None]:
    digest = hashlib.sha256()
    for path in sorted(BASE_DIR.glob("*.rego")) + sorted(BASE_DIR.glob("*.json")):
        try:
            digest.update(path.name.encode("utf-8"))
            digest.update(path.read_bytes())
        except OSError:
            continue
    policy_hash = digest.hexdigest()
    signature = None
    if _POLICY_SIGNING_KEY:
        signature = hmac.new(_POLICY_SIGNING_KEY, policy_hash.encode("utf-8"), hashlib.sha256).hexdigest()
    return {
        "policy_version": policy_hash[:12],
        "policy_hash": policy_hash,
        "policy_signature": signature,
    }


def _load_compliance_mappings() -> dict:
    if not COMPLIANCE_MAPPINGS_FILE.exists():
        return {"frameworks": [], "controls": {}}
    try:
        payload = json.loads(COMPLIANCE_MAPPINGS_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"frameworks": [], "controls": {}}
    if not isinstance(payload, dict):
        return {"frameworks": [], "controls": {}}
    frameworks = payload.get("frameworks", [])
    controls = payload.get("controls", {})
    return {
        "frameworks": frameworks if isinstance(frameworks, list) else [],
        "controls": controls if isinstance(controls, dict) else {},
    }


def _framework_coverage_summary(framework: str | None = None) -> dict:
    mappings = _load_compliance_mappings()
    frameworks: list[str] = mappings["frameworks"]
    controls: dict = mappings["controls"]

    requested_frameworks = [framework] if framework else frameworks
    summaries: list[dict] = []
    for name in requested_frameworks:
        if framework and name not in frameworks:
            raise HTTPException(status_code=404, detail=f"Unknown framework: {framework}")

        covered_clauses: set[str] = set()
        mapped_controls: list[dict] = []
        partial_controls = 0
        implemented_controls = 0

        for control_id, control in controls.items():
            if not isinstance(control, dict):
                continue
            control_mappings = control.get("mappings", {})
            if not isinstance(control_mappings, dict) or name not in control_mappings:
                continue
            clauses = control_mappings.get(name, [])
            if isinstance(clauses, list):
                covered_clauses.update(str(clause) for clause in clauses)
            status = str(control.get("status", "implemented"))
            if status == "partial":
                partial_controls += 1
            if status == "implemented":
                implemented_controls += 1
            mapped_controls.append(
                {
                    "control_id": control_id,
                    "title": control.get("title"),
                    "status": status,
                    "clauses": clauses,
                }
            )

        summaries.append(
            {
                "framework": name,
                "controls_total": len(controls),
                "controls_mapped": len(mapped_controls),
                "controls_implemented": implemented_controls,
                "controls_partial": partial_controls,
                "controls_unmapped": len(controls) - len(mapped_controls),
                "clauses_covered": sorted(covered_clauses),
                "controls": sorted(mapped_controls, key=lambda item: item["control_id"]),
            }
        )

    return {
        "frameworks": frameworks,
        "results": summaries,
    }


def _resolve_clause_mappings(
    failed_controls: list[str],
    selected_frameworks: list[str],
) -> tuple[dict[str, list[str]], dict[str, dict[str, list[str]]]]:
    if not failed_controls or not selected_frameworks:
        return {}, {}

    mappings = _load_compliance_mappings()
    controls = mappings.get("controls", {})
    if not isinstance(controls, dict):
        return {}, {}

    frameworks = list(dict.fromkeys(selected_frameworks))
    framework_clauses: dict[str, set[str]] = {name: set() for name in frameworks}
    control_clause_mappings: dict[str, dict[str, list[str]]] = {}

    for control_id in failed_controls:
        control = controls.get(control_id)
        if not isinstance(control, dict):
            continue
        control_mappings = control.get("mappings", {})
        if not isinstance(control_mappings, dict):
            continue

        per_control: dict[str, list[str]] = {}
        for framework in frameworks:
            clauses = control_mappings.get(framework)
            if not isinstance(clauses, list) or not clauses:
                continue
            normalized_clauses = sorted({str(clause) for clause in clauses})
            per_control[framework] = normalized_clauses
            framework_clauses[framework].update(normalized_clauses)

        if per_control:
            control_clause_mappings[control_id] = per_control

    return (
        {
            framework: sorted(clauses)
            for framework, clauses in framework_clauses.items()
            if clauses
        },
        control_clause_mappings,
    )


def _emit_event(event_type: str, payload: dict) -> None:
    if not _POLICY_WEBHOOK_URLS:
        return

    body = json.dumps({"event_type": event_type, "payload": payload}, separators=(",", ":")).encode("utf-8")

    def _post(url: str) -> None:
        parsed = urlparse(url)
        if parsed.scheme != "http" and parsed.scheme != "https":
            return
        conn_cls = http.client.HTTPSConnection if parsed.scheme == "https" else http.client.HTTPConnection
        conn = conn_cls(parsed.netloc, timeout=5)
        try:
            path = parsed.path or "/"
            if parsed.query:
                path += f"?{parsed.query}"
            conn.request("POST", path, body=body, headers={"Content-Type": "application/json"})
            conn.getresponse().read()
        except Exception:
            log.exception("webhook.dispatch_failed", event_type=event_type, url=url)
        finally:
            try:
                conn.close()
            except Exception:
                pass

    for webhook_url in _POLICY_WEBHOOK_URLS:
        threading.Thread(target=_post, args=(webhook_url,), daemon=True).start()


def _stable_json(value: object) -> str:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), default=str)


def _slack_record_latency(latency_ms: float | None) -> None:
    if latency_ms is None:
        return
    if latency_ms < 0:
        return
    with _SLACK_METRICS_LOCK:
        _SLACK_METRICS["dispatch_latency_count"] = int(_SLACK_METRICS.get("dispatch_latency_count", 0)) + 1
        _SLACK_METRICS["dispatch_latency_sum_ms"] = float(_SLACK_METRICS.get("dispatch_latency_sum_ms", 0.0)) + float(latency_ms)
        _SLACK_METRICS["dispatch_latency_last_ms"] = float(latency_ms)
        _SLACK_METRICS["dispatch_latency_max_ms"] = max(
            float(_SLACK_METRICS.get("dispatch_latency_max_ms", 0.0)),
            float(latency_ms),
        )
        _SLACK_LATENCIES_MS.append(float(latency_ms))
    _save_slack_state()


def _slack_record_success(*, latency_ms: float | None = None) -> None:
    with _SLACK_METRICS_LOCK:
        _SLACK_METRICS["dispatch_successes"] += 1
    _save_slack_state()
    _slack_record_latency(latency_ms)


def _slack_record_failure(error: Exception, *, latency_ms: float | None = None) -> None:
    with _SLACK_METRICS_LOCK:
        _SLACK_METRICS["dispatch_failures"] += 1
        _SLACK_METRICS["last_error"] = str(error)
        _SLACK_METRICS["last_error_at"] = datetime.now(timezone.utc).isoformat()
    _save_slack_state()
    _slack_record_latency(latency_ms)


def _slack_record_notification_sent(kind: str) -> None:
    with _SLACK_METRICS_LOCK:
        if kind == "decision":
            _SLACK_METRICS["decision_notifications_sent"] += 1
        elif kind == "batch":
            _SLACK_METRICS["batch_notifications_sent"] += 1
        elif kind == "digest":
            _SLACK_METRICS["digest_notifications_sent"] += 1
    _save_slack_state()


def _normalize_top_remediations(remediations: list[str], limit: int = 3) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in remediations:
        if not isinstance(item, str):
            continue
        cleaned = item.strip()
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
        if len(result) >= limit:
            break
    return result


def _dispatch_webhook_with_retry(url: str, body: bytes, timeout: float = 5.0, max_retries: int = 3) -> tuple[bool, str]:
    """
    POST body to webhook URL with exponential backoff retry on transient failures (5xx, timeout, connection errors).
    Permanent failures (4xx, invalid URL) do not retry.
    
    Args:
        url: Webhook URL
        body: JSON body as bytes
        timeout: Connection timeout in seconds
        max_retries: Maximum number of retry attempts
    
    Returns:
        (success, error_message): True if delivered, False if permanent failure or exhausted retries
    """
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False, f"Invalid scheme: {parsed.scheme}"
    
    conn_cls = http.client.HTTPSConnection if parsed.scheme == "https" else http.client.HTTPConnection
    last_error = None
    
    for attempt in range(max_retries + 1):
        conn = None
        try:
            conn = conn_cls(parsed.netloc, timeout=timeout)
            path = parsed.path or "/"
            if parsed.query:
                path += f"?{parsed.query}"
            conn.request("POST", path, body=body, headers={"Content-Type": "application/json"})
            response = conn.getresponse()
            response_body = response.read().decode("utf-8", "replace")
            
            # 2xx: success
            if response.status < 300:
                if response_body.strip().lower() not in {"", "ok"}:
                    log.warning("slack.dispatch_unexpected_body", body=response_body)
                return True, ""
            
            # 4xx: permanent failure, don't retry
            if response.status < 500:
                error_msg = f"HTTP {response.status}: {response_body[:100]}"
                return False, error_msg
            
            # 5xx: transient, retry
            last_error = f"HTTP {response.status}: {response_body[:100]}"
            if attempt < max_retries:
                backoff = _SLACK_WEBHOOK_BACKOFF_BASE_SECONDS * (_SLACK_WEBHOOK_BACKOFF_MULTIPLIER ** attempt)
                log.warning("slack.dispatch_retrying", attempt=attempt+1, max_retries=max_retries, backoff_seconds=backoff, error=last_error)
                time.sleep(backoff)
            continue
        
        except (http.client.HTTPException, OSError, TimeoutError, RuntimeError) as e:
            # Transient network/connection errors, retry with backoff
            last_error = str(e)
            if attempt < max_retries:
                backoff = _SLACK_WEBHOOK_BACKOFF_BASE_SECONDS * (_SLACK_WEBHOOK_BACKOFF_MULTIPLIER ** attempt)
                log.warning("slack.dispatch_retrying", attempt=attempt+1, max_retries=max_retries, backoff_seconds=backoff, error=type(e).__name__)
                time.sleep(backoff)
            continue
        
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
    
    # Exhausted retries
    return False, f"Exhausted retries after {max_retries} attempts: {last_error}"


    raw = f"{source}|{destination}|{protocol.lower()}|{port}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]


def _rule_fingerprint(source: str, destination: str, protocol: str, port: str) -> str:
    raw = f"{source}|{destination}|{protocol.lower()}|{port}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]


def _risk_rank(risk_key: str) -> int:
    ranking = {"LOW": 1, "MEDIUM": 2, "HIGH": 3, "CRITICAL": 4}
    return ranking.get(risk_key, 0)


def _slack_policy_allows(verdict_key: str, risk_key: str, *, realtime: bool = True) -> bool:
    if _SLACK_SEND_ONLY_DENY and verdict_key != "DENY":
        with _SLACK_METRICS_LOCK:
            _SLACK_METRICS["policy_suppressed"] += 1
        _save_slack_state()
        return False

    if realtime and _risk_rank(risk_key) < _risk_rank(_SLACK_REALTIME_MIN_RISK):
        with _SLACK_METRICS_LOCK:
            _SLACK_METRICS["policy_suppressed"] += 1
        _save_slack_state()
        return False

    if realtime and _SLACK_MAX_ALERTS_PER_MINUTE > 0:
        now = time.time()
        with _SLACK_RATE_LOCK:
            window_start = float(_SLACK_RATE_STATE.get("window_start", 0.0) or 0.0)
            if window_start <= 0 or (now - window_start) >= 60:
                _SLACK_RATE_STATE["window_start"] = now
                _SLACK_RATE_STATE["sent"] = 0
            if _SLACK_RATE_STATE["sent"] >= _SLACK_MAX_ALERTS_PER_MINUTE:
                with _SLACK_METRICS_LOCK:
                    _SLACK_METRICS["rate_limited"] += 1
                _save_slack_state()
                return False
            _SLACK_RATE_STATE["sent"] += 1
        _save_slack_state()

    return True


def _save_slack_state() -> None:
    payload = {
        "dedup_cache": _SLACK_DEDUP_CACHE,
        "digest_state": _SLACK_DIGEST_STATE,
        "metrics": _SLACK_METRICS,
        "rate_state": _SLACK_RATE_STATE,
    }
    try:
        _SLACK_STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(_SLACK_STATE_FILE, payload, separators=(",", ":"))
        _record_state_write("slack", success=True)
    except Exception:
        _record_state_write("slack", success=False)
        log.exception("slack.state_save_failed", file=str(_SLACK_STATE_FILE))


def _load_slack_state() -> None:
    if not _SLACK_STATE_FILE.exists():
        return
    try:
        payload = json.loads(_SLACK_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        log.exception("slack.state_load_failed", file=str(_SLACK_STATE_FILE))
        return
    if not isinstance(payload, dict):
        return

    dedup_cache = payload.get("dedup_cache", {})
    digest_state = payload.get("digest_state", {})
    metrics = payload.get("metrics", {})
    rate_state = payload.get("rate_state", {})
    now = time.time()

    if isinstance(dedup_cache, dict):
        with _SLACK_DEDUP_LOCK:
            _SLACK_DEDUP_CACHE.clear()
            for key, expiry in dedup_cache.items():
                if isinstance(key, str):
                    try:
                        expiry_f = float(expiry)
                    except (TypeError, ValueError):
                        continue
                    if expiry_f > now:
                        _SLACK_DEDUP_CACHE[key] = expiry_f

    if isinstance(digest_state, dict):
        with _SLACK_DIGEST_LOCK:
            window_start = digest_state.get("window_start", 0.0)
            items = digest_state.get("items", [])
            _SLACK_DIGEST_STATE["window_start"] = float(window_start) if isinstance(window_start, (int, float)) else 0.0
            _SLACK_DIGEST_STATE["items"] = items if isinstance(items, list) else []

    if isinstance(metrics, dict):
        with _SLACK_METRICS_LOCK:
            for key, value in metrics.items():
                if key not in _SLACK_METRICS:
                    continue
                if isinstance(_SLACK_METRICS[key], int):
                    try:
                        _SLACK_METRICS[key] = int(value)
                    except (TypeError, ValueError):
                        continue
                elif isinstance(_SLACK_METRICS[key], float):
                    try:
                        _SLACK_METRICS[key] = float(value)
                    except (TypeError, ValueError):
                        continue
                else:
                    _SLACK_METRICS[key] = value

    if isinstance(rate_state, dict):
        with _SLACK_RATE_LOCK:
            window_start = rate_state.get("window_start", 0.0)
            sent = rate_state.get("sent", 0)
            _SLACK_RATE_STATE["window_start"] = float(window_start) if isinstance(window_start, (int, float)) else 0.0
            _SLACK_RATE_STATE["sent"] = int(sent) if isinstance(sent, int) else 0


def _save_slo_state() -> None:
    with _SLO_LOCK:
        payload = {
            "counters": dict(_SLO_COUNTERS),
        }
    try:
        SLO_STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(SLO_STATE_FILE, payload, separators=(",", ":"))
        _record_state_write("slo", success=True)
    except Exception:
        _record_state_write("slo", success=False)
        log.exception("slo.state_save_failed", file=str(SLO_STATE_FILE))


def _record_state_write(component: str, *, success: bool) -> None:
    outcome = "success" if success else "failure"
    with _STATE_WRITE_METRICS_LOCK:
        if component not in _STATE_WRITE_METRICS:
            _STATE_WRITE_METRICS[component] = {"success": 0, "failure": 0}
        _STATE_WRITE_METRICS[component][outcome] += 1


def _state_write_metrics_snapshot() -> dict[str, dict[str, int]]:
    with _STATE_WRITE_METRICS_LOCK:
        return {
            component: {
                "success": int(outcomes.get("success", 0)),
                "failure": int(outcomes.get("failure", 0)),
            }
            for component, outcomes in _STATE_WRITE_METRICS.items()
        }


def _load_slo_state() -> None:
    """
    Load SLO state from disk. Counters are restored from snapshot; latencies are not persisted
    (they rebuild naturally from live requests after restart). This keeps the state file small
    and means percentiles reflect current post-restart performance. For historical latency
    metrics, use Prometheus/Grafana which stores time-series data.
    """
    if not SLO_STATE_FILE.exists():
        _bootstrap_slo_state_from_history()
        return
    try:
        payload = json.loads(SLO_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        log.exception("slo.state_load_failed", file=str(SLO_STATE_FILE))
        return
    if not isinstance(payload, dict):
        return

    counters = payload.get("counters", {})

    with _SLO_LOCK:
        if isinstance(counters, dict):
            for key in _SLO_COUNTERS:
                value = counters.get(key)
                try:
                    _SLO_COUNTERS[key] = int(value)
                except (TypeError, ValueError):
                    continue
        _SLO_LATENCIES_MS.clear()


def _bootstrap_slo_state_from_history() -> bool:
    history_path = _decision_history._history_path()
    if not history_path.exists():
        return False

    counters = {
        "requests_total": 0,
        "requests_error": 0,
        "decisions_total": 0,
        "decisions_deny": 0,
        "opa_unavailable": 0,
        "failed_standard_ms_nfr_total": 0,
        "failed_standard_iso_27001_total": 0,
        "failed_standard_cis_v81_total": 0,
        "failed_standard_pci_dss_total": 0,
    }
    latest_ts = 0
    try:
        lines = history_path.read_text(encoding="utf-8").splitlines()
    except Exception:
        return False

    parse_errors = 0
    for idx, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError as e:
            parse_errors += 1
            if parse_errors <= 5:  # Log first 5 errors to avoid spam
                log.warning("slo.history_parse_error", line_number=idx, error=str(e))
            continue
        if not isinstance(row, dict):
            continue

        counters["requests_total"] += 1
        counters["decisions_total"] += 1
        if row.get("decision_verdict") == "DENY":
            counters["decisions_deny"] += 1
        details = row.get("details", {})
        failed_standards = details.get("failed_standards", []) if isinstance(details, dict) else []
        if isinstance(failed_standards, list):
            for standard in failed_standards:
                if standard == "Clarisys NFR":
                    counters["failed_standard_ms_nfr_total"] += 1
                elif standard == "ISO 27001":
                    counters["failed_standard_iso_27001_total"] += 1
                elif standard == "CIS v8.1":
                    counters["failed_standard_cis_v81_total"] += 1
                elif standard == "PCI-DSS":
                    counters["failed_standard_pci_dss_total"] += 1
        ts_text = row.get("ts")
        if isinstance(ts_text, str):
            try:
                ts = int(datetime.fromisoformat(ts_text.replace("Z", "+00:00")).timestamp())
            except ValueError:
                ts = 0
            latest_ts = max(latest_ts, ts)

    if parse_errors > 5:
        log.warning("slo.history_parse_errors_truncated", total_errors=parse_errors, shown=5)
    if counters["decisions_total"] <= 0:
        return False

    with _SLO_LOCK:
        _SLO_COUNTERS.update(counters)
        _SLO_COUNTERS["opa_unavailable"] = 0
    _save_slo_state()
    return True


def _emit_pending_digest_if_due(force: bool = False) -> bool:
    if not _SLACK_DIGEST_MODE or _SLACK_DIGEST_WINDOW_SECONDS <= 0:
        return False

    now = time.time()
    with _SLACK_DIGEST_LOCK:
        window_start = float(_SLACK_DIGEST_STATE.get("window_start", 0.0) or 0.0)
        items = _SLACK_DIGEST_STATE.get("items", [])
        if not isinstance(items, list) or not items:
            return False
        if window_start <= 0:
            return False
        if not force and (now - window_start) < _SLACK_DIGEST_WINDOW_SECONDS:
            return False
        completed = list(items)
        _SLACK_DIGEST_STATE["window_start"] = now
        _SLACK_DIGEST_STATE["items"] = []

    with _SLACK_METRICS_LOCK:
        _SLACK_METRICS["digest_items_buffered"] = 0

    _save_slack_state()
    _emit_slack_digest(
        {
            "window_start": window_start,
            "window_end": now,
            "items": completed,
        },
        _SLACK_LOW_PRIORITY_WEBHOOK_URLS or _SLACK_WEBHOOK_URLS,
    )
    return True


def _audit_path_writable_check() -> None:
    backend = os.environ.get("AUDIT_BACKEND", "local").strip().lower()
    if backend != "local":
        return
    base = Path(os.environ.get("AUDIT_DIR", "/var/log/firewall-audit"))
    try:
        base.mkdir(parents=True, exist_ok=True)
        test_file = base / ".audit-write-check"
        with test_file.open("a", encoding="utf-8"):
            pass
        test_file.unlink(missing_ok=True)
    except Exception:
        log.warning("audit.path_not_writable", audit_dir=str(base))


def _build_slack_decision_links(decision_id: str, endpoint: str) -> dict[str, str]:
    if not _SLACK_API_BASE_URL or not decision_id:
        return {}
    return {
        "decision": f"{_SLACK_API_BASE_URL}/decisions/lifecycle/{decision_id}",
        "history": f"{_SLACK_API_BASE_URL}/decisions/history?limit=50",
        "explain_hint": f"{_SLACK_API_BASE_URL}{endpoint}/explain" if endpoint.startswith("/evaluate") else "",
    }


def _slack_webhooks_for_risk(risk_key: str) -> list[str]:
    if risk_key in {"HIGH", "CRITICAL"} and _SLACK_HIGH_PRIORITY_WEBHOOK_URLS:
        return _SLACK_HIGH_PRIORITY_WEBHOOK_URLS
    if risk_key in {"LOW", "MEDIUM"} and _SLACK_LOW_PRIORITY_WEBHOOK_URLS:
        return _SLACK_LOW_PRIORITY_WEBHOOK_URLS
    return _SLACK_WEBHOOK_URLS


def _slack_digest_enqueue(item: dict) -> dict | None:
    if not _SLACK_DIGEST_MODE or _SLACK_DIGEST_WINDOW_SECONDS <= 0:
        return None

    now = time.time()
    with _SLACK_DIGEST_LOCK:
        window_start = float(_SLACK_DIGEST_STATE.get("window_start", 0.0) or 0.0)
        items = _SLACK_DIGEST_STATE.get("items", [])
        if not isinstance(items, list):
            items = []

        if window_start <= 0:
            _SLACK_DIGEST_STATE["window_start"] = now
            _SLACK_DIGEST_STATE["items"] = [item]
            with _SLACK_METRICS_LOCK:
                _SLACK_METRICS["digest_items_buffered"] = 1
            _save_slack_state()
            return None

        if (now - window_start) < _SLACK_DIGEST_WINDOW_SECONDS:
            items.append(item)
            _SLACK_DIGEST_STATE["items"] = items
            with _SLACK_METRICS_LOCK:
                _SLACK_METRICS["digest_items_buffered"] = len(items)
            _save_slack_state()
            return None

        completed = list(items)
        _SLACK_DIGEST_STATE["window_start"] = now
        _SLACK_DIGEST_STATE["items"] = [item]
        with _SLACK_METRICS_LOCK:
            _SLACK_METRICS["digest_items_buffered"] = 1

    if not completed:
        return None

    _save_slack_state()

    return {
        "window_start": window_start,
        "window_end": now,
        "items": completed,
    }


def _emit_slack_digest(summary: dict, webhook_urls: list[str]) -> None:
    if not webhook_urls:
        return
    items = summary.get("items", []) if isinstance(summary.get("items"), list) else []
    if not items:
        return

    by_risk: dict[str, int] = {}
    lines: list[str] = []
    for item in items[:20]:
        risk = str(item.get("risk", "UNKNOWN"))
        by_risk[risk] = by_risk.get(risk, 0) + 1
        lines.append(
            f"- {item.get('endpoint', 'unknown')}: {item.get('source', 'unknown')} -> "
            f"{item.get('destination', 'unknown')} {item.get('protocol', 'unknown')}/{item.get('port', 'unknown')} "
            f"({item.get('verdict', 'UNKNOWN')}, {risk})"
        )

    text = (
        "Firewall low/medium digest\n"
        f"Window start: {datetime.fromtimestamp(summary['window_start'], timezone.utc).isoformat()}\n"
        f"Window end: {datetime.fromtimestamp(summary['window_end'], timezone.utc).isoformat()}\n"
        f"Items: {len(items)}\n"
        f"By risk: {json.dumps(by_risk, sort_keys=True)}\n"
        + "\n".join(lines)
    )
    body = json.dumps({"text": text}, separators=(",", ":")).encode("utf-8")

    def _post(url: str) -> bool:
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            return False
        conn_cls = http.client.HTTPSConnection if parsed.scheme == "https" else http.client.HTTPConnection
        conn = conn_cls(parsed.netloc, timeout=5)
        started_at = time.perf_counter()
        try:
            path = parsed.path or "/"
            if parsed.query:
                path += f"?{parsed.query}"
            conn.request("POST", path, body=body, headers={"Content-Type": "application/json"})
            response = conn.getresponse()
            response_body = response.read().decode("utf-8", "replace")
            if response.status >= 300:
                raise RuntimeError(f"Slack webhook returned HTTP {response.status}: {response_body}")
            _slack_record_success(latency_ms=(time.perf_counter() - started_at) * 1000.0)
            return True
        except Exception as exc:
            _slack_record_failure(exc, latency_ms=(time.perf_counter() - started_at) * 1000.0)
            log.exception("slack.digest_dispatch_failed", url=url)
            return False
        finally:
            try:
                conn.close()
            except Exception:
                pass

    sent_any = False
    for webhook_url in webhook_urls:
        sent_any = _post(webhook_url) or sent_any
    if sent_any:
        _slack_record_notification_sent("digest")


def _slack_metrics_snapshot() -> dict:
    with _SLACK_METRICS_LOCK:
        snapshot = dict(_SLACK_METRICS)
        latencies = list(_SLACK_LATENCIES_MS)

    if len(latencies) >= 2:
        quantiles = statistics.quantiles(latencies, n=20)
        p50 = quantiles[9]   # 50th percentile
        p95 = quantiles[18]  # 95th percentile
    else:
        p50 = 0.0
        p95 = 0.0

    snapshot.setdefault("dispatch_latency_count", 0)
    snapshot.setdefault("dispatch_latency_sum_ms", 0.0)
    snapshot.setdefault("dispatch_latency_last_ms", 0.0)
    snapshot.setdefault("dispatch_latency_max_ms", 0.0)
    snapshot["dispatch_latency_p50_ms"] = round(float(p50), 3)
    snapshot["dispatch_latency_p95_ms"] = round(float(p95), 3)
    snapshot["dispatch_latency_avg_ms"] = round(
        (float(snapshot.get("dispatch_latency_sum_ms", 0.0)) / int(snapshot.get("dispatch_latency_count", 0)))
        if int(snapshot.get("dispatch_latency_count", 0)) > 0
        else 0.0,
        3,
    )

    with _SLACK_DEDUP_LOCK:
        active_keys = sum(1 for expiry in _SLACK_DEDUP_CACHE.values() if expiry > time.time())
    snapshot["dedup_window_seconds"] = _SLACK_DEDUP_WINDOW_SECONDS
    snapshot["dedup_cache_active_keys"] = active_keys
    snapshot["digest_mode"] = _SLACK_DIGEST_MODE
    snapshot["digest_window_seconds"] = _SLACK_DIGEST_WINDOW_SECONDS
    snapshot["send_only_deny"] = _SLACK_SEND_ONLY_DENY
    snapshot["realtime_min_risk"] = _SLACK_REALTIME_MIN_RISK
    snapshot["max_alerts_per_minute"] = _SLACK_MAX_ALERTS_PER_MINUTE
    snapshot["state_file"] = str(_SLACK_STATE_FILE)
    snapshot["message_format"] = _SLACK_MESSAGE_FORMAT
    return snapshot


def _reset_slack_metrics(clear_dedup_cache: bool = True) -> dict:
    with _SLACK_METRICS_LOCK:
        _SLACK_METRICS["decision_notifications_sent"] = 0
        _SLACK_METRICS["batch_notifications_sent"] = 0
        _SLACK_METRICS["digest_notifications_sent"] = 0
        _SLACK_METRICS["digest_items_buffered"] = 0
        _SLACK_METRICS["notifications_dedup_suppressed"] = 0
        _SLACK_METRICS["dispatch_successes"] = 0
        _SLACK_METRICS["dispatch_failures"] = 0
        _SLACK_METRICS["last_error"] = None
        _SLACK_METRICS["last_error_at"] = None
        _SLACK_METRICS["dispatch_latency_count"] = 0
        _SLACK_METRICS["dispatch_latency_sum_ms"] = 0.0
        _SLACK_METRICS["dispatch_latency_last_ms"] = 0.0
        _SLACK_METRICS["dispatch_latency_max_ms"] = 0.0
        _SLACK_LATENCIES_MS.clear()

    with _SLACK_DIGEST_LOCK:
        _SLACK_DIGEST_STATE["window_start"] = 0.0
        _SLACK_DIGEST_STATE["items"] = []

    with _SLACK_RATE_LOCK:
        _SLACK_RATE_STATE["window_start"] = 0.0
        _SLACK_RATE_STATE["sent"] = 0

    _save_slack_state()

    if clear_dedup_cache:
        with _SLACK_DEDUP_LOCK:
            _SLACK_DEDUP_CACHE.clear()

    return {
        "reset": True,
        "cleared_dedup_cache": clear_dedup_cache,
        **_slack_metrics_snapshot(),
    }


def _slack_should_send(dedup_key: str) -> bool:
    if _SLACK_DEDUP_WINDOW_SECONDS <= 0:
        return True

    now = time.time()
    with _SLACK_DEDUP_LOCK:
        expires_at = _SLACK_DEDUP_CACHE.get(dedup_key, 0.0)
        if expires_at > now:
            with _SLACK_METRICS_LOCK:
                _SLACK_METRICS["notifications_dedup_suppressed"] += 1
            return False

        _SLACK_DEDUP_CACHE[dedup_key] = now + _SLACK_DEDUP_WINDOW_SECONDS

        # Keep cache bounded by removing expired entries opportunistically.
        if len(_SLACK_DEDUP_CACHE) > 10000:
            for key, expiry in list(_SLACK_DEDUP_CACHE.items()):
                if expiry <= now:
                    _SLACK_DEDUP_CACHE.pop(key, None)

    _save_slack_state()

    return True


def _emit_slack_decision(payload: dict) -> None:
    webhook_urls = _slack_webhooks_for_risk(str(payload.get("overall_risk", "UNKNOWN")).strip().upper())
    if not webhook_urls:
        return

    verdict = str(payload.get("decision_verdict", "UNKNOWN"))
    overall_status = str(payload.get("overall_status", "UNKNOWN"))
    overall_risk = str(payload.get("overall_risk", "UNKNOWN"))
    verdict_key = verdict.strip().upper()
    status_key = overall_status.strip().upper()
    risk_key = overall_risk.strip().upper()
    details = payload.get("details", {}) if isinstance(payload.get("details"), dict) else {}
    failure_reason = details.get("reason")
    remediations = details.get("remediations", []) if isinstance(details.get("remediations"), list) else []
    top_remediations = _normalize_top_remediations(remediations, limit=3)
    source_network = details.get("source") or details.get("source_name") or "unknown"
    destination_network = details.get("destination") or details.get("destination_name") or "unknown"
    protocol_value = details.get("protocol") or "unknown"
    port_value = details.get("port")
    if port_value is None:
        port_value = details.get("destination_port")
    port_display = str(port_value) if port_value is not None else "unknown"
    rule_fp = _rule_fingerprint(str(source_network), str(destination_network), str(protocol_value), port_display)

    if verdict_key == "ACCEPTABLE":
        verdict_label = ":white_check_mark: ACCEPTABLE"
    elif verdict_key == "DENY":
        verdict_label = ":x: DENY"
    else:
        verdict_label = verdict.strip() or "UNKNOWN"

    if status_key == "COMPLIANT":
        status_label = ":white_check_mark: COMPLIANT"
    elif status_key == "NON-COMPLIANT":
        status_label = ":x: NON-COMPLIANT"
    else:
        status_label = overall_status.strip() or "UNKNOWN"

    if risk_key == "LOW":
        risk_label = "🟢 LOW"
    elif risk_key == "MEDIUM":
        risk_label = "🟡 MEDIUM"
    elif risk_key == "HIGH":
        risk_label = "🔴 HIGH"
    elif risk_key == "CRITICAL":
        risk_label = "🚨 CRITICAL"
    else:
        risk_label = overall_risk.strip() or "UNKNOWN"

    should_show_failure_details = verdict_key == "DENY" or status_key == "NON-COMPLIANT"
    decision_links = _build_slack_decision_links(str(payload.get("decision_id", "")), str(payload.get("endpoint", "")))

    decision_dedup_key = _stable_json(
        {
            "kind": "decision",
            "endpoint": payload.get("endpoint", "unknown"),
            "decision_verdict": verdict_key,
            "overall_status": status_key,
            "overall_risk": risk_key,
            "source_network": source_network,
            "destination_network": destination_network,
            "protocol": str(protocol_value),
            "port": port_display,
            "rule_fingerprint": rule_fp,
            "reason": str(failure_reason).strip() if isinstance(failure_reason, str) else None,
            "remediations": top_remediations,
        }
    )
    if not _slack_should_send(decision_dedup_key):
        return

    if _SLACK_DIGEST_MODE and risk_key in {"LOW", "MEDIUM"}:
        if not _slack_policy_allows(verdict_key, risk_key, realtime=False):
            return
        digest_summary = _slack_digest_enqueue(
            {
                "endpoint": payload.get("endpoint", "unknown"),
                "source": source_network,
                "destination": destination_network,
                "protocol": str(protocol_value),
                "port": port_display,
                "verdict": verdict_key,
                "risk": risk_key,
            }
        )
        if digest_summary is None:
            return
        _emit_slack_digest(digest_summary, webhook_urls)
        return

    if not _slack_policy_allows(verdict_key, risk_key, realtime=True):
        return

    text = (
        "Firewall policy decision\n"
        f"Decision verdict: {verdict_label}\n"
        f"Overall status: {status_label}\n"
        f"Overall risk: {risk_label}\n"
        f"Source network: {source_network}\n"
        f"Destination network: {destination_network}\n"
        f"Protocol: {protocol_value}\n"
        f"Port: {port_display}\n"
        f"Rule fingerprint: {rule_fp}\n"
    )
    if decision_links.get("decision"):
        text += f"Decision details: {decision_links['decision']}\n"
    if decision_links.get("history"):
        text += f"Recent decisions: {decision_links['history']}\n"
    if decision_links.get("explain_hint"):
        text += f"Explain endpoint: {decision_links['explain_hint']}\n"
    if should_show_failure_details and isinstance(failure_reason, str) and failure_reason.strip():
        text += f"Failure reason: {failure_reason}\n"
    if should_show_failure_details and top_remediations:
        text += "Remediation:\n" + "\n".join(f"- {item}" for item in top_remediations) + "\n"
    header_fields = [
        {"type": "mrkdwn", "text": f"*Decision verdict*\n{verdict_label}"},
        {"type": "mrkdwn", "text": f"*Overall status*\n{status_label}"},
        {"type": "mrkdwn", "text": f"*Overall risk*\n{risk_label}"},
        {"type": "mrkdwn", "text": f"*Endpoint*\n{payload.get('endpoint', 'unknown')}"},
        {"type": "mrkdwn", "text": f"*Source network*\n{source_network}"},
        {"type": "mrkdwn", "text": f"*Destination network*\n{destination_network}"},
        {"type": "mrkdwn", "text": f"*Protocol*\n{protocol_value}"},
        {"type": "mrkdwn", "text": f"*Port*\n{port_display}"},
        {"type": "mrkdwn", "text": f"*Rule fingerprint*\n{rule_fp}"},
    ]
    if _SLACK_MESSAGE_FORMAT == "compact":
        header_fields = [
            {"type": "mrkdwn", "text": f"*Decision verdict*\n{verdict_label}"},
            {"type": "mrkdwn", "text": f"*Overall risk*\n{risk_label}"},
            {"type": "mrkdwn", "text": f"*Rule*\n{source_network} → {destination_network}"},
            {"type": "mrkdwn", "text": f"*Proto/Port*\n{protocol_value}/{port_display}"},
            {"type": "mrkdwn", "text": f"*FP*\n{rule_fp}"},
        ]

    blocks = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": "Firewall policy decision"},
        },
        {
            "type": "section",
            "fields": header_fields,
        },
    ]
    if should_show_failure_details and isinstance(failure_reason, str) and failure_reason.strip():
        blocks.append(
            {
                "type": "section",
                "text": {"type": "mrkdwn", "text": f"*Failure reason*\n{failure_reason}"},
            }
        )
    if should_show_failure_details and remediations:
        blocks.append(
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "*Remediation*\n" + "\n".join(f"• {item}" for item in remediations[:5]),
                },
            }
        )
    blocks.append(
        {
            "type": "context",
            "elements": [
                {
                    "type": "mrkdwn",
                    "text": (
                        f"Decision ID: `{payload.get('decision_id', 'unknown')}`. "
                        "Use the API for full details when needed."
                    ),
                }
            ],
        }
    )

    if decision_links.get("decision") or decision_links.get("history") or decision_links.get("explain_hint"):
        link_lines: list[str] = []
        if decision_links.get("decision"):
            link_lines.append(f"*Decision details*\n{decision_links['decision']}")
        if decision_links.get("history"):
            link_lines.append(f"*Recent decisions*\n{decision_links['history']}")
        if decision_links.get("explain_hint"):
            link_lines.append(f"*Explain endpoint*\n{decision_links['explain_hint']}")
        blocks.append(
            {
                "type": "section",
                "fields": [{"type": "mrkdwn", "text": line} for line in link_lines],
            }
        )

    if _SLACK_INCLUDE_JSON_DETAILS:
        details_json = json.dumps(payload, indent=2, default=str)
        blocks.extend(
            [
                {
                    "type": "context",
                    "elements": [
                        {
                            "type": "mrkdwn",
                            "text": "Full JSON details included because `SLACK_INCLUDE_JSON_DETAILS=true`.",
                        }
                    ],
                },
                {
                    "type": "section",
                    "text": {"type": "mrkdwn", "text": f"```{details_json}```"},
                },
            ]
        )

    message = {"text": text, "blocks": blocks}
    body = json.dumps(message, separators=(",", ":")).encode("utf-8")

    def _post(url: str) -> bool:
        started_at = time.perf_counter()
        success, error_msg = _dispatch_webhook_with_retry(
            url, body, 
            timeout=_SLACK_WEBHOOK_TIMEOUT_SECONDS,
            max_retries=_SLACK_WEBHOOK_MAX_RETRIES
        )
        latency_ms = (time.perf_counter() - started_at) * 1000.0
        if success:
            _slack_record_success(latency_ms=latency_ms)
        else:
            _slack_record_failure(RuntimeError(error_msg), latency_ms=latency_ms)
            log.error("slack.dispatch_failed", url=url, error=error_msg)
        return success

    sent_any = False
    for webhook_url in webhook_urls:
        sent_any = _post(webhook_url) or sent_any
    if sent_any:
        _slack_record_notification_sent("decision")


def _emit_slack_batch_summary(title: str, summary: dict) -> None:
    overall_status = str(summary.get("overall_status", "UNKNOWN"))
    status_key = overall_status.strip().upper()
    webhook_urls = _SLACK_HIGH_PRIORITY_WEBHOOK_URLS if status_key == "NON-COMPLIANT" and _SLACK_HIGH_PRIORITY_WEBHOOK_URLS else (
        _SLACK_LOW_PRIORITY_WEBHOOK_URLS if status_key == "COMPLIANT" and _SLACK_LOW_PRIORITY_WEBHOOK_URLS else _SLACK_WEBHOOK_URLS
    )
    if not webhook_urls:
        return

    if status_key == "COMPLIANT":
        status_label = ":white_check_mark: COMPLIANT"
    elif status_key == "NON-COMPLIANT":
        status_label = ":x: NON-COMPLIANT"
    else:
        status_label = overall_status.strip() or "UNKNOWN"

    try:
        denied_count = int(summary.get("denied", 0) or 0)
    except (TypeError, ValueError):
        denied_count = 0

    if _SLACK_SEND_ONLY_DENY and denied_count <= 0:
        with _SLACK_METRICS_LOCK:
            _SLACK_METRICS["policy_suppressed"] += 1
        return

    if _SLACK_MAX_ALERTS_PER_MINUTE > 0:
        now = time.time()
        with _SLACK_RATE_LOCK:
            window_start = float(_SLACK_RATE_STATE.get("window_start", 0.0) or 0.0)
            if window_start <= 0 or (now - window_start) >= 60:
                _SLACK_RATE_STATE["window_start"] = now
                _SLACK_RATE_STATE["sent"] = 0
            if _SLACK_RATE_STATE["sent"] >= _SLACK_MAX_ALERTS_PER_MINUTE:
                with _SLACK_METRICS_LOCK:
                    _SLACK_METRICS["rate_limited"] += 1
                return
            _SLACK_RATE_STATE["sent"] += 1

    should_show_failure_details = status_key == "NON-COMPLIANT" or denied_count > 0

    batch_dedup_key = _stable_json(
        {
            "kind": "batch_summary",
            "title": title,
            "overall_status": status_key,
            "total": summary.get("total"),
            "acceptable": summary.get("acceptable"),
            "denied": summary.get("denied"),
            "failed_controls": summary.get("failed_controls", []),
            "by_failed_standard": summary.get("by_failed_standard", {}),
            "by_failed_control": summary.get("by_failed_control", {}),
        }
    )
    if not _slack_should_send(batch_dedup_key):
        return

    lines = [
        title,
        f"Overall status: {status_label}",
        f"Total: {summary.get('total')}",
        f"Acceptable: {summary.get('acceptable')}",
        f"Denied: {summary.get('denied')}",
    ]

    if should_show_failure_details:
        failed_controls = summary.get("failed_controls")
        if isinstance(failed_controls, list) and failed_controls:
            lines.append(f"Failed controls: {', '.join(str(item) for item in failed_controls[:10])}")
        by_failed_standard = summary.get("by_failed_standard")
        if isinstance(by_failed_standard, dict) and by_failed_standard:
            lines.append(f"By failed standard: {json.dumps(by_failed_standard, sort_keys=True)}")
        by_failed_control = summary.get("by_failed_control")
        if isinstance(by_failed_control, dict) and by_failed_control:
            lines.append(f"By failed control: {json.dumps(by_failed_control, sort_keys=True)}")

    if _SLACK_INCLUDE_JSON_DETAILS and should_show_failure_details:
        lines.append(f"```{json.dumps(summary, indent=2, default=str)}```")

    text = "\n".join(lines)
    message = {"text": text}
    body = json.dumps(message, separators=(",", ":")).encode("utf-8")

    def _post(url: str) -> bool:
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            return False
        conn_cls = http.client.HTTPSConnection if parsed.scheme == "https" else http.client.HTTPConnection
        conn = conn_cls(parsed.netloc, timeout=5)
        started_at = time.perf_counter()
        try:
            path = parsed.path or "/"
            if parsed.query:
                path += f"?{parsed.query}"
            conn.request("POST", path, body=body, headers={"Content-Type": "application/json"})
            response = conn.getresponse()
            response_body = response.read().decode("utf-8", "replace")
            if response.status >= 300:
                raise RuntimeError(f"Slack webhook returned HTTP {response.status}: {response_body}")
            _slack_record_success(latency_ms=(time.perf_counter() - started_at) * 1000.0)
            return True
        except Exception as exc:
            _slack_record_failure(exc, latency_ms=(time.perf_counter() - started_at) * 1000.0)
            log.exception("slack.batch_summary_failed", title=title, url=url)
            return False
        finally:
            try:
                conn.close()
            except Exception:
                pass

    sent_any = False
    for webhook_url in webhook_urls:
        sent_any = _post(webhook_url) or sent_any
    if sent_any:
        _slack_record_notification_sent("batch")


def _decision_id(request_id: str, endpoint: str) -> str:
    endpoint_slug = endpoint.strip("/").replace("/", ".") or "root"
    return f"{request_id}:{endpoint_slug}"


def _record_slo(status_code: int, elapsed_ms: int) -> None:
    with _SLO_LOCK:
        _SLO_COUNTERS["requests_total"] += 1
        if status_code >= 500:
            _SLO_COUNTERS["requests_error"] += 1
        _SLO_LATENCIES_MS.append(elapsed_ms)
    _save_slo_state()


def _slo_snapshot() -> dict:
    state_write_metrics = _state_write_metrics_snapshot()
    with _SLO_LOCK:
        total = _SLO_COUNTERS["requests_total"]
        errors = _SLO_COUNTERS["requests_error"]
        latencies = list(_SLO_LATENCIES_MS)
        if len(latencies) >= 2:
            quantiles = statistics.quantiles(latencies, n=20)
            p50 = quantiles[9]   # 50th percentile
            p95 = quantiles[18]  # 95th percentile
        else:
            p50 = 0
            p95 = 0
        avg = round((sum(latencies) / len(latencies)), 3) if latencies else 0.0
        deny_rate = (
            _SLO_COUNTERS["decisions_deny"] / _SLO_COUNTERS["decisions_total"]
            if _SLO_COUNTERS["decisions_total"]
            else 0.0
        )
        error_rate = (errors / total) if total else 0.0
        # Stream end-to-end p95 (in-process approximation, separate from SLO_LATENCIES_MS
        # which captures time-to-StreamingResponse-object, not full stream duration)
        def _stream_p95(endpoint: str) -> float:
            lats = list(_STREAM_LATENCIES_MS.get(endpoint, []))
            if len(lats) >= 2:
                return round(statistics.quantiles(lats, n=20)[18], 3)
            return 0.0

        return {
            **_SLO_COUNTERS,
            "error_rate": round(error_rate, 6),
            "deny_rate": round(deny_rate, 6),
            "latency_avg_ms": avg,
            "latency_p50_ms": p50,
            "latency_p95_ms": p95,
            "stream_p95_ms": _stream_p95("/evaluate/bulk/stream"),
            "intake_stream_p95_ms": _stream_p95("/intake/evaluate/bulk/stream"),
            "state_write_metrics": state_write_metrics,
        }


def _severity_rank(value: str) -> int:
    return {"ok": 0, "warn": 1, "critical": 2}.get(value, 0)


def _slo_alerts_snapshot() -> dict:
    slo = _slo_snapshot()
    slack = _slack_metrics_snapshot()
    active_alerts: list[dict] = []

    if slo["error_rate"] >= SLO_ALERT_ERROR_RATE_THRESHOLD:
        active_alerts.append(
            {
                "id": "api.error-rate",
                "severity": "critical" if slo["error_rate"] >= (SLO_ALERT_ERROR_RATE_THRESHOLD * 2) else "warn",
                "value": slo["error_rate"],
                "threshold": SLO_ALERT_ERROR_RATE_THRESHOLD,
                "message": "API error rate exceeded threshold.",
            }
        )

    if slo["latency_p95_ms"] >= SLO_ALERT_P95_MS_THRESHOLD:
        active_alerts.append(
            {
                "id": "api.latency-p95",
                "severity": "warn",
                "value": slo["latency_p95_ms"],
                "threshold": SLO_ALERT_P95_MS_THRESHOLD,
                "message": "API p95 latency exceeded threshold.",
            }
        )

    if slo["opa_unavailable"] >= SLO_ALERT_OPA_UNAVAILABLE_THRESHOLD:
        active_alerts.append(
            {
                "id": "opa.unavailable",
                "severity": "critical",
                "value": slo["opa_unavailable"],
                "threshold": SLO_ALERT_OPA_UNAVAILABLE_THRESHOLD,
                "message": "OPA unavailable counter exceeded threshold.",
            }
        )

    if slack["dispatch_failures"] >= SLO_ALERT_SLACK_FAILURES_THRESHOLD:
        active_alerts.append(
            {
                "id": "slack.dispatch-failures",
                "severity": "warn",
                "value": slack["dispatch_failures"],
                "threshold": SLO_ALERT_SLACK_FAILURES_THRESHOLD,
                "message": "Slack webhook dispatch failures exceeded threshold.",
            }
        )

    if slack["digest_items_buffered"] >= SLO_ALERT_DIGEST_BACKLOG_THRESHOLD:
        active_alerts.append(
            {
                "id": "slack.digest-backlog",
                "severity": "warn",
                "value": slack["digest_items_buffered"],
                "threshold": SLO_ALERT_DIGEST_BACKLOG_THRESHOLD,
                "message": "Slack digest backlog exceeded threshold.",
            }
        )

    overall = "ok"
    for alert in active_alerts:
        if _severity_rank(alert["severity"]) > _severity_rank(overall):
            overall = alert["severity"]

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": overall,
        "active_alerts_count": len(active_alerts),
        "active_alerts": active_alerts,
        "thresholds": {
            "error_rate": SLO_ALERT_ERROR_RATE_THRESHOLD,
            "latency_p95_ms": SLO_ALERT_P95_MS_THRESHOLD,
            "opa_unavailable": SLO_ALERT_OPA_UNAVAILABLE_THRESHOLD,
            "slack_dispatch_failures": SLO_ALERT_SLACK_FAILURES_THRESHOLD,
            "slack_digest_backlog": SLO_ALERT_DIGEST_BACKLOG_THRESHOLD,
        },
    }


def _prometheus_escape_label(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")


def _prometheus_render_slo(snapshot: dict) -> str:
    state_write_metrics = snapshot.get("state_write_metrics", {})
    if not isinstance(state_write_metrics, dict):
        state_write_metrics = {}

    lines = [
        "# HELP firewall_requests_total Total API requests processed.",
        "# TYPE firewall_requests_total counter",
        f"firewall_requests_total {snapshot.get('requests_total', 0)}",
        "# HELP firewall_requests_error Total API 5xx responses.",
        "# TYPE firewall_requests_error counter",
        f"firewall_requests_error {snapshot.get('requests_error', 0)}",
        "# HELP firewall_rate_limited_total Total API requests rejected with HTTP 429.",
        "# TYPE firewall_rate_limited_total counter",
        f"firewall_rate_limited_total {snapshot.get('requests_rate_limited', 0)}",
        "# HELP firewall_decisions_total Total decisions evaluated.",
        "# TYPE firewall_decisions_total counter",
        f"firewall_decisions_total {snapshot.get('decisions_total', 0)}",
        "# HELP firewall_decisions_deny Total denied decisions.",
        "# TYPE firewall_decisions_deny counter",
        f"firewall_decisions_deny {snapshot.get('decisions_deny', 0)}",
        "# HELP firewall_opa_unavailable_total OPA unavailable events.",
        "# TYPE firewall_opa_unavailable_total counter",
        f"firewall_opa_unavailable_total {snapshot.get('opa_unavailable', 0)}",
        "# HELP firewall_error_rate API error rate.",
        "# TYPE firewall_error_rate gauge",
        f"firewall_error_rate {snapshot.get('error_rate', 0)}",
        "# HELP firewall_deny_rate Decision deny rate.",
        "# TYPE firewall_deny_rate gauge",
        f"firewall_deny_rate {snapshot.get('deny_rate', 0)}",
        "# HELP firewall_latency_p50_ms API latency p50 in milliseconds.",
        "# TYPE firewall_latency_p50_ms gauge",
        f"firewall_latency_p50_ms {snapshot.get('latency_p50_ms', 0)}",
        "# HELP firewall_latency_avg_ms API mean latency in milliseconds.",
        "# TYPE firewall_latency_avg_ms gauge",
        f"firewall_latency_avg_ms {snapshot.get('latency_avg_ms', 0)}",
        "# HELP firewall_latency_p95_ms API latency p95 in milliseconds.",
        "# TYPE firewall_latency_p95_ms gauge",
        f"firewall_latency_p95_ms {snapshot.get('latency_p95_ms', 0)}",
        "# HELP firewall_failed_standard_ms_nfr_total Total denied decisions failing Clarisys NFR.",
        "# TYPE firewall_failed_standard_ms_nfr_total counter",
        f"firewall_failed_standard_ms_nfr_total {snapshot.get('failed_standard_ms_nfr_total', 0)}",
        "# HELP firewall_failed_standard_iso_27001_total Total denied decisions failing ISO 27001.",
        "# TYPE firewall_failed_standard_iso_27001_total counter",
        f"firewall_failed_standard_iso_27001_total {snapshot.get('failed_standard_iso_27001_total', 0)}",
        "# HELP firewall_failed_standard_cis_v81_total Total denied decisions failing CIS v8.1.",
        "# TYPE firewall_failed_standard_cis_v81_total counter",
        f"firewall_failed_standard_cis_v81_total {snapshot.get('failed_standard_cis_v81_total', 0)}",
        "# HELP firewall_failed_standard_pci_dss_total Total denied decisions failing PCI-DSS.",
        "# TYPE firewall_failed_standard_pci_dss_total counter",
        f"firewall_failed_standard_pci_dss_total {snapshot.get('failed_standard_pci_dss_total', 0)}",
        "# HELP firewall_active_alerts_count Number of currently active alerts.",
        "# TYPE firewall_active_alerts_count gauge",
        f"firewall_active_alerts_count {snapshot.get('active_alerts_count', 0)}",
        "# HELP firewall_slack_dispatch_latency_count Total measured Slack dispatch attempts.",
        "# TYPE firewall_slack_dispatch_latency_count counter",
        f"firewall_slack_dispatch_latency_count {snapshot.get('slack_dispatch_latency_count', 0)}",
        "# HELP firewall_slack_dispatch_latency_avg_ms Average Slack dispatch latency in milliseconds.",
        "# TYPE firewall_slack_dispatch_latency_avg_ms gauge",
        f"firewall_slack_dispatch_latency_avg_ms {snapshot.get('slack_dispatch_latency_avg_ms', 0)}",
        "# HELP firewall_slack_dispatch_latency_p50_ms Slack dispatch latency p50 in milliseconds.",
        "# TYPE firewall_slack_dispatch_latency_p50_ms gauge",
        f"firewall_slack_dispatch_latency_p50_ms {snapshot.get('slack_dispatch_latency_p50_ms', 0)}",
        "# HELP firewall_slack_dispatch_latency_p95_ms Slack dispatch latency p95 in milliseconds.",
        "# TYPE firewall_slack_dispatch_latency_p95_ms gauge",
        f"firewall_slack_dispatch_latency_p95_ms {snapshot.get('slack_dispatch_latency_p95_ms', 0)}",
        "# HELP firewall_slack_dispatch_latency_max_ms Maximum Slack dispatch latency in milliseconds.",
        "# TYPE firewall_slack_dispatch_latency_max_ms gauge",
        f"firewall_slack_dispatch_latency_max_ms {snapshot.get('slack_dispatch_latency_max_ms', 0)}",
        "# HELP firewall_slack_dispatch_latency_last_ms Last measured Slack dispatch latency in milliseconds.",
        "# TYPE firewall_slack_dispatch_latency_last_ms gauge",
        f"firewall_slack_dispatch_latency_last_ms {snapshot.get('slack_dispatch_latency_last_ms', 0)}",
        "# HELP firewall_state_write_total Total state-file write attempts by component and outcome.",
        "# TYPE firewall_state_write_total counter",
    ]
    for component in sorted(state_write_metrics):
        outcomes = state_write_metrics.get(component)
        if not isinstance(outcomes, dict):
            continue
        component_label = _prometheus_escape_label(str(component))
        success = int(outcomes.get("success", 0))
        failure = int(outcomes.get("failure", 0))
        lines.append(
            f'firewall_state_write_total{{component="{component_label}",outcome="success"}} {success}'
        )
        lines.append(
            f'firewall_state_write_total{{component="{component_label}",outcome="failure"}} {failure}'
        )
    return "\n".join(lines) + "\n"


def _prometheus_render_alerts(snapshot: dict) -> str:
    status = str(snapshot.get("status", "ok"))
    alerts = snapshot.get("active_alerts", []) if isinstance(snapshot.get("active_alerts"), list) else []
    thresholds = snapshot.get("thresholds", {}) if isinstance(snapshot.get("thresholds"), dict) else {}

    lines = [
        "# HELP firewall_alerts_active_count Number of active alerts.",
        "# TYPE firewall_alerts_active_count gauge",
        f"firewall_alerts_active_count {snapshot.get('active_alerts_count', 0)}",
        "# HELP firewall_alert_status Current alert status by level.",
        "# TYPE firewall_alert_status gauge",
    ]
    for level in ("ok", "warn", "critical"):
        lines.append(f'firewall_alert_status{{level="{level}"}} {1 if status == level else 0}')

    lines.extend(
        [
            "# HELP firewall_alert_active Active alert indicator by alert id.",
            "# TYPE firewall_alert_active gauge",
            "# HELP firewall_alert_value Current observed value for alert.",
            "# TYPE firewall_alert_value gauge",
            "# HELP firewall_alert_threshold Configured threshold for alert.",
            "# TYPE firewall_alert_threshold gauge",
        ]
    )

    for alert in alerts:
        alert_id = _prometheus_escape_label(str(alert.get("id", "unknown")))
        severity = _prometheus_escape_label(str(alert.get("severity", "warn")))
        lines.append(f'firewall_alert_active{{alert_id="{alert_id}",severity="{severity}"}} 1')

        value = alert.get("value")
        if isinstance(value, (int, float)):
            lines.append(
                f'firewall_alert_value{{alert_id="{alert_id}",severity="{severity}"}} {float(value)}'
            )
        threshold = alert.get("threshold")
        if isinstance(threshold, (int, float)):
            lines.append(
                f'firewall_alert_threshold{{alert_id="{alert_id}",severity="{severity}"}} {float(threshold)}'
            )

    lines.extend(
        [
            "# HELP firewall_alert_threshold_config Configured alert thresholds.",
            "# TYPE firewall_alert_threshold_config gauge",
        ]
    )
    for name, value in thresholds.items():
        if isinstance(value, (int, float)):
            metric_name = _prometheus_escape_label(str(name))
            lines.append(f'firewall_alert_threshold_config{{name="{metric_name}"}} {float(value)}')

    return "\n".join(lines) + "\n"


def _ensure_evidence_store_ready() -> None:
    EVIDENCE_DIR.mkdir(parents=True, exist_ok=True)


def _safe_parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return None


def _prune_evidence_archive() -> None:
    _ensure_evidence_store_ready()
    cutoff = datetime.now(timezone.utc) - timedelta(days=max(1, EVIDENCE_RETENTION_DAYS))
    kept: list[dict] = []

    with _EVIDENCE_LOCK:
        if EVIDENCE_INDEX_FILE.exists():
            with EVIDENCE_INDEX_FILE.open("r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        item = json.loads(line)
                    except Exception:
                        continue
                    created_at = _safe_parse_dt(str(item.get("created_at", "")))
                    if created_at and created_at >= cutoff:
                        kept.append(item)
                    else:
                        file_name = str(item.get("file_name", "")).strip()
                        if file_name:
                            try:
                                (EVIDENCE_DIR / file_name).unlink(missing_ok=True)
                            except Exception:
                                pass

        with EVIDENCE_INDEX_FILE.open("w", encoding="utf-8") as fh:
            for item in kept:
                fh.write(json.dumps(item, separators=(",", ":")) + "\n")


def _archive_evidence_report(
    *,
    report_id: str,
    days: int,
    total: int,
    acceptable: int,
    denied: int,
    output_format: str,
    content: str,
) -> dict:
    _ensure_evidence_store_ready()
    _prune_evidence_archive()

    created_at = datetime.now(timezone.utc).isoformat()
    ext = "md" if output_format == "markdown" else output_format
    file_name = f"evidence-{report_id}.{ext}"
    report_path = EVIDENCE_DIR / file_name
    report_path.write_text(content, encoding="utf-8")

    digest = hashlib.sha256(content.encode("utf-8")).hexdigest()
    entry = {
        "report_id": report_id,
        "created_at": created_at,
        "window_days": days,
        "total": total,
        "acceptable": acceptable,
        "denied": denied,
        "format": output_format,
        "file_name": file_name,
        "sha256": digest,
        "size_bytes": report_path.stat().st_size,
    }

    with _EVIDENCE_LOCK:
        with EVIDENCE_INDEX_FILE.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, separators=(",", ":")) + "\n")

    return entry


def _load_evidence_index() -> list[dict]:
    _ensure_evidence_store_ready()
    if not EVIDENCE_INDEX_FILE.exists():
        return []

    rows: list[dict] = []
    with _EVIDENCE_LOCK:
        with EVIDENCE_INDEX_FILE.open("r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except Exception:
                    continue
    return rows


def _evidence_entry_by_report_id(report_id: str) -> dict | None:
    for row in reversed(_load_evidence_index()):
        if str(row.get("report_id")) == report_id:
            return row
    return None


# ── OPA evaluation ─────────────────────────────────────────────────────────────
# OPA runs as a long-lived server (see deploy/opa.service) and is reached over
# loopback HTTP. Per-request fork+exec and rego compilation of the previous
# `opa eval` subprocess approach are eliminated; batched evaluation collapses
# N bulk-items into a single OPA round-trip via policy.request_standards_batch.
OPA_HOST = os.environ.get("OPA_HOST", "127.0.0.1")
OPA_PORT = int(os.environ.get("OPA_PORT", "8181"))
OPA_TIMEOUT = float(os.environ.get("OPA_TIMEOUT", "30"))
_OPA_SINGLE_PATH = "/v1/data/policy/request_standards/decision"
_OPA_BATCH_PATH = "/v1/data/policy/request_standards_batch/decisions"

# === OPA decision cache ==========================================================
# In-process LRU cache keyed on SHA-256 of canonical-JSON of the OPA input.
# Invalidates automatically when any *.rego or *.json under the policy/ directory
# changes (mtime poll, capped at OPA_CACHE_TTL seconds between checks).
OPA_CACHE_SIZE = int(os.environ.get("OPA_CACHE_SIZE", "10000"))   # 0 = disabled
OPA_CACHE_TTL  = float(os.environ.get("OPA_CACHE_TTL", "5"))      # mtime-poll cooldown
_POLICY_DIR    = Path(__file__).parent.parent / "policy"


class _OpaCache:
    """Thread-safe LRU with lazy mtime-based invalidation."""

    def __init__(self, max_size: int, ttl: float, watch_dir: Path) -> None:
        from collections import OrderedDict
        self._max = max_size
        self._ttl = ttl
        self._watch_dir = watch_dir
        self._lock = threading.Lock()
        self._store: "OrderedDict[str, dict]" = OrderedDict()
        self._version: float = 0.0
        self._last_check: float = 0.0
        self.hits = 0
        self.misses = 0
        self.evictions = 0
        self.invalidations = 0
        self._refresh_version(force=True)

    def _current_max_mtime(self) -> float:
        m = 0.0
        try:
            for p in self._watch_dir.iterdir():
                if p.suffix in (".rego", ".json"):
                    try:
                        mt = p.stat().st_mtime
                        if mt > m:
                            m = mt
                    except OSError:
                        pass
        except OSError:
            pass
        return m

    def _refresh_version(self, force: bool = False) -> None:
        now = time.monotonic()
        if not force and (now - self._last_check) < self._ttl:
            return
        self._last_check = now
        current = self._current_max_mtime()
        if current > self._version:
            if self._version != 0.0:
                self.invalidations += 1
                self._store.clear()
            self._version = current

    @staticmethod
    def key_for(opa_input: dict) -> str:
        canonical = json.dumps(opa_input, sort_keys=True, separators=(",", ":"))
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    def get(self, key: str):
        if self._max <= 0:
            return None
        with self._lock:
            self._refresh_version()
            val = self._store.get(key)
            if val is None:
                self.misses += 1
                return None
            self._store.move_to_end(key)
            self.hits += 1
            return val

    def put(self, key: str, value: dict) -> None:
        if self._max <= 0:
            return
        with self._lock:
            self._refresh_version()
            self._store[key] = value
            self._store.move_to_end(key)
            while len(self._store) > self._max:
                self._store.popitem(last=False)
                self.evictions += 1

    def clear(self) -> None:
        with self._lock:
            self._store.clear()

    def stats(self) -> dict:
        with self._lock:
            return {
                "enabled": self._max > 0,
                "size": len(self._store),
                "max_size": self._max,
                "ttl_seconds": self._ttl,
                "hits": self.hits,
                "misses": self.misses,
                "evictions": self.evictions,
                "invalidations": self.invalidations,
                "policy_version_mtime": self._version,
            }


_opa_cache = _OpaCache(OPA_CACHE_SIZE, OPA_CACHE_TTL, _POLICY_DIR)
# === end OPA decision cache =====================================================


# One HTTP connection per worker thread for keep-alive.
_opa_tls = threading.local()


def _opa_conn() -> http.client.HTTPConnection:
    conn = getattr(_opa_tls, "conn", None)
    if conn is None:
        conn = http.client.HTTPConnection(OPA_HOST, OPA_PORT, timeout=OPA_TIMEOUT)
        _opa_tls.conn = conn
    return conn


def _opa_post(path: str, body: bytes) -> dict:
    global _CB_FAILURES, _CB_OPEN_UNTIL
    mode = "batch" if path == _OPA_BATCH_PATH else "single"
    started_at = time.perf_counter()

    with _CB_LOCK:
        if _CB_OPEN_UNTIL > time.monotonic():
            with _SLO_LOCK:
                _SLO_COUNTERS["opa_unavailable"] += 1
            _save_slo_state()
            record_opa_request(mode=mode, outcome="circuit_open")
            raise HTTPException(status_code=503, detail="OPA circuit breaker open")

    headers = {"Content-Type": "application/json", "Content-Length": str(len(body))}
    last_exc: Exception | None = None
    for _ in (1, 2):
        conn = _opa_conn()
        try:
            conn.request("POST", path, body=body, headers=headers)
            resp = conn.getresponse()
            data = resp.read()
            if resp.status != 200:
                raise HTTPException(
                    status_code=502,
                    detail=f"OPA returned HTTP {resp.status}: {data[:200].decode('utf-8', 'replace')}",
                )
            try:
                payload = json.loads(data)
                record_opa_request(
                    mode=mode,
                    outcome="success",
                    latency_seconds=time.perf_counter() - started_at,
                )
                return payload
            except json.JSONDecodeError as exc:
                record_opa_request(
                    mode=mode,
                    outcome="error",
                    latency_seconds=time.perf_counter() - started_at,
                )
                raise HTTPException(
                    status_code=500,
                    detail=f"Unexpected OPA response: {data[:200]!r} ({exc})",
                )
        except HTTPException:
            record_opa_request(
                mode=mode,
                outcome="error",
                latency_seconds=time.perf_counter() - started_at,
            )
            raise
        except (http.client.HTTPException, ConnectionError, OSError, TimeoutError) as exc:
            last_exc = exc
            with _CB_LOCK:
                _CB_FAILURES += 1
                if _CB_FAILURES >= _OPA_CB_FAILURE_THRESHOLD:
                    _CB_OPEN_UNTIL = time.monotonic() + _OPA_CB_COOLDOWN_SECONDS
            try:
                conn.close()
            except Exception:
                pass
            _opa_tls.conn = None
    timed_out = isinstance(last_exc, TimeoutError) or "timed out" in str(last_exc).lower()
    record_opa_request(
        mode=mode,
        outcome="timeout" if timed_out else "error",
        latency_seconds=time.perf_counter() - started_at,
        timed_out=timed_out,
    )
    with _SLO_LOCK:
        _SLO_COUNTERS["opa_unavailable"] += 1
    _save_slo_state()
    raise HTTPException(status_code=503, detail=f"OPA unavailable: {last_exc}")


def _evaluate_opa_input(opa_input: dict) -> dict:
    global _CB_FAILURES
    body = json.dumps({"input": opa_input}).encode("utf-8")
    payload = _opa_post(_OPA_SINGLE_PATH, body)
    result = payload.get("result")
    if not isinstance(result, dict):
        raise HTTPException(status_code=500, detail=f"Unexpected OPA response shape: {payload!r}")
    with _CB_LOCK:
        _CB_FAILURES = 0
    return result


def _evaluate(request: TrafficRequest) -> dict:
    """Build standards-only OPA input and call the OPA server for a single decision.

    Result is cached per-input until any policy/*.{rego,json} file changes.
    """
    opa_input = _build_standards_input(request)
    cache_key = _OpaCache.key_for(opa_input)
    cached = _opa_cache.get(cache_key)
    if cached is not None:
        return cached
    result = _evaluate_opa_input(opa_input)
    _opa_cache.put(cache_key, result)
    return result


def _evaluate_batch(requests_list: list) -> list[dict]:
    """Evaluate a batch of requests in a single OPA round-trip.

    Cached inputs are served from the in-process cache; only uncached inputs
    are sent to OPA, then the results are spliced back into their original
    positions. Falls back to per-item evaluation if OPA returns an unexpected
    shape.
    """
    if not requests_list:
        return []
    opa_inputs = [_build_standards_input(r) for r in requests_list]
    results: list[dict | None] = [None] * len(requests_list)
    miss_indices: list[int] = []
    miss_inputs: list[dict] = []
    miss_keys: list[str] = []
    for i, opa_input in enumerate(opa_inputs):
        key = _OpaCache.key_for(opa_input)
        cached = _opa_cache.get(key)
        if cached is not None:
            results[i] = cached
        else:
            miss_indices.append(i)
            miss_inputs.append(opa_input)
            miss_keys.append(key)
    if miss_inputs:
        body = json.dumps({"input": {"requests": miss_inputs}}).encode("utf-8")
        payload = _opa_post(_OPA_BATCH_PATH, body)
        opa_results = payload.get("result")
        if not isinstance(opa_results, list) or len(opa_results) != len(miss_inputs):
            fallback = [_evaluate(requests_list[i]) for i in miss_indices]
            for slot, dec in zip(miss_indices, fallback):
                results[slot] = dec
        else:
            for slot, key, dec in zip(miss_indices, miss_keys, opa_results):
                results[slot] = dec
                _opa_cache.put(key, dec)
    return results  # type: ignore[return-value]


# ── Endpoints ──────────────────────────────────────────────────────────────────
def _record_audit(
    request: Request,
    caller: CallerIdentity,
    endpoint: str,
    payload_summary: dict,
    verdict_summary: dict,
) -> None:
    """Append a single evaluation record to the immutable audit trail."""
    timer: RequestTimer | None = getattr(request.state, "timer", None)
    elapsed = timer.elapsed_ms() if timer else 0
    event = make_event(
        request_id=getattr(request.state, "request_id", "unknown"),
        endpoint=endpoint,
        caller_sub=caller.sub if caller else None,
        payload_summary=payload_summary,
        verdict_summary=verdict_summary,
        elapsed_ms=elapsed,
    )
    try:
        get_audit_store().record(event)
    except Exception:  # noqa: BLE001 — audit-trail failure must not break the request
        log.exception("audit_store.record_failed", endpoint=endpoint)


def _record_decision_history(
    request: Request,
    caller: CallerIdentity,
    endpoint: str,
    action_requested: str,
    verdict: str,
    overall_status: str,
    overall_risk: str,
    details: dict,
    policy_input: dict | None = None,
    reason: str | None = None,
    remediations: list[str] | None = None,
) -> None:
    """Append one accept/decline decision for historical context."""
    try:
        request_id = getattr(request.state, "request_id", "unknown")
        decision_id = _decision_id(request_id, endpoint)
        merged_details = dict(details)
        if policy_input is not None:
            merged_details["policy_input"] = policy_input
        if reason:
            merged_details["reason"] = reason
        if remediations:
            merged_details["remediations"] = remediations

        append_decision_history(
            {
                "decision_id": decision_id,
                "request_id": request_id,
                "endpoint": endpoint,
                "caller_sub": caller.sub if caller else "anonymous",
                "action_requested": action_requested,
                "decision_verdict": verdict,
                "overall_status": overall_status,
                "overall_risk": overall_risk,
                "details": merged_details,
            }
        )
        slack_payload = {
            "decision_id": decision_id,
            "request_id": request_id,
            "endpoint": endpoint,
            "caller_sub": caller.sub if caller else "anonymous",
            "action_requested": action_requested,
            "decision_verdict": verdict,
            "overall_status": overall_status,
            "overall_risk": overall_risk,
            "details": merged_details,
        }
        _emit_slack_decision(slack_payload)
        set_decision_lifecycle(
            decision_id=decision_id,
            status="evaluated",
            actor=caller.sub if caller else "anonymous",
            notes="Auto-created on evaluation",
        )
        with _SLO_LOCK:
            _SLO_COUNTERS["decisions_total"] += 1
            if verdict == "DENY":
                _SLO_COUNTERS["decisions_deny"] += 1
                failed_standards = details.get("failed_standards", [])
                if isinstance(failed_standards, list):
                    for standard in failed_standards:
                        if not isinstance(standard, str):
                            continue
                        counter_key = _canonical_standard_key(standard)
                        if counter_key is not None:
                            _SLO_COUNTERS[counter_key] = int(_SLO_COUNTERS.get(counter_key, 0)) + 1
        _save_slo_state()

        if verdict == "DENY" or overall_risk in {"HIGH", "CRITICAL"}:
            _emit_event(
                "decision.recorded",
                {
                    "decision_id": decision_id,
                    "request_id": request_id,
                    "endpoint": endpoint,
                    "verdict": verdict,
                    "overall_risk": overall_risk,
                    "overall_status": overall_status,
                },
            )
    except Exception:  # noqa: BLE001 — history write failures should not fail evaluation
        log.exception("decision_history.record_failed", endpoint=endpoint)


def _build_verdict(request: TrafficRequest, decision: dict) -> StandardsVerdict:
    summary = decision.get("summary", {})
    violations = _filter_violations_by_standards(decision.get("violations", []), request.standards)
    failed_controls = _filter_controls_by_standards(
        sorted({v.get("control") for v in violations if isinstance(v, dict) and v.get("control")}),
        request.standards,
    )
    failed_standards = _collect_failed_standards(violations, selected_standards=request.standards)
    framework_scope = failed_standards or request.standards
    framework_clauses, control_clause_mappings = _resolve_clause_mappings(
        failed_controls,
        framework_scope,
    )
    policy_meta = _policy_metadata()
    local_request_id = request.request_id or "generated"

    return StandardsVerdict(
        decision_id=_decision_id(local_request_id, "/evaluate"),
        verdict="ACCEPTABLE" if decision.get("compliant") else "DENY",
        allow=decision.get("compliant", False),
        reason=_build_reason(request, decision),
        overall_status=summary.get("status", "COMPLIANT" if decision.get("compliant") else "NON-COMPLIANT"),
        overall_risk=summary.get("overall_risk", "LOW"),
        failed_standards=failed_standards,
        failed_controls=failed_controls,
        framework_clauses=framework_clauses,
        control_clause_mappings=control_clause_mappings,
        violations_count=decision.get("violations_count", 0),
        violations=violations,
        request=request.model_dump(),
        policy_version=policy_meta["policy_version"],
        policy_hash=policy_meta["policy_hash"],
        policy_signature=policy_meta["policy_signature"],
    )


@app.post(
    "/evaluate",
    response_model=StandardsVerdict,
    summary="Evaluate a proposed request against standards",
    description=(
        "Pass any source IP/FQDN and destination IP/FQDN with protocol and port. "
        "The API evaluates the proposed addition against security standards only, "
        "not the existing firewall ruleset, and returns an immediate standards-based verdict."
    ),
)
def evaluate_traffic(
    request: TrafficRequest,
    http_request: Request,
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> StandardsVerdict:
    verdict = _build_verdict(request, _evaluate(request))
    verdict.decision_id = _decision_id(getattr(http_request.state, "request_id", "unknown"), "/evaluate")
    req_id = getattr(http_request.state, "request_id", "unknown")
    include_in_dashboard_metrics = not _is_synthetic_request(http_request)
    if include_in_dashboard_metrics:
        # Record ROI metric for non-synthetic traffic only.
        record_rule_processed(req_id, "/evaluate", verdict.verdict)
    _record_audit(
        http_request,
        caller,
        endpoint="/evaluate",
        payload_summary={
            "source": request.source,
            "destination": request.destination,
            "protocol": request.protocol,
            "port": request.port,
        },
        verdict_summary={
            "verdict": verdict.verdict,
            "overall_status": verdict.overall_status,
            "overall_risk": verdict.overall_risk,
            "failed_controls": verdict.failed_controls,
        },
    )
    if include_in_dashboard_metrics:
        _record_decision_history(
            http_request,
            caller,
            endpoint="/evaluate",
            action_requested=request.action,
            verdict=verdict.verdict,
            overall_status=verdict.overall_status,
            overall_risk=verdict.overall_risk,
            details={
                "source": request.source,
                "destination": request.destination,
                "protocol": request.protocol,
                "port": request.port,
                "failed_controls": verdict.failed_controls,
                "failed_standards": verdict.failed_standards,
            },
            policy_input=_build_standards_input(request),
            reason=verdict.reason,
            remediations=_collect_remediations(verdict.violations),
        )
    return verdict



# Maximum parallel OPA subprocesses for bulk evaluation.
# Tuned to the 2 uvicorn workers × typical CPU headroom; raise via env var.
_BULK_MAX_WORKERS = int(os.environ.get("BULK_MAX_WORKERS", "8"))


def _evaluate_one(req: "TrafficRequest") -> "StandardsVerdict":
    """Evaluate a single request — retained for legacy single-item callers."""
    return _build_verdict(req, _evaluate(req))


def _compute_evaluate_bulk(payload: BulkRequest) -> BulkResponse:
    by_standard: dict[str, int] = {}
    by_control: dict[str, int] = {}
    failed_controls: set[str] = set()
    acceptable = 0

    decisions = _evaluate_batch(payload.requests)
    results: list[StandardsVerdict] = [
        _build_verdict(req, dec) for req, dec in zip(payload.requests, decisions)
    ]

    for verdict in results:
        if verdict.allow:
            acceptable += 1
        for std in verdict.failed_standards:
            by_standard[std] = by_standard.get(std, 0) + 1
        for ctrl in verdict.failed_controls:
            by_control[ctrl] = by_control.get(ctrl, 0) + 1
            failed_controls.add(ctrl)

    total = len(results)
    return BulkResponse(
        summary=BulkSummary(
            total=total,
            acceptable=acceptable,
            denied=total - acceptable,
            by_failed_standard=dict(sorted(by_standard.items(), key=lambda kv: -kv[1])),
            by_failed_control=dict(sorted(by_control.items(), key=lambda kv: -kv[1])),
            failed_controls=sorted(failed_controls),
            overall_status="COMPLIANT" if acceptable == total else "NON-COMPLIANT",
        ),
        results=results,
    )


@app.post(
    "/evaluate/bulk",
    response_model=BulkResponse,
    summary="Evaluate many proposed requests in one call",
    description=(
        "Submit `{\"requests\": [...]}` containing 1–500 proposed traffic requests. "
        "Each item is evaluated against the security-standards policy and returned with "
        "its own verdict, plus an aggregated summary across the batch."
    ),
)
def evaluate_bulk(
    payload: BulkRequest,
    http_request: Request,
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> BulkResponse:
    # Capture caller for rate limiting
    http_request.state.caller_id = caller.sub
    response = _compute_evaluate_bulk(payload)
    summary = response.summary
    _record_audit(
        http_request,
        caller,
        endpoint="/evaluate/bulk",
        payload_summary={"total": summary.total},
        verdict_summary={
            "acceptable": summary.acceptable,
            "denied": summary.denied,
            "overall_status": summary.overall_status,
            "failed_controls": summary.failed_controls,
        },
    )
    req_id = getattr(http_request.state, "request_id", "unknown")
    if not _is_synthetic_request(http_request):
        for idx, (req, verdict) in enumerate(zip(payload.requests, response.results)):
            # Composite dedup key matches /evaluate/bulk/stream so a request
            # processed twice (e.g. nginx retry) is counted once.
            record_rule_processed(f"{req_id}:{idx}", "/evaluate/bulk", verdict.verdict)
            _record_decision_history(
                http_request,
                caller,
                endpoint="/evaluate/bulk",
                action_requested=req.action,
                verdict=verdict.verdict,
                overall_status=verdict.overall_status,
                overall_risk=verdict.overall_risk,
                details={
                    "source": req.source,
                    "destination": req.destination,
                    "protocol": req.protocol,
                    "port": req.port,
                    "failed_controls": verdict.failed_controls,
                    "failed_standards": verdict.failed_standards,
                },
                policy_input=_build_standards_input(req),
                reason=verdict.reason,
                remediations=_collect_remediations(verdict.violations),
            )
    return response


# ── Streaming bulk endpoints ────────────────────────────────────────────
# Each chunk is one OPA batch call; verdicts are flushed as NDJSON as soon as
# each chunk completes, so TTFB scales with the chunk size rather than the
# whole batch size.
_BULK_STREAM_CHUNK = int(os.environ.get("BULK_STREAM_CHUNK", "100"))


def _bulk_stream_generator(requests_list, decision_recorder=None):
    by_standard: dict[str, int] = {}
    by_control: dict[str, int] = {}
    failed_controls: set[str] = set()
    acceptable = 0
    total = len(requests_list)
    chunk = max(1, _BULK_STREAM_CHUNK)

    for start in range(0, total, chunk):
        batch = requests_list[start:start + chunk]
        decisions = _evaluate_batch(batch)
        for req, dec in zip(batch, decisions):
            verdict = _build_verdict(req, dec)
            if verdict.allow:
                acceptable += 1
            for std in verdict.failed_standards:
                by_standard[std] = by_standard.get(std, 0) + 1
            for ctrl in verdict.failed_controls:
                by_control[ctrl] = by_control.get(ctrl, 0) + 1
                failed_controls.add(ctrl)
            if decision_recorder is not None:
                decision_recorder(req, verdict)
            yield (json.dumps({"type": "verdict", "data": verdict.model_dump()}) + "\n").encode("utf-8")

    summary = {
        "type": "summary",
        "data": {
            "total": total,
            "acceptable": acceptable,
            "denied": total - acceptable,
            "by_failed_standard": dict(sorted(by_standard.items(), key=lambda kv: -kv[1])),
            "by_failed_control": dict(sorted(by_control.items(), key=lambda kv: -kv[1])),
            "failed_controls": sorted(failed_controls),
            "overall_status": "COMPLIANT" if acceptable == total else "NON-COMPLIANT",
        },
    }
    yield (json.dumps(summary) + "\n").encode("utf-8")


def _intake_bulk_stream_generator(intake_requests, decision_recorder=None):
    by_standard: dict[str, int] = {}
    by_control: dict[str, int] = {}
    failed_controls: set[str] = set()
    acceptable = 0
    total_risk_score = 0
    max_risk_score = 0
    total = len(intake_requests)
    chunk = max(1, _BULK_STREAM_CHUNK)

    for start in range(0, total, chunk):
        batch = intake_requests[start:start + chunk]
        traffic_inputs = [_intake_to_traffic(intake) for intake in batch]
        decisions = _evaluate_batch(traffic_inputs)
        for intake, traffic, dec in zip(batch, traffic_inputs, decisions):
            verdict = _build_intake_verdict(intake, traffic, dec)
            if verdict.allow:
                acceptable += 1
            total_risk_score += verdict.risk_score
            if verdict.risk_score > max_risk_score:
                max_risk_score = verdict.risk_score
            for std in verdict.failed_standards:
                by_standard[std] = by_standard.get(std, 0) + 1
            for ctrl in verdict.failed_controls:
                by_control[ctrl] = by_control.get(ctrl, 0) + 1
                failed_controls.add(ctrl)
            if decision_recorder is not None:
                decision_recorder(intake, verdict)
            yield (json.dumps({"type": "verdict", "data": verdict.model_dump()}) + "\n").encode("utf-8")

    summary = {
        "type": "summary",
        "data": {
            "total": total,
            "acceptable": acceptable,
            "denied": total - acceptable,
            "total_risk_score": total_risk_score,
            "max_risk_score": max_risk_score,
            "by_failed_standard": dict(sorted(by_standard.items(), key=lambda kv: -kv[1])),
            "by_failed_control": dict(sorted(by_control.items(), key=lambda kv: -kv[1])),
            "failed_controls": sorted(failed_controls),
            "overall_status": "COMPLIANT" if acceptable == total else "NON-COMPLIANT",
        },
    }
    yield (json.dumps(summary) + "\n").encode("utf-8")


@app.post(
    "/evaluate/bulk/stream",
    summary="Stream verdicts as NDJSON (one JSON object per line)",
    description=(
        "Same input as `/evaluate/bulk` but the response is streamed line-by-line "
        "as `application/x-ndjson`. Each line is `{\"type\":\"verdict\",\"data\":{...}}` "
        "until a final `{\"type\":\"summary\",\"data\":{...}}` line. "
        "Accepts up to 5000 items per call. Useful for very large batches where "
        "time-to-first-byte matters."
    ),
)
def evaluate_bulk_stream(
    payload: BulkStreamRequest,
    http_request: Request,
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
):
    http_request.state.caller_id = caller.sub
    total = len(payload.requests)
    req_id = getattr(http_request.state, "request_id", "unknown")
    _record_audit(
        http_request,
        caller,
        endpoint="/evaluate/bulk/stream",
        payload_summary={"total": total},
        verdict_summary={"streamed": True},
    )
    summary_payload = {
        "endpoint": "/evaluate/bulk/stream",
        "request_id": req_id,
        "total": total,
    }
    include_in_dashboard_metrics = not _is_synthetic_request(http_request)
    item_counter = [0]  # Use list to allow mutation in nested function
    def _record_stream_item(req: TrafficRequest, verdict: StandardsVerdict) -> None:
        if not include_in_dashboard_metrics:
            return
        # Composite dedup key: parent request_id + item index
        item_key = f"{req_id}:{item_counter[0]}"
        item_counter[0] += 1
        record_rule_processed(item_key, "/evaluate/bulk/stream", verdict.verdict)
        _record_decision_history(
            http_request,
            caller,
            endpoint="/evaluate/bulk/stream",
            action_requested=req.action,
            verdict=verdict.verdict,
            overall_status=verdict.overall_status,
            overall_risk=verdict.overall_risk,
            details={
                "source": req.source,
                "destination": req.destination,
                "protocol": req.protocol,
                "port": req.port,
                "failed_controls": verdict.failed_controls,
                "failed_standards": verdict.failed_standards,
            },
            policy_input=_build_standards_input(req),
            reason=verdict.reason,
            remediations=_collect_remediations(verdict.violations),
        )
    def _generator_with_summary():
        _t0 = time.monotonic()
        for chunk in _bulk_stream_generator(list(payload.requests), decision_recorder=_record_stream_item):
            try:
                parsed = json.loads(chunk.decode("utf-8"))
            except Exception:
                yield chunk
                continue
            if parsed.get("type") == "summary" and isinstance(parsed.get("data"), dict):
                _emit_slack_batch_summary(
                    "Firewall bulk stream summary",
                    {**summary_payload, **parsed["data"]},
                )
            yield chunk
        # Record total stream wall-clock duration after last byte is yielded.
        if include_in_dashboard_metrics:
            elapsed_s = time.monotonic() - _t0
            stream_duration_histogram.labels(endpoint="/evaluate/bulk/stream").observe(elapsed_s)
            with _SLO_LOCK:
                _STREAM_LATENCIES_MS["/evaluate/bulk/stream"].append(elapsed_s * 1000)
    return StreamingResponse(
        _generator_with_summary(),
        media_type="application/x-ndjson",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-store"},
    )


@app.post(
    "/intake/evaluate/bulk/stream",
    summary="Stream intake verdicts as NDJSON (one JSON object per line)",
    description=(
        "Same input as `/intake/evaluate/bulk` but streamed as NDJSON. "
        "Accepts up to 5000 items per call. Each `verdict` line is followed by a "
        "final `summary` line."
    ),
)
def evaluate_intake_bulk_stream(
    payload: IntakeBulkStreamRequest,
    http_request: Request,
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
):
    http_request.state.caller_id = caller.sub
    total = len(payload.requests)
    _record_audit(
        http_request,
        caller,
        endpoint="/intake/evaluate/bulk/stream",
        payload_summary={"total": total},
        verdict_summary={"streamed": True},
    )
    summary_payload = {
        "endpoint": "/intake/evaluate/bulk/stream",
        "request_id": getattr(http_request.state, "request_id", "unknown"),
        "total": total,
    }
    include_in_dashboard_metrics = not _is_synthetic_request(http_request)
    item_counter = [0]
    def _record_stream_item(intake: IntakeRequest, verdict: IntakeVerdict) -> None:
        if not include_in_dashboard_metrics:
            return
        item_key = f"{summary_payload['request_id']}:{item_counter[0]}"
        item_counter[0] += 1
        record_rule_processed(item_key, "/intake/evaluate/bulk/stream", verdict.verdict)
        _record_decision_history(
            http_request,
            caller,
            endpoint="/intake/evaluate/bulk/stream",
            action_requested=intake.action,
            verdict=verdict.verdict,
            overall_status=verdict.overall_status,
            overall_risk=verdict.overall_risk,
            details={
                "app_id": intake.app_id,
                "source_name": intake.source_name,
                "destination_name": intake.destination_name,
                "protocol": intake.protocol,
                "destination_port": intake.destination_port,
                "risk_score": verdict.risk_score,
                "failed_controls": verdict.failed_controls,
                "failed_standards": verdict.failed_standards,
            },
            policy_input=_build_standards_input(_intake_to_traffic(intake)),
            reason=verdict.reason,
            remediations=_collect_remediations(verdict.violations),
        )
    def _generator_with_summary():
        _t0 = time.monotonic()
        for chunk in _intake_bulk_stream_generator(list(payload.requests), decision_recorder=_record_stream_item):
            try:
                parsed = json.loads(chunk.decode("utf-8"))
            except Exception:
                yield chunk
                continue
            if parsed.get("type") == "summary" and isinstance(parsed.get("data"), dict):
                _emit_slack_batch_summary(
                    "Firewall intake bulk stream summary",
                    {**summary_payload, **parsed["data"]},
                )
            yield chunk
        # Record total stream wall-clock duration after last byte is yielded.
        if include_in_dashboard_metrics:
            elapsed_s = time.monotonic() - _t0
            stream_duration_histogram.labels(endpoint="/intake/evaluate/bulk/stream").observe(elapsed_s)
            with _SLO_LOCK:
                _STREAM_LATENCIES_MS["/intake/evaluate/bulk/stream"].append(elapsed_s * 1000)
    return StreamingResponse(
        _generator_with_summary(),
        media_type="application/x-ndjson",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-store"},
    )


def _intake_to_traffic(intake: IntakeRequest) -> TrafficRequest:
    """Map a logical IntakeRequest onto the raw TrafficRequest expected by _evaluate."""
    protocol_map = {"TCP": "tcp", "UDP": "udp", "ICMP": "icmp", "ANY": "any"}
    protocol = protocol_map[intake.protocol]

    port = intake.destination_port if intake.destination_port is not None else 0
    action = "accept" if intake.action == "ALLOW" else "deny"

    return TrafficRequest(
        source=intake.source_name,
        destination=intake.destination_name,
        protocol=protocol,
        port=port,
        log="all",
        action=action,
        source_interface=f"{intake.environment}-src",
        destination_interface=f"{intake.environment}-dst",
        standards=intake.standards,
    )


def _compute_risk_score(overall_risk: str) -> int:
    model = _risk_scores()
    return model.get(overall_risk.upper(), model["LOW"])


def _build_intake_verdict(intake: IntakeRequest, traffic: TrafficRequest, decision: dict) -> IntakeVerdict:
    summary = decision.get("summary", {})
    violations = _filter_violations_by_standards(decision.get("violations", []), intake.standards)
    failed_controls = _filter_controls_by_standards(
        sorted({v.get("control") for v in violations if isinstance(v, dict) and v.get("control")}),
        intake.standards,
    )
    failed_standards = _collect_failed_standards(violations, selected_standards=intake.standards)
    framework_scope = failed_standards or intake.standards
    framework_clauses, control_clause_mappings = _resolve_clause_mappings(
        failed_controls,
        framework_scope,
    )
    overall_risk = summary.get("overall_risk", "LOW")
    policy_meta = _policy_metadata()
    local_request_id = intake.request_id or "generated"

    return IntakeVerdict(
        decision_id=_decision_id(local_request_id, "/intake/evaluate"),
        verdict="ACCEPTABLE" if decision.get("compliant") else "DENY",
        allow=decision.get("compliant", False),
        reason=_build_reason(traffic, decision),
        overall_status=summary.get("status", "COMPLIANT" if decision.get("compliant") else "NON-COMPLIANT"),
        overall_risk=overall_risk,
        risk_score=_compute_risk_score(overall_risk),
        failed_standards=failed_standards,
        failed_controls=failed_controls,
        framework_clauses=framework_clauses,
        control_clause_mappings=control_clause_mappings,
        violations_count=decision.get("violations_count", 0),
        violations=violations,
        intake=intake.model_dump(mode="json"),
        policy_version=policy_meta["policy_version"],
        policy_hash=policy_meta["policy_hash"],
        policy_signature=policy_meta["policy_signature"],
    )


@app.post(
    "/intake/evaluate",
    response_model=IntakeVerdict,
    summary="Evaluate a logical intake request",
    description=(
        "Submit a structured intake request using CMDB/logical identifiers. "
        "The request is evaluated against the Clarisys security standards and returns a verdict. "
        "Risk is derived from the resulting standards classification only."
    ),
)
def evaluate_intake(
    http_request: Request,
    intake: IntakeRequest = Body(
        ...,
        openapi_examples=INTAKE_EVALUATE_EXAMPLES,
    ),
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> IntakeVerdict:
    # Capture caller for rate limiting
    http_request.state.caller_id = caller.sub
    traffic = _intake_to_traffic(intake)
    verdict = _build_intake_verdict(intake, traffic, _evaluate(traffic))
    verdict.decision_id = _decision_id(getattr(http_request.state, "request_id", "unknown"), "/intake/evaluate")
    include_in_dashboard_metrics = not _is_synthetic_request(http_request)
    if http_request is not None:
        _record_audit(
            http_request,
            caller,
            endpoint="/intake/evaluate",
            payload_summary={
                "app_id": intake.app_id,
                "environment": intake.environment,
                "source_name": intake.source_name,
                "destination_name": intake.destination_name,
                "protocol": intake.protocol,
                "destination_port": intake.destination_port,
            },
            verdict_summary={
                "verdict": verdict.verdict,
                "overall_status": verdict.overall_status,
                "overall_risk": verdict.overall_risk,
                "risk_score": verdict.risk_score,
                "failed_controls": verdict.failed_controls,
            },
        )
        if include_in_dashboard_metrics:
            record_rule_processed(
                getattr(http_request.state, "request_id", "unknown"),
                "/intake/evaluate",
                verdict.verdict,
            )
            _record_decision_history(
                http_request,
                caller,
                endpoint="/intake/evaluate",
                action_requested=intake.action,
                verdict=verdict.verdict,
                overall_status=verdict.overall_status,
                overall_risk=verdict.overall_risk,
                details={
                    "app_id": intake.app_id,
                    "source_name": intake.source_name,
                    "destination_name": intake.destination_name,
                    "protocol": intake.protocol,
                    "destination_port": intake.destination_port,
                    "risk_score": verdict.risk_score,
                    "failed_controls": verdict.failed_controls,
                    "failed_standards": verdict.failed_standards,
                },
                policy_input=_build_standards_input(traffic),
                reason=verdict.reason,
                remediations=_collect_remediations(verdict.violations),
            )
    return verdict


def _evaluate_one_intake(intake: "IntakeRequest") -> "IntakeVerdict":
    """Evaluate a single intake request — retained for legacy single-item callers."""
    traffic = _intake_to_traffic(intake)
    return _build_intake_verdict(intake, traffic, _evaluate(traffic))


def _compute_evaluate_intake_bulk(payload: IntakeBulkRequest) -> IntakeBulkResponse:
    by_standard: dict[str, int] = {}
    by_control: dict[str, int] = {}
    failed_controls: set[str] = set()
    acceptable = 0
    total_risk_score = 0
    max_risk_score = 0

    traffic_inputs = [_intake_to_traffic(intake) for intake in payload.requests]
    decisions = _evaluate_batch(traffic_inputs)
    results: list[IntakeVerdict] = [
        _build_intake_verdict(intake, traffic, dec)
        for intake, traffic, dec in zip(payload.requests, traffic_inputs, decisions)
    ]

    for verdict in results:
        if verdict.allow:
            acceptable += 1
        total_risk_score += verdict.risk_score
        max_risk_score = max(max_risk_score, verdict.risk_score)
        for std in verdict.failed_standards:
            by_standard[std] = by_standard.get(std, 0) + 1
        for ctrl in verdict.failed_controls:
            by_control[ctrl] = by_control.get(ctrl, 0) + 1
            failed_controls.add(ctrl)

    total = len(results)
    return IntakeBulkResponse(
        summary=IntakeBulkSummary(
            total=total,
            acceptable=acceptable,
            denied=total - acceptable,
            total_risk_score=total_risk_score,
            max_risk_score=max_risk_score,
            by_failed_standard=dict(sorted(by_standard.items(), key=lambda kv: -kv[1])),
            by_failed_control=dict(sorted(by_control.items(), key=lambda kv: -kv[1])),
            failed_controls=sorted(failed_controls),
            overall_status="COMPLIANT" if acceptable == total else "NON-COMPLIANT",
        ),
        results=results,
    )


@app.post(
    "/intake/evaluate/bulk",
    response_model=IntakeBulkResponse,
    summary="Evaluate many logical intake requests in one call",
    description=(
        "Submit `{\"requests\": [...]}` containing 1–500 logical intake requests. "
        "Each item is evaluated against the Clarisys security standards and returned with "
        "its own verdict, standards-derived numeric risk score, and an aggregated summary across the batch."
    ),
)
def evaluate_intake_bulk(
    http_request: Request,
    payload: IntakeBulkRequest = Body(
        ...,
        openapi_examples=INTAKE_BULK_EVALUATE_EXAMPLES,
    ),
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> IntakeBulkResponse:
    # Capture caller for rate limiting
    http_request.state.caller_id = caller.sub
    response = _compute_evaluate_intake_bulk(payload)
    summary = response.summary
    if http_request is not None:
        _record_audit(
            http_request,
            caller,
            endpoint="/intake/evaluate/bulk",
            payload_summary={"total": summary.total},
            verdict_summary={
                "acceptable": summary.acceptable,
                "denied": summary.denied,
                "overall_status": summary.overall_status,
                "total_risk_score": summary.total_risk_score,
                "max_risk_score": summary.max_risk_score,
                "failed_controls": summary.failed_controls,
            },
        )
        if not _is_synthetic_request(http_request):
            for idx, (intake, verdict) in enumerate(zip(payload.requests, response.results)):
                record_rule_processed(
                    f"{getattr(http_request.state, 'request_id', 'unknown')}:{idx}",
                    "/intake/evaluate/bulk",
                    verdict.verdict,
                )
                _record_decision_history(
                    http_request,
                    caller,
                    endpoint="/intake/evaluate/bulk",
                    action_requested=intake.action,
                    verdict=verdict.verdict,
                    overall_status=verdict.overall_status,
                    overall_risk=verdict.overall_risk,
                    details={
                        "app_id": intake.app_id,
                        "source_name": intake.source_name,
                        "destination_name": intake.destination_name,
                        "protocol": intake.protocol,
                        "destination_port": intake.destination_port,
                        "risk_score": verdict.risk_score,
                        "failed_controls": verdict.failed_controls,
                        "failed_standards": verdict.failed_standards,
                    },
                    policy_input=_build_standards_input(_intake_to_traffic(intake)),
                    reason=verdict.reason,
                    remediations=_collect_remediations(verdict.violations),
                )
    return response


@app.get(
    "/decisions/history",
    summary="Recent accept/decline decision history",
)
def decisions_history(
    limit: int = Query(100, ge=1, le=1000),
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    items = list_recent_decisions(limit=limit)
    return JSONResponse({"total": len(items), "items": items})


@app.get(
    "/policy/metadata",
    summary="Active policy metadata",
)
def policy_metadata(_: CallerIdentity = Depends(require_scope("firewall.evaluate"))) -> JSONResponse:
    return JSONResponse(_policy_metadata())


@app.get(
    "/auth/whoami",
    summary="Current authenticated user details",
)
def auth_whoami(
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    claims = dict(caller.raw_claims) if caller.raw_claims else {}
    return JSONResponse({
        "username": claims.get("username", caller.sub),
        "email": claims.get("email", ""),
        "scopes": sorted(caller.scopes),
        "sub": caller.sub,
    })


@app.get(
    "/compliance/coverage",
    summary="Framework coverage summary",
)
def compliance_coverage(
    framework: str | None = Query(None),
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    return JSONResponse(_framework_coverage_summary(framework=framework))


@app.get(
    "/metrics",
    summary="Prometheus metrics export (ROI tracking)",
    include_in_schema=False,
)
def metrics() -> Response:
    """
    Prometheus-format metrics for rule processing and ROI tracking.
    Accessible without authentication for Prometheus scraping.
    """
    # Update OPA cache metrics before generating output
    update_opa_cache_metrics(_opa_cache.stats())
    return Response(content=get_prometheus_metrics(), media_type="text/plain; version=0.0.4")


@app.get(
    "/metrics/roi",
    summary="Current ROI metrics summary",
)
def metrics_roi(_: CallerIdentity = Depends(require_scope("firewall.evaluate"))) -> JSONResponse:
    """
    Human-readable ROI metrics summary: rules processed, HIPS freed, cost savings.
    """
    return JSONResponse(get_current_metrics())


@app.post(
    "/evaluate/explain",
    summary="Evaluate with detailed explanation",
)
def evaluate_explain(
    request: TrafficRequest,
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    decision = _evaluate(request)
    verdict = _build_verdict(request, decision).model_dump()
    violations = decision.get("violations", []) if isinstance(decision, dict) else []
    explanation = {
        "triggered_violations": [
            {
                "control": item.get("control"),
                "standard": item.get("standard"),
                "severity": item.get("severity"),
                "violation": item.get("violation"),
                "details": item.get("details"),
                "remediation": item.get("remediation"),
            }
            for item in violations
            if isinstance(item, dict)
        ]
    }
    return JSONResponse({"verdict": verdict, "explanation": explanation})


@app.get(
    "/decisions/lifecycle/{decision_id}",
    summary="Get decision lifecycle state",
)
def decision_lifecycle_get(
    decision_id: str,
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    lifecycle = get_decision_lifecycle(decision_id)
    if lifecycle is None:
        raise HTTPException(status_code=404, detail="Decision lifecycle not found")
    return JSONResponse(lifecycle)


@app.put(
    "/decisions/lifecycle/{decision_id}",
    summary="Update decision lifecycle state",
)
def decision_lifecycle_put(
    decision_id: str,
    payload: LifecycleUpdate,
    caller: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    lifecycle = set_decision_lifecycle(
        decision_id=decision_id,
        status=payload.status,
        actor=caller.sub,
        notes=payload.notes,
        expires_at=payload.expires_at,
    )
    _emit_event("decision.lifecycle.updated", lifecycle)
    return JSONResponse(lifecycle)


@app.post(
    "/decisions/drift/recheck",
    summary="Re-check recent decisions for policy drift",
)
def decisions_drift_recheck(
    limit: int = Query(200, ge=1, le=1000),
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    items = list_recent_decisions(limit=limit)
    checked = 0
    drifted: list[dict] = []
    for item in items:
        details = item.get("details", {})
        if not isinstance(details, dict):
            continue
        policy_input = details.get("policy_input")
        if not isinstance(policy_input, dict):
            continue
        checked += 1
        current = _evaluate_opa_input(policy_input)
        current_verdict = "ACCEPTABLE" if current.get("compliant") else "DENY"
        previous_verdict = item.get("decision_verdict")
        if current_verdict != previous_verdict:
            drift = {
                "decision_id": item.get("decision_id"),
                "request_id": item.get("request_id"),
                "endpoint": item.get("endpoint"),
                "previous_verdict": previous_verdict,
                "current_verdict": current_verdict,
                "checked_at": datetime.now(timezone.utc).isoformat(),
            }
            drifted.append(drift)
            _emit_event("decision.drift.detected", drift)
    return JSONResponse({"checked": checked, "drifted": len(drifted), "items": drifted})


@app.get(
    "/compliance/evidence",
    summary="Generate evidence report from decision history",
)
def compliance_evidence(
    days: int = Query(30, ge=1, le=3650),
    format: str = Query("markdown", pattern="^(markdown|json|csv)$"),
    persist: bool = Query(
        True,
        description="When true, stores generated evidence in the archive and index.",
    ),
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
):
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    items = [
        row
        for row in list_recent_decisions(limit=5000)
        if isinstance(row.get("ts"), str)
        and datetime.fromisoformat(row["ts"].replace("Z", "+00:00")) >= cutoff
    ]
    total = len(items)
    denied = sum(1 for row in items if row.get("decision_verdict") == "DENY")
    acceptable = total - denied
    controls: dict[str, int] = {}
    for row in items:
        details = row.get("details", {})
        if not isinstance(details, dict):
            continue
        for control in details.get("failed_controls", []):
            controls[str(control)] = controls.get(str(control), 0) + 1

    report_id_seed = f"{datetime.now(timezone.utc).isoformat()}:{days}:{total}:{denied}:{format}"
    report_id = hashlib.sha256(report_id_seed.encode("utf-8")).hexdigest()[:16]

    if format == "json":
        payload = {
            "report_id": report_id,
            "window_days": days,
            "total": total,
            "acceptable": acceptable,
            "denied": denied,
            "top_failed_controls": dict(sorted(controls.items(), key=lambda kv: -kv[1])[:20]),
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        if persist:
            _archive_evidence_report(
                report_id=report_id,
                days=days,
                total=total,
                acceptable=acceptable,
                denied=denied,
                output_format="json",
                content=json.dumps(payload, separators=(",", ":")),
            )
        return JSONResponse(payload)

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ts", "decision_id", "request_id", "endpoint", "verdict", "overall_risk"])
        for row in items:
            writer.writerow(
                [
                    row.get("ts"),
                    row.get("decision_id"),
                    row.get("request_id"),
                    row.get("endpoint"),
                    row.get("decision_verdict"),
                    row.get("overall_risk"),
                ]
            )
        csv_text = output.getvalue()
        if persist:
            _archive_evidence_report(
                report_id=report_id,
                days=days,
                total=total,
                acceptable=acceptable,
                denied=denied,
                output_format="csv",
                content=csv_text,
            )
        return PlainTextResponse(csv_text, media_type="text/csv", headers={"X-Evidence-Report-Id": report_id})

    lines = [
        "# Compliance Evidence Report",
        "",
        f"- Generated: {datetime.now(timezone.utc).isoformat()}",
        f"- Window: last {days} days",
        f"- Total decisions: {total}",
        f"- Acceptable: {acceptable}",
        f"- Denied: {denied}",
        "",
        "## Top failed controls",
    ]
    for control, count in sorted(controls.items(), key=lambda kv: -kv[1])[:20]:
        lines.append(f"- {control}: {count}")
    markdown_text = "\n".join(lines)
    if persist:
        _archive_evidence_report(
            report_id=report_id,
            days=days,
            total=total,
            acceptable=acceptable,
            denied=denied,
            output_format="markdown",
            content=markdown_text,
        )
    return PlainTextResponse(
        markdown_text,
        media_type="text/markdown",
        headers={"X-Evidence-Report-Id": report_id},
    )


@app.get(
    "/compliance/evidence/archive",
    summary="List archived compliance evidence reports",
)
def compliance_evidence_archive(
    limit: int = Query(50, ge=1, le=500),
    days: int = Query(3650, ge=1, le=3650),
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    _prune_evidence_archive()
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    rows = []
    for row in reversed(_load_evidence_index()):
        created_at = _safe_parse_dt(str(row.get("created_at", "")))
        if created_at and created_at >= cutoff:
            rows.append(row)
        if len(rows) >= limit:
            break
    return JSONResponse({"total": len(rows), "items": rows})


@app.get(
    "/compliance/evidence/archive/{report_id}",
    summary="Fetch archived compliance evidence report content",
)
def compliance_evidence_archive_item(
    report_id: str,
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
):
    _prune_evidence_archive()
    row = _evidence_entry_by_report_id(report_id)
    if row is None:
        raise HTTPException(status_code=404, detail="evidence report not found")

    report_path = EVIDENCE_DIR / str(row.get("file_name", ""))
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="evidence report content missing")

    output_format = str(row.get("format", "")).lower()
    if output_format == "json":
        with report_path.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)
        return JSONResponse(payload)
    if output_format == "csv":
        return PlainTextResponse(report_path.read_text(encoding="utf-8"), media_type="text/csv")
    return PlainTextResponse(report_path.read_text(encoding="utf-8"), media_type="text/markdown")


@app.get(
    "/metrics/slo",
    summary="Operational SLO snapshot",
)
def slo_metrics(
    format: str = Query("json", pattern="^(json|prometheus)$"),
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
):
    snapshot = _slo_snapshot()
    slack_snapshot = _slack_metrics_snapshot()
    snapshot["active_alerts_count"] = _slo_alerts_snapshot()["active_alerts_count"]
    snapshot["slack_dispatch_latency_count"] = slack_snapshot.get("dispatch_latency_count", 0)
    snapshot["slack_dispatch_latency_avg_ms"] = slack_snapshot.get("dispatch_latency_avg_ms", 0.0)
    snapshot["slack_dispatch_latency_p50_ms"] = slack_snapshot.get("dispatch_latency_p50_ms", 0.0)
    snapshot["slack_dispatch_latency_p95_ms"] = slack_snapshot.get("dispatch_latency_p95_ms", 0.0)
    snapshot["slack_dispatch_latency_max_ms"] = slack_snapshot.get("dispatch_latency_max_ms", 0.0)
    snapshot["slack_dispatch_latency_last_ms"] = slack_snapshot.get("dispatch_latency_last_ms", 0.0)
    if format == "prometheus":
        return PlainTextResponse(
            _prometheus_render_slo(snapshot),
            media_type="text/plain; version=0.0.4",
        )
    return JSONResponse(snapshot)


@app.get(
    "/metrics/alerts",
    summary="SLO and delivery alert status",
)
def metrics_alerts(
    format: str = Query("json", pattern="^(json|prometheus)$"),
    _: CallerIdentity = Depends(require_scope("firewall.ops")),
):
    snapshot = _slo_alerts_snapshot()
    if format == "prometheus":
        return PlainTextResponse(
            _prometheus_render_alerts(snapshot),
            media_type="text/plain; version=0.0.4",
        )
    return JSONResponse(snapshot)


@app.get(
    "/notifications/slack/metrics",
    summary="Slack notification delivery metrics",
)
def slack_notification_metrics(_: CallerIdentity = Depends(require_scope("firewall.ops"))) -> JSONResponse:
    return JSONResponse(_slack_metrics_snapshot())


@app.post(
    "/notifications/slack/metrics/reset",
    summary="Reset Slack notification metrics",
)
def reset_slack_notification_metrics(
    clear_dedup_cache: bool = Query(
        True,
        description="When true, clears active Slack dedup cache entries as well as counters.",
    ),
    _: CallerIdentity = Depends(require_scope("firewall.ops")),
) -> JSONResponse:
    return JSONResponse(_reset_slack_metrics(clear_dedup_cache=clear_dedup_cache))


@app.post(
    "/notifications/slack/digest/flush",
    summary="Flush pending low/medium Slack digest",
)
def flush_slack_digest(_: CallerIdentity = Depends(require_scope("firewall.ops"))) -> JSONResponse:
    flushed = _emit_pending_digest_if_due(force=True)
    return JSONResponse({"flushed": flushed, **_slack_metrics_snapshot()})


@app.get(
    "/cache/stats",
    summary="OPA decision cache statistics",
)
def cache_stats(_: CallerIdentity = Depends(require_scope("firewall.evaluate"))) -> JSONResponse:
    return JSONResponse(_opa_cache.stats())


@app.post(
    "/cache/clear",
    summary="Clear the OPA decision cache",
)
def cache_clear(_: CallerIdentity = Depends(require_scope("firewall.evaluate"))) -> JSONResponse:
    _opa_cache.clear()
    return JSONResponse({"cleared": True, **_opa_cache.stats()})


@app.get(
    "/decisions/history/prune/stats",
    summary="Decision history pruning statistics",
)
def decision_history_prune_stats_endpoint(
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    """Return current pruning state and history file size."""
    return JSONResponse(decision_history_prune_stats())


@app.post(
    "/decisions/history/prune",
    summary="Force immediate decision history prune",
)
def decision_history_prune_endpoint(
    _: CallerIdentity = Depends(require_scope("firewall.evaluate")),
) -> JSONResponse:
    """Force an immediate prune of expired decision history entries."""
    return JSONResponse(decision_history_force_prune())


@app.get(
    "/health",
    response_model=None,
    summary="Health check",
    description=(
        "Liveness probe. Returns minimal status by default. "
        "Pass `?verbose=true` for a one-shot triage view including OPA "
        "cache stats, decision-history pruning, SLO snapshot, and Slack "
        "dispatcher metrics."
    ),
)
def health(verbose: bool = False) -> JSONResponse:
    opa_ok  = Path(OPA_BINARY).exists()
    data_ok = DATA_FILE.exists()
    base = {
        "status": "ok" if (opa_ok and data_ok) else "degraded",
        "opa_available": opa_ok,
        "data_file_loaded": data_ok,
    }
    if not verbose:
        return JSONResponse(base)

    # Each sub-block is wrapped so a single subsystem error never kills the
    # whole triage view — verbose health is exactly the endpoint ops will
    # hit when something is already wrong.
    try:
        cache = _opa_cache.stats()
    except Exception:
        cache = {"error": "stats_unavailable"}

    try:
        history = decision_history_prune_stats()
    except Exception:
        history = {"error": "stats_unavailable"}

    try:
        snap = _slo_snapshot()
        slo = {
            "requests_total": snap.get("requests_total"),
            "requests_error": snap.get("requests_error"),
            "error_rate": snap.get("error_rate"),
            "latency_p50_ms": snap.get("latency_p50_ms"),
            "latency_p95_ms": snap.get("latency_p95_ms"),
        }
    except Exception:
        slo = {"error": "snapshot_unavailable"}

    try:
        s = _slack_metrics_snapshot()
        slack = {
            "dispatch_successes": s.get("dispatch_successes", 0),
            "dispatch_failures": s.get("dispatch_failures", 0),
            "last_error": s.get("last_error"),
            "last_error_at": s.get("last_error_at"),
        }
    except Exception:
        slack = {"error": "snapshot_unavailable"}

    # Compose verbose-only warnings list. We don't downgrade `status`
    # (kubelet/load-balancer liveness logic already consumes it); ops can
    # alert on `warnings` independently.
    warnings: list[str] = []
    last_error_at = slack.get("last_error_at") if isinstance(slack, dict) else None
    if isinstance(last_error_at, str):
        try:
            ts = datetime.fromisoformat(last_error_at.replace("Z", "+00:00"))
            age_seconds = (datetime.now(timezone.utc) - ts).total_seconds()
            if 0 <= age_seconds <= _HEALTH_SLACK_RECENT_FAILURE_WINDOW_SECONDS:
                warnings.append("slack.recent_failure")
        except (TypeError, ValueError):
            pass

    payload = {
        **base,
        "opa_cache": cache,
        "decision_history": history,
        "slo": slo,
        "slack": slack,
    }
    if warnings:
        payload["warnings"] = warnings
    return JSONResponse(payload)


@app.get(
    "/rules/summary",
    summary="Policy rule count",
    description="Returns metadata from data.json: rule count, ADOM, policy package.",
)
def rules_summary() -> JSONResponse:
    if not DATA_FILE.exists():
        raise HTTPException(status_code=503, detail="data.json not found")
    with DATA_FILE.open() as f:
        data = json.load(f)
    meta = data.get("_metadata", {})
    return JSONResponse({
        "adom":           meta.get("adom"),
        "policy_package": meta.get("policy_package"),
        "total_rules":    meta.get("total_rules", len(data.get("rules", []))),
        "exported":       meta.get("exported"),
    })


# ── CSV audit ──────────────────────────────────────────────────────────────────
_CSV_BOOL_FIELDS = {"approved_external_sharing", "encryption_required"}
_CSV_INT_FIELDS_RAW = {"port"}
_CSV_INT_FIELDS_INTAKE = {"destination_port"}
_CSV_RAW_SIGNALS = {"source", "destination", "port"}
_CSV_INTAKE_SIGNALS = {"app id", "source name", "destination name"}
_CSV_PALOALTO_REQUIRED_ALIAS_GROUPS = [
    {"source zone", "from zone", "src zone"},
    {"source address", "src address", "source", "src", "from address", "source ip"},
    {"destination zone", "to zone", "dst zone"},
    {"destination address", "dst address", "destination", "dst", "to address", "destination ip"},
    {"application", "app"},
    {"service", "application service", "application service port", "service port"},
    {"action", "rule action", "policy action"},
]


class AuditInvalidRow(BaseModel):
    row: int = Field(..., description="1-based row number in the CSV (header is row 1).")
    errors: list[str] = Field(default_factory=list)
    raw: dict = Field(default_factory=dict, description="Original row values as parsed from the CSV.")


class AuditCsvResponse(BaseModel):
    kind: Literal["raw", "intake"] = Field(..., description="Detected CSV schema family.")
    summary: dict = Field(default_factory=dict, description="Aggregate summary identical to the matching bulk endpoint.")
    results: list[dict] = Field(default_factory=list, description="Per-row verdicts, in CSV order (excluding invalid rows).")
    invalid_rows: list[AuditInvalidRow] = Field(default_factory=list)


def _detect_csv_kind(headers: set[str]) -> str:
    if all(any(alias in headers for alias in group) for group in _CSV_PALOALTO_REQUIRED_ALIAS_GROUPS):
        return "raw_paloalto"

    # Heuristic fallback: many Palo Alto exports vary column names by UI version.
    # If the header set looks policy-like, parse as Palo Alto instead of rejecting.
    paloalto_hints = {
        "name",
        "rule name",
        "policy name",
        "source zone",
        "from zone",
        "destination zone",
        "to zone",
        "source address",
        "src address",
        "destination address",
        "dst address",
        "application",
        "app",
        "service",
        "action",
        "rule action",
        "policy action",
    }
    hint_matches = len(headers & paloalto_hints)
    has_action_hint = bool(headers & {"action", "rule action", "policy action"})
    if hint_matches >= 4 and has_action_hint:
        return "raw_paloalto"

    has_intake = bool(headers & _CSV_INTAKE_SIGNALS)
    has_raw = bool(headers & _CSV_RAW_SIGNALS)
    if has_intake and not has_raw:
        return "intake"
    if has_raw and not has_intake:
        return "raw"
    if has_raw and has_intake:
        raise HTTPException(
            status_code=400,
            detail="CSV header mixes raw and intake columns; pick one schema.",
        )
    raise HTTPException(
        status_code=400,
        detail=(
            "Unrecognised CSV header. Expected raw standards columns, logical intake columns, "
            "or Palo Alto security rule export columns."
        ),
    )


def _paloalto_first_token(value: str, default: str = "any") -> str:
    text = (value or "").strip()
    if not text:
        return default
    for token in text.split(";"):
        cleaned = token.strip()
        if cleaned:
            return cleaned
    return default


def _normalize_csv_header(name: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", " ", (name or "").strip().lower())
    return re.sub(r"\s+", " ", normalized).strip()


def _paloalto_row_get(row: dict[str, str], aliases: set[str], default: str = "") -> str:
    for alias in aliases:
        value = row.get(alias)
        if value is None:
            continue
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                return cleaned
        else:
            return str(value)
    return default


def _paloalto_parse_log(options: str) -> str:
    lowered = (options or "").lower()
    if "no log" in lowered or "disable" in lowered:
        return "no_log"
    return "log_all_sessions"


def _paloalto_parse_action(action: str) -> str:
    lowered = (action or "").strip().lower()
    if lowered in {"allow", "accept", "permit"}:
        return "accept"
    return "deny"


def _paloalto_parse_protocol_port(service: str, application: str) -> tuple[str, int]:
    service_token = _paloalto_first_token(service, default="application-default").lower()
    app_token = _paloalto_first_token(application, default="any").lower()

    # app-default style services rely on application context where possible.
    if service_token in {"application-default", "app-default", "application_default", "any"}:
        if app_token in {"ssl", "service-https", "https"}:
            return "tcp", 443
        if app_token in {"web-browsing", "service-http", "http"}:
            return "tcp", 80
        if app_token == "dns":
            return "udp", 53
        return "any", 0

    if service_token in {"service-http", "http"}:
        return "tcp", 80
    if service_token in {"service-https", "https"}:
        return "tcp", 443

    match = re.search(r"(tcp|udp)[-_ ]?(\d{1,5})", service_token)
    if match:
        port = int(match.group(2))
        if 0 <= port <= 65535:
            return match.group(1), port

    number_match = re.search(r"\b(\d{1,5})\b", service_token)
    if number_match:
        port = int(number_match.group(1))
        if 0 <= port <= 65535:
            return "tcp", port

    return "any", 0


def _coerce_paloalto_csv_row(row: dict) -> dict:
    normalized_row: dict[str, str] = {}
    for key, value in row.items():
        if key is None:
            continue
        norm_key = _normalize_csv_header(str(key))
        if not norm_key:
            continue
        normalized_row[norm_key] = "" if value is None else str(value)

    rule_name = _paloalto_row_get(normalized_row, {"name", "rule name", "policy name"}, default="")
    source_interface = _paloalto_first_token(
        _paloalto_row_get(normalized_row, {"source zone", "from zone", "src zone"}, default=""),
        default="any",
    )
    destination_interface = _paloalto_first_token(
        _paloalto_row_get(normalized_row, {"destination zone", "to zone", "dst zone"}, default=""),
        default="any",
    )
    source = _paloalto_first_token(
        _paloalto_row_get(
            normalized_row,
            {"source address", "src address", "source", "src", "from address", "source ip"},
            default="",
        ),
        default="any",
    )
    destination = _paloalto_first_token(
        _paloalto_row_get(
            normalized_row,
            {"destination address", "dst address", "destination", "dst", "to address", "destination ip"},
            default="",
        ),
        default="any",
    )
    protocol, port = _paloalto_parse_protocol_port(
        _paloalto_row_get(
            normalized_row,
            {"service", "application service", "application service port", "service port"},
            default="",
        ),
        _paloalto_row_get(normalized_row, {"application", "app"}, default=""),
    )

    return {
        "source": source,
        "destination": destination,
        "protocol": protocol,
        "port": port,
        "log": _paloalto_parse_log(
            _paloalto_row_get(
                normalized_row,
                {"options", "log", "logging", "traffic log", "log forwarding profile"},
                default="",
            )
        ),
        "action": _paloalto_parse_action(
            _paloalto_row_get(normalized_row, {"action", "rule action", "policy action"}, default="")
        ),
        "source_interface": source_interface,
        "destination_interface": destination_interface,
        "rule_name": rule_name,
    }


def _safe_json_get(obj: dict, path: list[str], default: str = "") -> str:
    """Navigate nested dict path, return string or default."""
    current = obj
    for key in path:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
    if current is None:
        return default
    return str(current).strip() or default


def _safe_json_get_first(items: list, default: str = "any") -> str:
    """Get first item's 'data' field from list, return string or default."""
    if not isinstance(items, list) or not items:
        return default
    first = items[0]
    if isinstance(first, dict):
        data = first.get("data")
        if data is not None:
            return str(data).strip() or default
    return default


def _parse_juniper_srx_json(text: str) -> tuple[list[TrafficRequest], list[AuditInvalidRow]]:
    """Extract policies from Juniper SRX JSON export to internal schema."""
    valid: list[TrafficRequest] = []
    invalid: list[AuditInvalidRow] = []

    try:
        obj = json.loads(text)
    except json.JSONDecodeError as e:
        invalid.append(AuditInvalidRow(row=1, errors=[f"Invalid JSON: {e}"], raw={}))
        return valid, invalid

    if not isinstance(obj, dict):
        invalid.append(AuditInvalidRow(row=1, errors=["Root is not a dict"], raw={}))
        return valid, invalid

    policies_list = obj.get("policies", [])
    if not policies_list:
        return valid, invalid

    row_idx = 1
    for policy_group in policies_list:
        if not isinstance(policy_group, dict):
            continue

        for zone_pair in policy_group.get("policy", []):
            if not isinstance(zone_pair, dict):
                continue

            from_zone = _safe_json_get(zone_pair, ["from-zone-name", "data"], "any")
            to_zone = _safe_json_get(zone_pair, ["to-zone-name", "data"], "any")
            rules_list = zone_pair.get("policy", [])

            if not isinstance(rules_list, list):
                continue

            for rule in rules_list:
                row_idx += 1
                if not isinstance(rule, dict):
                    continue

                try:
                    name = _safe_json_get(rule, ["name", "data"], "")
                    match_items = rule.get("match", [])
                    if not isinstance(match_items, list) or not match_items:
                        continue

                    match_obj = match_items[0]
                    if not isinstance(match_obj, dict):
                        continue

                    source = _safe_json_get_first(match_obj.get("source-address", []), "any")
                    destination = _safe_json_get_first(match_obj.get("destination-address", []), "any")

                    then_items = rule.get("then", [])
                    action = "deny"
                    if then_items and isinstance(then_items[0], dict):
                        if "permit" in then_items[0]:
                            action = "accept"

                    req = TrafficRequest(
                        source=source,
                        destination=destination,
                        protocol="any",
                        port=0,
                        log="log_all_sessions",
                        action=action,
                        source_interface=from_zone,
                        destination_interface=to_zone,
                        rule_name=name,
                    )
                    valid.append(req)
                except Exception as e:
                    invalid.append(AuditInvalidRow(row=row_idx, errors=[str(e)], raw=rule))

    return valid, invalid


def _xml_local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[-1]
    return tag


def _xml_child_text(parent: ET.Element, name: str, default: str = "") -> str:
    for child in list(parent):
        if _xml_local_name(child.tag) == name:
            text = (child.text or "").strip()
            return text or default
    return default


def _parse_juniper_srx_xml(text: str) -> tuple[list[TrafficRequest], list[AuditInvalidRow]]:
    """Extract policies from Juniper SRX XML export to internal schema."""
    valid: list[TrafficRequest] = []
    invalid: list[AuditInvalidRow] = []

    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        invalid.append(AuditInvalidRow(row=1, errors=[f"Invalid XML: {e}"], raw={}))
        return valid, invalid

    policies_nodes: list[ET.Element]
    if _xml_local_name(root.tag) == "policies":
        policies_nodes = [root]
    else:
        policies_nodes = [n for n in root.iter() if _xml_local_name(n.tag) == "policies"]

    row_idx = 1
    for policies_node in policies_nodes:
        zone_pairs = [n for n in list(policies_node) if _xml_local_name(n.tag) == "policy"]
        for zone_pair in zone_pairs:
            from_zone = _xml_child_text(zone_pair, "from-zone-name", "any")
            to_zone = _xml_child_text(zone_pair, "to-zone-name", "any")

            rules = []
            for maybe_rule in list(zone_pair):
                if _xml_local_name(maybe_rule.tag) != "policy":
                    continue
                if any(_xml_local_name(c.tag) == "match" for c in list(maybe_rule)):
                    rules.append(maybe_rule)

            for rule in rules:
                row_idx += 1
                try:
                    name = _xml_child_text(rule, "name", "")

                    match_node = next((c for c in list(rule) if _xml_local_name(c.tag) == "match"), None)
                    source = _xml_child_text(match_node, "source-address", "any") if match_node is not None else "any"
                    destination = _xml_child_text(match_node, "destination-address", "any") if match_node is not None else "any"

                    then_node = next((c for c in list(rule) if _xml_local_name(c.tag) == "then"), None)
                    action = "deny"
                    if then_node is not None:
                        then_children = {_xml_local_name(c.tag) for c in list(then_node)}
                        if "permit" in then_children:
                            action = "accept"

                    req = TrafficRequest(
                        source=source,
                        destination=destination,
                        protocol="any",
                        port=0,
                        log="log_all_sessions",
                        action=action,
                        source_interface=from_zone,
                        destination_interface=to_zone,
                        rule_name=name,
                    )
                    valid.append(req)
                except Exception as e:
                    invalid.append(AuditInvalidRow(row=row_idx, errors=[str(e)], raw={"rule": "xml"}))

    return valid, invalid


def _coerce_csv_row(row: dict, kind: str) -> dict:
    int_fields = _CSV_INT_FIELDS_RAW if kind == "raw" else _CSV_INT_FIELDS_INTAKE
    out: dict = {}
    for raw_key, raw_val in row.items():
        if raw_key is None:
            continue
        key = raw_key.strip()
        if not key:
            continue
        val = (raw_val or "").strip()
        if val == "":
            continue
        if key in _CSV_BOOL_FIELDS:
            lowered = val.lower()
            if lowered in {"true", "1", "yes", "y"}:
                out[key] = True
            elif lowered in {"false", "0", "no", "n"}:
                out[key] = False
            else:
                raise ValueError(f"{key} must be boolean, got {val!r}")
        elif key in int_fields:
            out[key] = int(val)
        else:
            out[key] = val
    return out


def _xlsx_cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _xlsx_first_line(value: object, default: str = "") -> str:
    text = _xlsx_cell_text(value)
    if not text:
        return default
    return text.splitlines()[0].strip() or default


def _xlsx_parse_service(service_name: str, service_value: str) -> tuple[str, int]:
    name = (service_name or "").strip()
    value = (service_value or "").strip()

    if name.upper() == "ALL":
        return "any", 0

    if value.startswith("IP/"):
        proto_num = value.split("/", 1)[1].strip() if "/" in value else "0"
        proto_map = {
            "0": "any",
            "1": "icmp",
            "6": "tcp",
            "17": "udp",
        }
        return proto_map.get(proto_num, "tcp"), 0

    lower_name = name.lower()
    known_ports = {
        "http": 80,
        "https": 443,
        "ssh": 22,
        "dns": 53,
        "ntp": 123,
        "smtp": 25,
        "pop3": 110,
        "imap": 143,
        "snmp": 161,
        "rdp": 3389,
    }
    if lower_name in known_ports:
        return "tcp", known_ports[lower_name]

    port_match = re.search(r"\b(\d{1,5})\b", name)
    if port_match:
        port = int(port_match.group(1))
        if 0 <= port <= 65535:
            return "tcp", port

    return "tcp", 0


def _load_xlsx_worksheet(body_bytes: bytes):
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=503,
            detail="XLSX parsing is unavailable. Install dependency 'openpyxl'.",
        ) from exc

    try:
        workbook = load_workbook(filename=io.BytesIO(body_bytes), read_only=True, data_only=True)
        return workbook.worksheets[0]
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Unable to parse XLSX: {exc}") from exc


def _parse_raw_xlsx_sheet(sheet) -> tuple[list[TrafficRequest], list[AuditInvalidRow], list[dict]]:
    valid: list[TrafficRequest] = []
    invalid: list[AuditInvalidRow] = []
    normalized_rows: list[dict] = []

    header_values = [
        _xlsx_cell_text(cell).lower()
        for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    ]
    headers = [h.strip() for h in header_values]
    if not headers or not (set(headers) & _CSV_RAW_SIGNALS):
        return valid, invalid, normalized_rows

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_dict: dict[str, str] = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            val = _xlsx_cell_text(row[idx] if idx < len(row) else "")
            if val:
                row_dict[key] = val

        if not row_dict:
            continue

        try:
            cleaned = _coerce_csv_row(row_dict, "raw")
            model = TrafficRequest(**cleaned)
            valid.append(model)
            normalized_rows.append(model.model_dump(exclude_none=True))
        except Exception as exc:  # noqa: BLE001
            invalid.append(AuditInvalidRow(row=row_idx, errors=[str(exc)], raw=row_dict))

    return valid, invalid, normalized_rows


def _parse_fortinet_xlsx_sheet(sheet) -> tuple[list[TrafficRequest], list[AuditInvalidRow], list[dict]]:
    valid: list[TrafficRequest] = []
    invalid: list[AuditInvalidRow] = []
    normalized_rows: list[dict] = []

    header_seq = _xlsx_cell_text(sheet.cell(row=2, column=1).value).lower()
    if header_seq != "seq #":
        return valid, invalid, normalized_rows

    for row_idx in range(6, sheet.max_row + 1):
        seq_val = sheet.cell(row=row_idx, column=1).value
        if seq_val is None:
            continue
        try:
            int(seq_val)
        except Exception:
            continue

        rule_name = _xlsx_cell_text(sheet.cell(row=row_idx, column=3).value) or f"Rule-{seq_val}"
        action = _xlsx_cell_text(sheet.cell(row=row_idx, column=4).value).lower() or "deny"
        if action not in {"accept", "deny"}:
            action = "deny"

        source_interface = _xlsx_first_line(sheet.cell(row=row_idx, column=5).value, default="any")
        source = _xlsx_first_line(sheet.cell(row=row_idx, column=7).value, default="0.0.0.0/0")

        destination_interface = _xlsx_first_line(sheet.cell(row=row_idx, column=8).value, default="any")
        destination = _xlsx_first_line(sheet.cell(row=row_idx, column=10).value, default="0.0.0.0/0")

        service_name = _xlsx_cell_text(sheet.cell(row=row_idx, column=13).value)
        service_value = _xlsx_cell_text(sheet.cell(row=row_idx, column=14).value)
        protocol, port = _xlsx_parse_service(service_name, service_value)

        log_value_raw = _xlsx_cell_text(sheet.cell(row=row_idx, column=16).value).lower()
        if "no log" in log_value_raw or "disable" in log_value_raw:
            log_value = "no_log"
        else:
            log_value = "log_all_sessions"

        raw_payload = {
            "source": source,
            "destination": destination,
            "protocol": protocol,
            "port": port,
            "log": log_value,
            "action": action,
            "source_interface": source_interface,
            "destination_interface": destination_interface,
            "rule_name": rule_name,
        }

        try:
            model = TrafficRequest(**raw_payload)
            valid.append(model)
            normalized_rows.append(model.model_dump(exclude_none=True))
        except Exception as exc:  # noqa: BLE001
            invalid.append(AuditInvalidRow(row=row_idx, errors=[str(exc)], raw=raw_payload))

    return valid, invalid, normalized_rows


def _apply_standards_to_requests(requests: list[TrafficRequest], standards: list[str] | None) -> None:
    """
    Apply selected standards to all traffic requests in place.
    If standards is None or empty, uses default ["ISO 27001", "CIS v8.1", "PCI-DSS"].
    """
    if not standards:
        standards = ["ISO 27001", "CIS v8.1", "PCI-DSS"]
    for req in requests:
        req.standards = standards


def _parse_xlsx_to_raw_requests(body_bytes: bytes) -> tuple[list[TrafficRequest], list[AuditInvalidRow], list[dict]]:
    sheet = _load_xlsx_worksheet(body_bytes)

    valid_raw, invalid, normalized_rows = _parse_raw_xlsx_sheet(sheet)
    if valid_raw or invalid:
        return valid_raw, invalid, normalized_rows

    return _parse_fortinet_xlsx_sheet(sheet)


def _render_cleaned_raw_csv(rows: list[dict]) -> str:
    if not rows:
        return ""

    def _has_value(value: object) -> bool:
        return value is not None and value != "" and value != []

    ordered_headers = [
        "source",
        "destination",
        "protocol",
        "port",
        "log",
        "action",
        "source_interface",
        "destination_interface",
        "rule_name",
        "data_classification",
        "approved_external_sharing",
        "contract_reference",
        "encryption_required",
        "tls_version_minimum",
        "source_name",
        "destination_name",
        "app_id",
        "environment",
        "business_owner",
        "ticket_id",
    ]
    active_headers = [h for h in ordered_headers if any(_has_value(r.get(h)) for r in rows)]
    extras = sorted(
        {k for row in rows for k, v in row.items() if k not in active_headers and _has_value(v)}
    )
    fieldnames = active_headers + extras

    serializable_rows: list[dict] = []
    for row in rows:
        cleaned_row: dict = {}
        for key in fieldnames:
            value = row.get(key)
            if isinstance(value, (list, dict)):
                cleaned_row[key] = json.dumps(value, ensure_ascii=False)
            else:
                cleaned_row[key] = value
        serializable_rows.append(cleaned_row)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(serializable_rows)
    return buf.getvalue()


def _evaluate_csv_text(text: str, standards: list[str] | None = None) -> tuple[str, dict, list[dict], list[AuditInvalidRow], list[dict]]:
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV is missing a header row.")

    headers = {_normalize_csv_header(h) for h in reader.fieldnames if h is not None}
    kind = _detect_csv_kind(headers)

    valid_raw: list[TrafficRequest] = []
    valid_intake: list[IntakeRequest] = []
    invalid: list[AuditInvalidRow] = []
    normalized_rows: list[dict] = []

    for idx, row in enumerate(reader, start=2):  # header is row 1
        try:
            cleaned = _coerce_paloalto_csv_row(row) if kind == "raw_paloalto" else _coerce_csv_row(row, kind)
            if kind in {"raw", "raw_paloalto"}:
                model = TrafficRequest(**cleaned)
                valid_raw.append(model)
                normalized_rows.append(model.model_dump(exclude_none=True))
            else:
                model = IntakeRequest(**cleaned)
                valid_intake.append(model)
                normalized_rows.append(model.model_dump(exclude_none=True))
        except Exception as exc:  # noqa: BLE001 — surface validation/coercion errors per row
            invalid.append(AuditInvalidRow(row=idx, errors=[str(exc)], raw=row))

    if not valid_raw and not valid_intake and not invalid:
        raise HTTPException(status_code=400, detail="CSV contained no data rows.")

    # Apply standards to traffic requests if provided
    if valid_raw and standards:
        _apply_standards_to_requests(valid_raw, standards)

    summary: dict = {}
    results: list[dict] = []
    if kind in {"raw", "raw_paloalto"} and valid_raw:
        bulk = _compute_evaluate_bulk(BulkRequest(requests=valid_raw))
        summary = bulk.summary.model_dump()
        results = [r.model_dump() for r in bulk.results]
    elif kind == "intake" and valid_intake:
        bulk = _compute_evaluate_intake_bulk(IntakeBulkRequest(requests=valid_intake))
        summary = bulk.summary.model_dump()
        results = [r.model_dump() for r in bulk.results]

    return kind, summary, results, invalid, normalized_rows


@app.post(
    "/audit/csv",
    summary="Audit deployed rules from a CSV export",
    response_class=Response,
    responses={
        200: {
            "description": "Markdown compliance report.",
            "content": {"text/markdown": {"schema": {"type": "string"}}},
        }
    },
    description=(
        "Submit an existing ruleset as `text/csv` in the request body. The endpoint auto-detects "
        "whether the CSV uses the raw standards columns (source, destination, protocol, port, …) "
        "or the logical intake columns (app_id, source_name, destination_name, …), evaluates every "
        "row against the security-standards policy, and returns a Markdown compliance report "
        "(`text/markdown`) suitable for saving as a `.md` file. Malformed rows are listed in the "
        "report but do not abort the audit."
    ),
)
async def audit_csv(
    request: Request,
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    # Capture caller for rate limiting
    request.state.caller_id = caller.sub
    content_type = request.headers.get("content-type", "")
    if content_type and "text/csv" not in content_type and "text/plain" not in content_type:
        raise HTTPException(
            status_code=415,
            detail="Content-Type must be text/csv (or text/plain).",
        )

    body_bytes = await request.body()
    if not body_bytes.strip():
        raise HTTPException(status_code=400, detail="Empty CSV body.")

    text = body_bytes.decode("utf-8-sig", errors="replace")
    kind, summary, results, invalid, _normalized_rows = _evaluate_csv_text(text)

    report = _render_markdown_report(kind, summary, results, invalid)
    _record_audit(
        request,
        caller,
        endpoint="/audit/csv",
        payload_summary={
            "kind": kind,
            "valid_rows": len(results),
            "invalid_rows": len(invalid),
        },
        verdict_summary={
            "overall_status": summary.get("overall_status", "UNKNOWN"),
            "acceptable": summary.get("acceptable", 0),
            "denied": summary.get("denied", 0),
            "failed_controls": summary.get("failed_controls", []),
        },
    )
    return Response(
        content=report,
        media_type="text/markdown; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="compliance-report.md"',
        },
    )


@app.post(
    "/audit/csv/html",
    summary="Audit deployed rules from a CSV upload and return HTML report",
    response_class=Response,
    responses={
        200: {
            "description": "HTML compliance report.",
            "content": {"text/html": {"schema": {"type": "string"}}},
        }
    },
    description=(
        "Upload a `.csv` firewall rules export as multipart form data. The endpoint auto-detects "
        "raw standards schema vs intake schema, evaluates all rows, and returns a downloadable "
        "HTML compliance report."
    ),
)
async def audit_csv_html(
    request: Request,
    file: UploadFile = File(..., description="Firewall rules CSV file (.csv)"),
    standards: list[str] = Query(None, description="Compliance standards to evaluate (ISO 27001, CIS v8.1, PCI-DSS)"),
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    request.state.caller_id = caller.sub

    filename = file.filename or "firewall-policy.csv"
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=415, detail="Only .csv files are supported for this endpoint.")

    body_bytes = await file.read()
    if not body_bytes.strip():
        raise HTTPException(status_code=400, detail="Uploaded CSV file is empty.")

    text = body_bytes.decode("utf-8-sig", errors="replace")
    kind, summary, results, invalid, _normalized_rows = _evaluate_csv_text(text, standards=standards if standards else None)

    report_markdown = _render_markdown_report(kind, summary, results, invalid)
    report_html = _markdown_to_html_document(report_markdown)
    base_name = Path(filename).stem

    _record_audit(
        request,
        caller,
        endpoint="/audit/csv/html",
        payload_summary={
            "kind": kind,
            "valid_rows": len(results),
            "invalid_rows": len(invalid),
            "source_filename": filename,
        },
        verdict_summary={
            "overall_status": summary.get("overall_status", "UNKNOWN"),
            "acceptable": summary.get("acceptable", 0),
            "denied": summary.get("denied", 0),
            "failed_controls": summary.get("failed_controls", []),
        },
    )

    return Response(
        content=report_html,
        media_type="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{base_name}-compliance-report.html"',
            "X-Audit-Valid-Rows": str(len(results)),
            "X-Audit-Invalid-Rows": str(len(invalid)),
            "X-Audit-Acceptable": str(summary.get("acceptable", 0)),
            "X-Audit-Denied": str(summary.get("denied", 0)),
        },
    )


@app.post(
    "/audit/csv/cleaned",
    summary="Extract normalized rules from a CSV upload",
    response_class=Response,
    responses={
        200: {
            "description": "Cleaned normalized artifact.",
            "content": {
                "text/csv": {"schema": {"type": "string"}},
                "application/json": {"schema": {"type": "object"}},
            },
        }
    },
)
async def audit_csv_cleaned(
    request: Request,
    file: UploadFile = File(..., description="Firewall rules CSV file (.csv)"),
    format: Literal["csv", "json"] = Query(
        default="csv",
        description="Artifact format to return: csv or json.",
    ),
    standards: list[str] = Query(None, description="Compliance standards to evaluate (ISO 27001, CIS v8.1, PCI-DSS)"),
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    request.state.caller_id = caller.sub

    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=415, detail="Uploaded file must be .csv")

    body_bytes = await file.read()
    if not body_bytes:
        raise HTTPException(status_code=400, detail="Empty CSV upload.")
    if len(body_bytes) > MAX_REQUEST_BODY_BYTES:
        raise HTTPException(status_code=413, detail="Uploaded CSV exceeds configured size limit.")

    text = body_bytes.decode("utf-8-sig", errors="replace")
    kind, _summary, _results, invalid, normalized_rows = _evaluate_csv_text(text, standards=standards if standards else None)

    report_name = Path(filename).stem.replace(" ", "-")[:60] or "firewall-rules"
    _record_audit(
        request,
        caller,
        endpoint="/audit/csv/cleaned",
        payload_summary={
            "kind": f"{kind}-csv-cleaned",
            "filename": filename,
            "valid_rows": len(normalized_rows),
            "invalid_rows": len(invalid),
            "format": format,
        },
        verdict_summary={
            "overall_status": "NORMALIZED",
            "acceptable": 0,
            "denied": 0,
            "failed_controls": [],
        },
    )

    if format == "json":
        payload = {
            "source_filename": filename,
            "kind": kind,
            "valid_rows": len(normalized_rows),
            "invalid_rows": [i.model_dump() for i in invalid],
            "rows": normalized_rows,
        }
        return Response(
            content=json.dumps(payload, ensure_ascii=False, indent=2),
            media_type="application/json; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{report_name}-cleaned.json"',
                "X-Audit-Valid-Rows": str(len(normalized_rows)),
                "X-Audit-Invalid-Rows": str(len(invalid)),
                "X-Audit-Cleaned-Format": "json",
            },
        )

    csv_payload = _render_cleaned_raw_csv(normalized_rows)
    return Response(
        content=csv_payload,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{report_name}-cleaned.csv"',
            "X-Audit-Valid-Rows": str(len(normalized_rows)),
            "X-Audit-Invalid-Rows": str(len(invalid)),
            "X-Audit-Cleaned-Format": "csv",
        },
    )


@app.post(
    "/audit/xlsx",
    summary="Audit deployed rules from an XLSX export",
    response_class=Response,
    responses={
        200: {
            "description": "HTML compliance report.",
            "content": {"text/html": {"schema": {"type": "string"}}},
        }
    },
    description=(
        "Upload an `.xlsx` workbook (first worksheet is used). The endpoint strips "
        "non-essential export columns, maps rows to the raw standards schema, runs compliance "
        "evaluation, and returns a downloadable HTML report."
    ),
)
async def audit_xlsx(
    request: Request,
    file: UploadFile = File(..., description="Firewall rules workbook (.xlsx)"),
    standards: list[str] = Query(None, description="Compliance standards to evaluate (ISO 27001, CIS v8.1, PCI-DSS)"),
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    request.state.caller_id = caller.sub

    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Uploaded file must be .xlsx")

    body_bytes = await file.read()
    if not body_bytes:
        raise HTTPException(status_code=400, detail="Empty XLSX upload.")
    if len(body_bytes) > MAX_REQUEST_BODY_BYTES:
        raise HTTPException(status_code=413, detail="Uploaded XLSX exceeds configured size limit.")

    valid_raw, invalid, _normalized_rows = _parse_xlsx_to_raw_requests(body_bytes)
    if not valid_raw and not invalid:
        raise HTTPException(
            status_code=400,
            detail=(
                "Could not detect supported worksheet format. Provide either a raw-schema sheet "
                "(source, destination, protocol, port, ...) or a Fortinet policy export worksheet."
            ),
        )

    # Apply standards to traffic requests if provided
    if valid_raw and standards:
        _apply_standards_to_requests(valid_raw, standards)

    summary: dict = {}
    results: list[dict] = []
    if valid_raw:
        bulk = _compute_evaluate_bulk(BulkRequest(requests=valid_raw))
        summary = bulk.summary.model_dump()
        results = [r.model_dump() for r in bulk.results]

    report_markdown = _render_markdown_report("raw", summary, results, invalid)
    report_html = _markdown_to_html_document(report_markdown)
    _record_audit(
        request,
        caller,
        endpoint="/audit/xlsx",
        payload_summary={
            "kind": "raw-xlsx",
            "filename": filename,
            "valid_rows": len(results),
            "invalid_rows": len(invalid),
        },
        verdict_summary={
            "overall_status": summary.get("overall_status", "UNKNOWN"),
            "acceptable": summary.get("acceptable", 0),
            "denied": summary.get("denied", 0),
            "failed_controls": summary.get("failed_controls", []),
        },
    )

    report_name = Path(filename).stem.replace(" ", "-")[:60] or "compliance-report"
    return Response(
        content=report_html,
        media_type="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{report_name}-compliance-report.html"',
            "X-Audit-Valid-Rows": str(len(results)),
            "X-Audit-Invalid-Rows": str(len(invalid)),
            "X-Audit-Acceptable": str(summary.get("acceptable", 0)),
            "X-Audit-Denied": str(summary.get("denied", 0)),
        },
    )


@app.post(
    "/audit/xlsx/cleaned",
    summary="Extract normalized rules from an XLSX export",
    response_class=Response,
    responses={
        200: {
            "description": "Cleaned normalized artifact.",
            "content": {
                "text/csv": {"schema": {"type": "string"}},
                "application/json": {"schema": {"type": "object"}},
            },
        }
    },
)
async def audit_xlsx_cleaned(
    request: Request,
    file: UploadFile = File(..., description="Firewall rules workbook (.xlsx)"),
    format: Literal["csv", "json"] = Query(
        default="csv",
        description="Artifact format to return: csv or json.",
    ),
    standards: list[str] = Query(None, description="Compliance standards to evaluate (ISO 27001, CIS v8.1, PCI-DSS)"),
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    request.state.caller_id = caller.sub

    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Uploaded file must be .xlsx")

    body_bytes = await file.read()
    if not body_bytes:
        raise HTTPException(status_code=400, detail="Empty XLSX upload.")
    if len(body_bytes) > MAX_REQUEST_BODY_BYTES:
        raise HTTPException(status_code=413, detail="Uploaded XLSX exceeds configured size limit.")

    valid_raw, invalid, normalized_rows = _parse_xlsx_to_raw_requests(body_bytes)
    if not valid_raw and not invalid:
        raise HTTPException(
            status_code=400,
            detail=(
                "Could not detect supported worksheet format. Provide either a raw-schema sheet "
                "(source, destination, protocol, port, ...) or a Fortinet policy export worksheet."
            ),
        )

    report_name = Path(filename).stem.replace(" ", "-")[:60] or "firewall-rules"
    _record_audit(
        request,
        caller,
        endpoint="/audit/xlsx/cleaned",
        payload_summary={
            "kind": "raw-xlsx-cleaned",
            "filename": filename,
            "valid_rows": len(normalized_rows),
            "invalid_rows": len(invalid),
            "format": format,
        },
        verdict_summary={
            "overall_status": "NORMALIZED",
            "acceptable": 0,
            "denied": 0,
            "failed_controls": [],
        },
    )

    if format == "json":
        payload = {
            "source_filename": filename,
            "valid_rows": len(normalized_rows),
            "invalid_rows": [i.model_dump() for i in invalid],
            "rows": normalized_rows,
        }
        return Response(
            content=json.dumps(payload, ensure_ascii=False, indent=2),
            media_type="application/json; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{report_name}-cleaned.json"',
                "X-Audit-Valid-Rows": str(len(normalized_rows)),
                "X-Audit-Invalid-Rows": str(len(invalid)),
                "X-Audit-Cleaned-Format": "json",
            },
        )

    csv_payload = _render_cleaned_raw_csv(normalized_rows)
    return Response(
        content=csv_payload,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{report_name}-cleaned.csv"',
            "X-Audit-Valid-Rows": str(len(normalized_rows)),
            "X-Audit-Invalid-Rows": str(len(invalid)),
            "X-Audit-Cleaned-Format": "csv",
        },
    )


@app.post(
    "/audit/json/html",
    summary="Audit Juniper SRX JSON export and return HTML report",
    response_class=Response,
)
async def audit_json_html(
    request: Request,
    file: UploadFile = File(..., description="Juniper SRX policy JSON or XML export"),
    standards: list[str] = Query(None, description="Compliance standards to evaluate (ISO 27001, CIS v8.1, PCI-DSS)"),
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    request.state.caller_id = caller.sub

    filename = (file.filename or "").strip()
    lower_name = filename.lower()
    if not (lower_name.endswith(".json") or lower_name.endswith(".xml")):
        raise HTTPException(status_code=415, detail="Uploaded file must be .json or .xml")

    body_bytes = await file.read()
    if not body_bytes:
        raise HTTPException(status_code=400, detail="Empty Juniper upload.")
    if len(body_bytes) > MAX_REQUEST_BODY_BYTES:
        raise HTTPException(status_code=413, detail="Uploaded file exceeds configured size limit.")

    try:
        text = body_bytes.decode("utf-8")
    except UnicodeDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Upload must be UTF-8 encoded: {e}")

    if lower_name.endswith(".json"):
        valid_raw, invalid = _parse_juniper_srx_json(text)
        source_kind = "juniper-srx-json"
    else:
        valid_raw, invalid = _parse_juniper_srx_xml(text)
        source_kind = "juniper-srx-xml"
    if not valid_raw and not invalid:
        raise HTTPException(
            status_code=400,
            detail="Could not extract policies from Juniper SRX JSON/XML export.",
        )

    # Apply standards to traffic requests if provided
    if valid_raw and standards:
        _apply_standards_to_requests(valid_raw, standards)

    summary: dict = {}
    results: list[dict] = []
    if valid_raw:
        bulk = _compute_evaluate_bulk(BulkRequest(requests=valid_raw))
        summary = bulk.summary.model_dump()
        results = [r.model_dump() for r in bulk.results]

    report_markdown = _render_markdown_report("raw", summary, results, invalid)
    report_html = _markdown_to_html_document(report_markdown)
    _record_audit(
        request,
        caller,
        endpoint="/audit/json/html",
        payload_summary={
            "kind": source_kind,
            "filename": filename,
            "valid_rows": len(results),
            "invalid_rows": len(invalid),
        },
        verdict_summary={
            "overall_status": summary.get("overall_status", "UNKNOWN"),
            "acceptable": summary.get("acceptable", 0),
            "denied": summary.get("denied", 0),
            "failed_controls": summary.get("failed_controls", []),
        },
    )

    report_name = Path(filename).stem.replace(" ", "-")[:60] or "juniper-compliance-report"
    return Response(
        content=report_html,
        media_type="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{report_name}-compliance-report.html"',
            "X-Audit-Valid-Rows": str(len(results)),
            "X-Audit-Invalid-Rows": str(len(invalid)),
            "X-Audit-Acceptable": str(summary.get("acceptable", 0)),
            "X-Audit-Denied": str(summary.get("denied", 0)),
        },
    )


@app.post(
    "/audit/json/cleaned",
    summary="Extract normalized rules from Juniper SRX JSON export",
    response_class=Response,
)
async def audit_json_cleaned(
    request: Request,
    file: UploadFile = File(..., description="Juniper SRX policy JSON or XML export"),
    format: str = Query("csv", description="Output format: csv or json"),
    fmt: str | None = Query(None, description="Alias for format"),
    standards: list[str] = Query(None, description="Compliance standards to evaluate (ISO 27001, CIS v8.1, PCI-DSS)"),
    caller: CallerIdentity = Depends(require_scope("firewall.audit")),
) -> Response:
    request.state.caller_id = caller.sub

    filename = (file.filename or "").strip()
    lower_name = filename.lower()
    if not (lower_name.endswith(".json") or lower_name.endswith(".xml")):
        raise HTTPException(status_code=415, detail="Uploaded file must be .json or .xml")

    selected_fmt = (fmt or format or "csv").lower()

    if selected_fmt not in ("csv", "json"):
        raise HTTPException(status_code=400, detail='Format must be "csv" or "json"')

    body_bytes = await file.read()
    if not body_bytes:
        raise HTTPException(status_code=400, detail="Empty Juniper upload.")
    if len(body_bytes) > MAX_REQUEST_BODY_BYTES:
        raise HTTPException(status_code=413, detail="Uploaded file exceeds configured size limit.")

    try:
        text = body_bytes.decode("utf-8")
    except UnicodeDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Upload must be UTF-8 encoded: {e}")

    if lower_name.endswith(".json"):
        valid_raw, invalid = _parse_juniper_srx_json(text)
        source_kind = "juniper-srx-json"
    else:
        valid_raw, invalid = _parse_juniper_srx_xml(text)
        source_kind = "juniper-srx-xml"

    normalized_rows = [r.model_dump() for r in valid_raw]

    if selected_fmt == "json":
        payload = [r.model_dump() for r in valid_raw]
        content = json.dumps(payload, indent=2)
        media_type = "application/json"
        ext = "json"
    else:
        content = _render_cleaned_raw_csv(normalized_rows)
        media_type = "text/csv; charset=utf-8"
        ext = "csv"

    _record_audit(
        request,
        caller,
        endpoint="/audit/json/cleaned",
        payload_summary={
            "kind": source_kind,
            "filename": filename,
            "format": selected_fmt,
            "valid_rows": len(normalized_rows),
            "invalid_rows": len(invalid),
        },
        verdict_summary={"format": selected_fmt},
    )

    report_name = Path(filename).stem.replace(" ", "-")[:60] or "juniper-rules"
    if selected_fmt == "json":
        return Response(
            content=content,
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{report_name}-cleaned.json"',
                "X-Audit-Valid-Rows": str(len(normalized_rows)),
                "X-Audit-Invalid-Rows": str(len(invalid)),
                "X-Audit-Cleaned-Format": "json",
            },
        )
    else:
        return Response(
            content=content,
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{report_name}-cleaned.csv"',
                "X-Audit-Valid-Rows": str(len(normalized_rows)),
                "X-Audit-Invalid-Rows": str(len(invalid)),
                "X-Audit-Cleaned-Format": "csv",
            },
        )


@app.get(
    "/audit/ui",
    summary="Upload UI for firewall compliance audit",
    response_class=HTMLResponse,
)
def audit_ui() -> HTMLResponse:
    template = Path(__file__).parent.parent / "templates" / "audit-upload.html"
    if not template.exists():
        raise HTTPException(status_code=404, detail="UI template not found.")
    return HTMLResponse(
        template.read_text(encoding="utf-8"),
        headers={
            # This UI has inline CSS/JS and same-origin API calls.
            # Override the API-wide default-src 'none' CSP for this page only.
            "Content-Security-Policy": (
                "default-src 'self'; "
                "script-src 'self' 'unsafe-inline'; "
                "style-src 'self' 'unsafe-inline'; "
                "img-src 'self' data:; "
                "connect-src 'self'; "
                "form-action 'self'; "
                "frame-ancestors 'none'"
            )
        },
    )


@app.get(
    "/guide",
    summary="User guide for firewall compliance",
    response_class=HTMLResponse,
)
def user_guide() -> HTMLResponse:
    template = Path(__file__).parent.parent / "templates" / "user-guide.html"
    if not template.exists():
        raise HTTPException(status_code=404, detail="User guide template not found.")
    return HTMLResponse(
        template.read_text(encoding="utf-8"),
        headers={
            "Content-Security-Policy": (
                "default-src 'self'; "
                "script-src 'self' 'unsafe-inline'; "
                "style-src 'self' 'unsafe-inline'; "
                "img-src 'self' data:; "
                "connect-src 'self'; "
                "form-action 'self'; "
                "frame-ancestors 'none'"
            )
        },
    )


def _md_escape(value: object) -> str:
    s = "" if value is None else str(value)
    return s.replace("|", "\\|").replace("\n", " ").strip()


def _format_raw_target(req: dict) -> str:
    return f"{req.get('source', '?')} → {req.get('destination', '?')} {req.get('protocol', '?')}/{req.get('port', '?')}"


def _format_intake_target(intake: dict) -> str:
    return (
        f"{intake.get('source_name', '?')} → {intake.get('destination_name', '?')} "
        f"{intake.get('protocol', '?')}/{intake.get('destination_port', '?')} "
        f"[app={intake.get('app_id', '?')}, env={intake.get('environment', '?')}]"
    )


def _add_rag_badges_to_html(html_content: str) -> str:
    """Post-process HTML to add RAG status badges based on risk levels."""
    import re

    def replace_h3(match):
        h3_content = match.group(1)
        risk_match = re.search(r'\((LOW|MEDIUM|HIGH|CRITICAL)\)', h3_content, re.IGNORECASE)
        if risk_match:
            risk_level = risk_match.group(1).upper()
            if risk_level == "LOW":
                badge = '<span class="rag-badge rag-low">LOW</span>'
            elif risk_level == "MEDIUM":
                badge = '<span class="rag-badge rag-medium">MEDIUM</span>'
            elif risk_level in {"HIGH", "CRITICAL"}:
                badge = f'<span class="rag-badge rag-{risk_level.lower()}">{risk_level}</span>'
            else:
                badge = ''
            if badge:
                return f'<h3>{h3_content} {badge}</h3>'
        return match.group(0)

    return re.sub(r'<h3>(.*?)</h3>', replace_h3, html_content)


def _markdown_to_html_document(markdown_text: str) -> str:
    """Convert markdown report text to a styled Clarisys-branded HTML document."""
    body_html: str
    try:
        import markdown  # type: ignore

        body_html = markdown.markdown(
            markdown_text,
            extensions=["tables", "fenced_code", "sane_lists"],
        )
    except Exception:
        body_html = f"<pre>{html.escape(markdown_text)}</pre>"

    return (
        "<!DOCTYPE html>"
        "<html lang=\"en\">"
        "<head>"
        "<meta charset=\"UTF-8\">"
        "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">"
        "<title>Firewall Ruleset Compliance Report</title>"
        "<style>"
        ":root{\n"
        "--ms-black:#04342C;\n"
        "--ms-charcoal:#085041;\n"
        "--ms-text:#2C2C2A;\n"
        "--ms-muted:#5F5E5A;\n"
        "--ms-line:#D3D1C7;\n"
        "--ms-bg:#E1F5EE;\n"
        "--ms-surface:#ffffff;\n"
        "--ms-good:#1D9E75;\n"
        "--ms-bad:#b02a2a;\n"
        "--ms-primary:#0F6E56;\n"
        "}\n"
        "*{margin:0;padding:0;box-sizing:border-box;}\n"
        "body{font-family:\"Helvetica Neue\",Helvetica,Arial,sans-serif;line-height:1.55;color:var(--ms-text);background:var(--ms-bg);min-height:100vh;padding:14px;}\n"
        ".container{max-width:1200px;margin:0 auto;background:var(--ms-surface);border:1px solid var(--ms-line);box-shadow:0 10px 28px rgba(17,17,17,0.08);overflow:hidden;}\n"
        ".utility-strip{display:flex;justify-content:space-between;gap:12px;align-items:center;background:var(--ms-primary);color:#E1F5EE;padding:9px 28px;font-size:0.72rem;letter-spacing:0.06em;text-transform:uppercase;}\n"
        ".utility-links{display:flex;gap:18px;flex-wrap:wrap;}\n"
        ".utility-links span{opacity:0.9;}\n"
        ".brand-row{display:flex;justify-content:space-between;align-items:center;padding:18px 28px;border-bottom:1px solid var(--ms-line);background:#fff;}\n"
        ".brand{font-family:Georgia,\"Times New Roman\",serif;font-size:1.2rem;letter-spacing:0.16em;color:var(--ms-black);text-transform:uppercase;white-space:nowrap;}\n"
        ".top-nav{display:flex;gap:18px;align-items:center;font-size:0.78rem;letter-spacing:0.06em;text-transform:uppercase;color:var(--ms-charcoal);}\n"
        ".top-nav a{color:inherit;text-decoration:none;border-bottom:1px solid transparent;transition:border-color 0.2s ease;}\n"
        ".top-nav a:hover{border-bottom-color:var(--ms-primary);}\n"
        "header{background:#fff;color:var(--ms-charcoal);padding:20px 28px 22px;border-bottom:1px solid var(--ms-line);}\n"
        "header h1{font-family:Georgia,\"Times New Roman\",serif;font-size:2.05rem;font-weight:500;letter-spacing:0.01em;margin-bottom:4px;}\n"
        "header p{color:var(--ms-muted);font-size:0.92rem;}\n"
        ".metadata{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;padding:16px 28px;background:#fafafa;border-bottom:1px solid var(--ms-line);}\n"
        ".meta-item{display:flex;justify-content:space-between;align-items:center;gap:10px;padding:12px;background:#fff;border:1px solid var(--ms-line);}\n"
        ".meta-label{font-weight:600;color:var(--ms-muted);font-size:0.72rem;text-transform:uppercase;letter-spacing:0.08em;}\n"
        ".meta-value{font-size:0.98rem;font-weight:700;color:var(--ms-charcoal);}\n"
        ".status-non-compliant,.status-denied{color:var(--ms-bad);}\n"
        ".status-compliant,.status-acceptable{color:var(--ms-good);}\n"
        ".content{padding:26px 28px 32px;}\n"
        "h2{font-family:Georgia,\"Times New Roman\",serif;font-size:1.5rem;color:var(--ms-charcoal);margin-top:30px;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid var(--ms-primary);}\n"
        "h2:first-child{margin-top:0;}\n"
        "h3{font-size:1rem;color:var(--ms-charcoal);margin-top:24px;margin-bottom:12px;padding:10px 12px;background:#f5f5f5;border-left:3px solid var(--ms-primary);display:flex;align-items:center;gap:10px;flex-wrap:wrap;}\n"
        ".rag-badge{display:inline-flex;align-items:center;padding:4px 10px;border-radius:4px;font-weight:700;font-size:0.7rem;letter-spacing:0.06em;text-transform:uppercase;white-space:nowrap;}\n"
        ".rag-badge.rag-low{background:#1D9E75;color:#fff;}\n"
        ".rag-badge.rag-medium{background:#e8a500;color:#fff;}\n"
        ".rag-badge.rag-high{background:#b02a2a;color:#fff;}\n"
        ".rag-badge.rag-critical{background:#8b0000;color:#fff;}\n"
        "table{width:100%;border-collapse:collapse;margin:14px 0;border:1px solid var(--ms-line);}\n"
        "thead{background:var(--ms-primary);color:#fff;}\n"
        "th{padding:10px 12px;text-align:left;font-size:0.76rem;text-transform:uppercase;letter-spacing:0.07em;font-weight:700;}\n"
        "td{padding:10px 12px;border-bottom:1px solid #ececec;vertical-align:top;font-size:0.94rem;}\n"
        "tbody tr:nth-child(even){background:#fcfcfc;}\n"
        ".stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:10px;margin:12px 0;}\n"
        ".stat-box{padding:14px;background:#fff;text-align:center;border:1px solid var(--ms-line);border-top:3px solid var(--ms-primary);}\n"
        ".stat-box.acceptable{border-top-color:var(--ms-good);}\n"
        ".stat-box.denied{border-top-color:var(--ms-bad);}\n"
        ".stat-number{font-size:1.9rem;font-weight:700;color:var(--ms-charcoal);margin:4px 0;}\n"
        ".stat-box.acceptable .stat-number{color:var(--ms-good);}\n"
        ".stat-box.denied .stat-number{color:var(--ms-bad);}\n"
        ".stat-label{font-size:0.72rem;color:var(--ms-muted);text-transform:uppercase;letter-spacing:0.08em;font-weight:700;}\n"
        "footer{background:#f5f5f5;padding:14px 28px;text-align:center;color:#4f4f4f;font-size:0.8rem;border-top:1px solid var(--ms-line);}\n"
        "@media (max-width:768px){.utility-strip,.brand-row,header,.metadata,.content,footer{padding-left:14px;padding-right:14px;}.utility-strip,.brand-row{flex-direction:column;align-items:flex-start;}.brand{font-size:1.05rem;}header h1{font-size:1.55rem;}.metadata{grid-template-columns:1fr;}}\n"
        "</style>"
        "</head>"
        "<body>"
        "<div class=\"container\">"
        "<div class=\"utility-strip\"><div>Clarisys</div></div>"
        "<div class=\"brand-row\"><div class=\"brand\">Clarisys</div><nav class=\"top-nav\" aria-label=\"Report Navigation\"><a href=\"#summary\">Summary</a></nav></div>"
        "<header><h1>Firewall Ruleset Compliance Report</h1><p>Compliance audit aligned to Clarisys security standards framework.</p></header>"
        "<div class=\"metadata\" id=\"summary\"></div>"
        "<div class=\"content\">"
        f"{_add_rag_badges_to_html(body_html)}"
        "</div>"
        "<footer>Firewall Compliance Report &bullet; Clarisys Security Standards Framework &bullet; Automated Audit</footer>"
        "</div>"
        "</body>"
        "</html>"
    )


def _render_markdown_report(
    kind: str,
    summary: dict,
    results: list[dict],
    invalid: list[AuditInvalidRow],
) -> str:
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    total = summary.get("total", 0)
    acceptable = summary.get("acceptable", 0)
    denied = summary.get("denied", 0)
    overall_status = summary.get("overall_status", "COMPLIANT" if denied == 0 and not invalid else "NON-COMPLIANT")

    lines: list[str] = []
    lines.append("# Firewall Ruleset Compliance Report")
    lines.append("")
    lines.append(f"- **Generated:** {generated}")
    lines.append(f"- **Schema detected:** `{kind}`")
    lines.append(f"- **Total rules evaluated:** {total}")
    lines.append(f"- **Acceptable:** {acceptable}")
    lines.append(f"- **Requires remediation:** {denied}")
    lines.append(f"- **Invalid rows:** {len(invalid)}")
    lines.append(f"- **Overall status:** **{overall_status}**")
    selected_standards: list[str] = []
    seen_standards: set[str] = set()
    for item in results:
        request_payload = item.get("request") if isinstance(item, dict) else None
        if not isinstance(request_payload, dict):
            request_payload = item.get("intake") if isinstance(item, dict) else None
        if not isinstance(request_payload, dict):
            continue
        for standard in request_payload.get("standards") or []:
            if isinstance(standard, str) and standard not in seen_standards:
                seen_standards.add(standard)
                selected_standards.append(standard)
    if selected_standards:
        lines.append(f"- **Selected optional standards:** {', '.join(_md_escape(s) for s in selected_standards)}")
    lines.append("- **Evaluated against selected standards only**")
    if kind == "intake":
        lines.append(f"- **Total risk score:** {summary.get('total_risk_score', 0)}")
        lines.append(f"- **Max risk score:** {summary.get('max_risk_score', 0)}")
    lines.append("")

    failed_controls = summary.get("failed_controls") or []
    if failed_controls:
        lines.append("## Failed controls")
        lines.append("")
        mappings_data = _load_compliance_mappings().get("controls", {})
        lines.append("| Control | Clause | Occurrences |")
        lines.append("|---|---|---|")
        breakdown = summary.get("by_failed_control") or {}
        for ctrl in failed_controls:
            ctrl_info = mappings_data.get(ctrl, {})
            title = ctrl_info.get("title", ctrl)
            # Show clause from selected standard if available
            clause_str = ""
            if selected_standards and ctrl_info.get("mappings"):
                for std in selected_standards:
                    clauses = ctrl_info["mappings"].get(std, [])
                    if clauses:
                        clause_str = ", ".join(str(c) for c in clauses)
                        break
            lines.append(f"| {_md_escape(title)} | {_md_escape(clause_str)} | {breakdown.get(ctrl, '')} |")
        lines.append("")

    failed_standards = summary.get("by_failed_standard") or {}
    if failed_standards:
        lines.append("## Failed standards")
        lines.append("")
        lines.append("| Standard | Occurrences |")
        lines.append("|---|---|")
        for std, count_ in failed_standards.items():
            lines.append(f"| {_md_escape(std)} | {count_} |")
        lines.append("")

    lines.append("## Per-rule findings")
    lines.append("")
    if not results:
        lines.append("_No valid rows were evaluated._")
        lines.append("")
    else:
        for offset, item in enumerate(results, start=2):
            verdict_raw = str(item.get("verdict", "?")).upper()
            verdict = "REQUIRES REMEDIATION" if verdict_raw in {"DENY", "DENIED"} else verdict_raw
            risk = item.get("overall_risk", "?")
            if kind in {"raw", "raw_paloalto"}:
                target = _format_raw_target(item.get("request", {}))
            else:
                target = _format_intake_target(item.get("intake", {}))
            heading = f"### Row {offset} — {verdict} ({risk})"
            if kind == "intake":
                heading += f" — risk score {item.get('risk_score', '?')}"
            lines.append(heading)
            lines.append("")
            lines.append(f"- **Target:** {target}")
            lines.append(f"- **Reason:** {_md_escape(item.get('reason', ''))}")
            ctrls = item.get("failed_controls") or []
            if ctrls:
                ctrl_labels = [mappings_data.get(c, {}).get("title", c) for c in ctrls]
                lines.append(f"- **Failed controls:** {', '.join(ctrl_labels)}")
            stds = item.get("failed_standards") or []
            if stds:
                lines.append(f"- **Failed standards:** {', '.join(stds)}")
            violations = item.get("violations") or []
            if violations:
                lines.append("- **Violations:**")
                for v in violations:
                    if not isinstance(v, dict):
                        continue
                    ctrl_id = v.get("control") or ""
                    ctrl_label = mappings_data.get(ctrl_id, {}).get("title", ctrl_id)
                    lines.append(
                        f"  - **{_md_escape(ctrl_label)}** ({_md_escape(v.get('severity'))}) — "
                        f"{_md_escape(v.get('violation'))}"
                    )
                    if v.get("remediation"):
                        lines.append(f"    - Remediation: {_md_escape(v.get('remediation'))}")
            lines.append("")

    if invalid:
        lines.append("## Invalid rows")
        lines.append("")
        lines.append("| Row | Errors | Raw |")
        lines.append("|---|---|---|")
        for item in invalid:
            errors = "; ".join(_md_escape(e) for e in item.errors)
            raw_preview = _md_escape(json.dumps(item.raw, ensure_ascii=False))
            lines.append(f"| {item.row} | {errors} | `{raw_preview}` |")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


# ── Frontend static files (React SPA) ───────────────────────────────────────

_FRONTEND_DIR = Path(__file__).parent.parent / "frontend" / "dist"


@app.get("/app/{path:path}", include_in_schema=False)
async def serve_frontend(path: str) -> Response:
    """Serve React SPA static assets from frontend/dist/."""
    if not _FRONTEND_DIR.is_dir():
        raise HTTPException(status_code=404, detail="Frontend not built. Run: cd frontend && npm run build")

    # Serve the requested file if it exists
    file_path = _FRONTEND_DIR / path
    # Prevent directory traversal
    try:
        file_path = file_path.resolve()
        _FRONTEND_DIR.resolve()
        if not str(file_path).startswith(str(_FRONTEND_DIR.resolve())):
            raise HTTPException(status_code=403, detail="Forbidden")
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="Invalid path")

    if file_path.is_file():
        content_type = _guess_mime(file_path.suffix)
        return Response(
            content=file_path.read_bytes(),
            media_type=content_type,
            headers={"Cache-Control": "public, max-age=31536000, immutable"}
            if file_path.suffix in {".js", ".css", ".woff2", ".woff", ".ttf"}
            else {},
        )

    # Fallback to index.html for SPA client-side routing
    index = _FRONTEND_DIR / "index.html"
    if index.is_file():
        return HTMLResponse(index.read_text(encoding="utf-8"))
    raise HTTPException(status_code=404, detail="Not found")


def _guess_mime(suffix: str) -> str:
    return {
        ".js": "application/javascript",
        ".css": "text/css",
        ".html": "text/html",
        ".json": "application/json",
        ".svg": "image/svg+xml",
        ".png": "image/png",
        ".ico": "image/x-icon",
        ".woff": "font/woff",
        ".woff2": "font/woff2",
        ".ttf": "font/ttf",
    }.get(suffix.lower(), "application/octet-stream")
