import openpyxl, json, re

wb = openpyxl.load_workbook('Firewall Policy.xlsx')

# ── helpers ──────────────────────────────────────────────────────────────────

def mask_to_prefix(mask):
    mask = mask.rstrip(')')  # Remove trailing paren if present
    return sum(bin(int(x)).count('1') for x in mask.split('.'))

def parse_cidr(ip, mask):
    if '.' in mask:          # dotted netmask
        return f"{ip}/{mask_to_prefix(mask)}"
    return f"{ip}/{mask}"   # already prefix len

def split_cell(val):
    if not val:
        return []
    return [v.strip() for v in str(val).split('\n') if v.strip()]

# ── Firewall rules ────────────────────────────────────────────────────────────

fw_ws = wb['Firewall Policy']
rows = list(fw_ws.iter_rows(values_only=True))

rules = []
for row in rows[5:]:                    # data starts at row 6 (index 5)
    if row[0] is None:
        continue
    seq = row[0]
    if not isinstance(seq, (int, float)):
        continue
    rule = {
        "seq": int(seq),
        "id": row[1],
        "name": row[2],
        "action": str(row[3]).lower() if row[3] else "deny",
        "source": {
            "interfaces": split_cell(row[4]),
            "addresses":  split_cell(row[5])
        },
        "destination": {
            "interfaces": split_cell(row[7]),
            "addresses":  split_cell(row[8])
        },
        "services":    split_cell(row[12]),
        "log":  str(row[15]).lower().replace(' ', '_') if row[15] else "no_log",
        "nat":  True if str(row[16]).strip() == 'Y' else False,
        "install_on": split_cell(row[20]),
        "comments": str(row[19]) if row[19] else ""
    }
    rules.append(rule)

# ── Address groups ────────────────────────────────────────────────────────────

addr_ws = wb['Address Group']
address_groups = {}

ag_rows = list(addr_ws.iter_rows(values_only=True))
for row in ag_rows[1:]:
    name, details = row[0], row[1] if len(row) > 1 else None
    if not name or not details:
        continue
    entry = {"cidrs": [], "ip_ranges": [], "fqdns": [], "members": []}
    items_str = re.sub(r'^Total\(\d+\):\s*', '', str(details))
    for item in re.split(r',\s*(?=[A-Za-z0-9_\*\.\-]+\()', items_str):
        item = item.strip().rstrip(')')
        m_ip = re.search(r'IP/Netmask:\s+(\S+)/(\S+?)(?:\)|$)', item)
        m_range = re.search(r'IP Range:\s+(.+)\)', item)
        m_fqdn = re.search(r'FQDN:([^\)]+)', item)
        m_group = re.search(r'Group Members', item)
        if m_ip:
            entry["cidrs"].append(parse_cidr(m_ip.group(1), m_ip.group(2)))
        elif m_range:
            entry["ip_ranges"].append(m_range.group(1).strip())
        elif m_fqdn:
            entry["fqdns"].append(m_fqdn.group(1).strip())
        elif m_group:
            gname = re.match(r'([^\(]+)\(', item)
            if gname:
                entry["members"].append(gname.group(1).strip())
    address_groups[name] = entry

# Inline addresses and service definitions
inline_addresses = {
    "all":                    {"cidrs": ["0.0.0.0/0"], "ip_ranges": [], "fqdns": [], "members": []},
    "any":                    {"cidrs": ["0.0.0.0/0"], "ip_ranges": [], "fqdns": [], "members": []},
    "10.0.0.0_8":             {"cidrs": ["10.0.0.0/8"], "ip_ranges": [], "fqdns": [], "members": []},
    "TRAS_10.157.26.0_24":    {"cidrs": ["10.157.26.0/24"], "ip_ranges": [], "fqdns": [], "members": []},
    "GPU_svr_2099":           {"cidrs": [], "ip_ranges": ["10.221.126.33-10.221.126.34"], "fqdns": [], "members": []},
}
address_groups.update(inline_addresses)

address_variables = {
    "STORE_MAIN_DATA_SUMMARY":       {"cidrs": [], "_note": "Populate per-store"},
    "STORE_MAIN_MGMT_SUMMARY":       {"cidrs": [], "_note": "Populate per-store"},
    "STORE_MAIN_VOICE_SUMMARY":      {"cidrs": [], "_note": "Populate per-store"},
    "store_main_data_summary":       {"cidrs": [], "_note": "Populate per-store"},
    "store_main_mgmt_summary":       {"cidrs": [], "_note": "Populate per-store"},
    "store_main_voice_summary":      {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan101_dns_server": {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan102_net":        {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan104_net":        {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan111_net":        {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan114_net":        {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan116_net":        {"cidrs": [], "_note": "Populate per-store"},
    "store_main_vlan1432_net":       {"cidrs": [], "_note": "Populate per-store"},
}

service_definitions = {
    "ALL":      {"match_all": True,  "entries": []},
    "ALL_ICMP": {"match_all": False, "entries": [{"protocol": "icmp"}]},
    "ALL_TCP":  {"match_all": False, "entries": [{"protocol": "tcp", "port_min": 1, "port_max": 65535}]},
    "DNS":      {"match_all": False, "entries": [{"protocol": "tcp", "port": 53}, {"protocol": "udp", "port": 53}]},
    "NTP":      {"match_all": False, "entries": [{"protocol": "tcp", "port": 123}, {"protocol": "udp", "port": 123}]},
    "HTTP":     {"match_all": False, "entries": [{"protocol": "tcp", "port": 80}]},
    "HTTPS":    {"match_all": False, "entries": [{"protocol": "tcp", "port": 443}]},
    "FTP":      {"match_all": False, "entries": [{"protocol": "tcp", "port": 21}]},
}

data = {
    "_metadata": {
        "adom": "Stores",
        "policy_package": "STORE-MAIN-root",
        "exported": "2026-05-15",
        "total_rules": len(rules)
    },
    "rules": rules,
    "address_groups": address_groups,
    "address_variables": address_variables,
    "service_definitions": service_definitions
}

import os
os.makedirs('policy', exist_ok=True)
with open('policy/data.json', 'w') as f:
    json.dump(data, f, indent=2)

print(f"Generated policy/data.json")
print(f"  Rules:           {len(rules)}")
print(f"  Address groups:  {len(address_groups)}")
print(f"  Addr variables:  {len(address_variables)}")
print(f"  Service defs:    {len(service_definitions)}")
